from __future__ import annotations

from datetime import date

import pandas as pd
from openpyxl import Workbook

from backend.valuation import xlsx_formatting


def _sheet_from_dataframe(df: pd.DataFrame):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(list(df.columns))
    for row in df.itertuples(index=False, name=None):
        sheet.append(list(row))
    return workbook, sheet


def test_xlsx_apply_header_style_formats_first_row() -> None:
    _workbook, sheet = _sheet_from_dataframe(pd.DataFrame([{"Player": "A", "DynastyValue": 1.0}]))

    xlsx_formatting._xlsx_apply_header_style(sheet)

    assert sheet.row_dimensions[1].height == 22
    assert sheet["A1"].font.bold is True
    assert sheet["A1"].fill.fgColor.rgb in {"001F4E79", "1F4E79"}
    assert sheet["A1"].alignment.horizontal == "center"


def test_xlsx_set_freeze_filters_and_view_adds_and_clears_autofilter() -> None:
    _workbook, sheet = _sheet_from_dataframe(pd.DataFrame([{"Player": "A"}, {"Player": "B"}]))

    xlsx_formatting._xlsx_set_freeze_filters_and_view(sheet, freeze_panes="B2", add_autofilter=True)
    assert sheet.freeze_panes == "B2"
    assert sheet.sheet_view.showGridLines is False
    assert sheet.auto_filter.ref == "A1:A3"

    xlsx_formatting._xlsx_set_freeze_filters_and_view(sheet, freeze_panes="A2", add_autofilter=False)
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is None


def test_xlsx_add_table_noops_for_insufficient_data_rows() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Player", "DynastyValue"])

    xlsx_formatting._xlsx_add_table(sheet, table_name="TooSmall")

    assert len(sheet.tables) == 0


def test_xlsx_add_table_creates_table_and_clears_sheet_autofilter() -> None:
    _workbook, sheet = _sheet_from_dataframe(pd.DataFrame([{"Player": "A"}, {"Player": "B"}]))
    sheet.auto_filter.ref = "A1:A3"

    xlsx_formatting._xlsx_add_table(sheet, table_name="PlayerValuesTbl")

    assert "PlayerValuesTbl" in sheet.tables
    assert sheet.auto_filter.ref is None


def test_xlsx_set_column_widths_honors_overrides_and_type_defaults() -> None:
    df = pd.DataFrame(
        [
            {
                "Player": "Very Long Player Name",
                "Age": 27,
                "Eligible": True,
                "OldestProjectionDate": date(2026, 2, 25),
                "Value_2026": 12.34,
            }
        ]
    )
    workbook = Workbook()
    sheet = workbook.active

    xlsx_formatting._xlsx_set_column_widths(
        sheet,
        df,
        overrides={"Player": 30.0},
        sample_rows=100,
        min_width=8.0,
        max_width=40.0,
    )

    assert sheet.column_dimensions["A"].width == 30.0
    assert sheet.column_dimensions["B"].width <= 16.0
    assert sheet.column_dimensions["C"].width <= 16.0
    assert sheet.column_dimensions["D"].width == 14.0
    assert sheet.column_dimensions["E"].width == 10.0


def test_xlsx_apply_number_formats_applies_to_data_rows_only() -> None:
    df = pd.DataFrame([{"Age": 22, "DynastyValue": 1.234}, {"Age": 23, "DynastyValue": 5.678}])
    _workbook, sheet = _sheet_from_dataframe(df)

    xlsx_formatting._xlsx_apply_number_formats(
        sheet,
        df,
        {"Age": "0", "DynastyValue": "0.00", "MissingCol": "0.0"},
    )

    assert sheet["A1"].number_format != "0"
    assert sheet["A2"].number_format == "0"
    assert sheet["A3"].number_format == "0"
    assert sheet["B2"].number_format == "0.00"
    assert sheet["B3"].number_format == "0.00"


def test_xlsx_add_value_color_scale_only_when_column_and_rows_are_valid() -> None:
    df = pd.DataFrame([{"DynastyValue": 1.0}, {"DynastyValue": 2.0}])
    _workbook, sheet = _sheet_from_dataframe(df)

    xlsx_formatting._xlsx_add_value_color_scale(sheet, df, "DynastyValue")
    assert len(sheet.conditional_formatting) == 1

    xlsx_formatting._xlsx_add_value_color_scale(sheet, df, "Missing")
    assert len(sheet.conditional_formatting) == 1

    tiny_df = pd.DataFrame([{"DynastyValue": 1.0}])
    _workbook2, tiny_sheet = _sheet_from_dataframe(tiny_df)
    xlsx_formatting._xlsx_add_value_color_scale(tiny_sheet, tiny_df, "DynastyValue")
    assert len(tiny_sheet.conditional_formatting) == 0


def test_xlsx_format_player_values_smoke() -> None:
    df = pd.DataFrame(
        [
            {
                "Player": "Jane Roe",
                "Team": "SEA",
                "MLBTeam": "SEA",
                "Pos": "OF",
                "Age": 26,
                "OldestProjectionDate": date(2026, 2, 24),
                "DynastyValue": 7.12,
                "RawDynastyValue": 8.55,
                "CenteringBaselineMean": 1.23,
                "Value_2026": 2.34,
            },
            {
                "Player": "Janet Roe",
                "Team": "SEA",
                "MLBTeam": "SEA",
                "Pos": "OF",
                "Age": 27,
                "OldestProjectionDate": date(2026, 2, 24),
                "DynastyValue": 6.01,
                "RawDynastyValue": 7.22,
                "CenteringBaselineMean": 1.23,
                "Value_2026": 2.01,
            }
        ]
    )
    _workbook, sheet = _sheet_from_dataframe(df)

    xlsx_formatting._xlsx_format_player_values(sheet, df, table_name="PlayerValuesTbl")

    assert sheet.freeze_panes == "B2"
    assert "PlayerValuesTbl" in sheet.tables
    assert sheet["H2"].number_format == "0.00"
    assert sheet["J2"].number_format == "0.00"
    assert len(sheet.conditional_formatting) == 1


def test_xlsx_format_detail_sheet_smoke() -> None:
    df = pd.DataFrame(
        [
            {
                "Player": "Jane Roe",
                "Year": 2026,
                "Team": "SEA",
                "MLBTeam": "SEA",
                "Pos": "OF",
                "Age": 26,
                "BestSlot": "OF",
                "OldestProjectionDate": date(2026, 2, 24),
                "YearValue": 3.21,
                "DynastyValue": 7.12,
                "RawDynastyValue": 8.55,
                "AVG": 0.301,
            },
            {
                "Player": "Jane Roe",
                "Year": 2027,
                "Team": "SEA",
                "MLBTeam": "SEA",
                "Pos": "OF",
                "Age": 27,
                "BestSlot": "OF",
                "OldestProjectionDate": date(2026, 2, 24),
                "YearValue": 2.88,
                "DynastyValue": 6.01,
                "RawDynastyValue": 7.22,
                "AVG": 0.297,
            },
        ]
    )
    _workbook, sheet = _sheet_from_dataframe(df)

    xlsx_formatting._xlsx_format_detail_sheet(
        sheet,
        df,
        table_name="BatAggTbl",
        is_pitch=False,
    )

    assert sheet.freeze_panes == "C2"
    assert "BatAggTbl" in sheet.tables
    assert sheet["B2"].number_format == "0"
    assert sheet["J2"].number_format == "0.00"
    assert len(sheet.conditional_formatting) == 2
