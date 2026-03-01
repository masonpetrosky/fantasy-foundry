import io
import ipaddress
import sys
import time
import types
import unittest
from collections import deque
from unittest.mock import Mock, patch

import pandas as pd
from fastapi import HTTPException
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import backend.app as app_module


class CalculatorValidationTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls) -> None:
        cls.client = TestClient(app_module.app)

    def setUp(self) -> None:
        app_module._calculate_common_dynasty_frame_cached.cache_clear()
        app_module._calculate_points_dynasty_frame_cached.cache_clear()
        app_module._playable_pool_counts_by_year.cache_clear()
        with app_module.CALCULATOR_JOB_LOCK:
            app_module.CALCULATOR_JOBS.clear()
        with app_module.CALC_RESULT_CACHE_LOCK:
            app_module.CALC_RESULT_CACHE.clear()
            app_module.CALC_RESULT_CACHE_ORDER.clear()
        with app_module.REQUEST_RATE_LIMIT_LOCK:
            app_module.REQUEST_RATE_LIMIT_BUCKETS.clear()
            app_module._REQUEST_RATE_LIMIT_LAST_SWEEP_TS = 0.0

    def test_local_result_cache_reorders_on_get_for_lru_eviction(self) -> None:
        with patch.object(app_module, "_redis_client", return_value=None), patch.object(
            app_module,
            "CALC_RESULT_CACHE_MAX_ENTRIES",
            2,
        ):
            app_module._result_cache_set("a", {"value": "A"})
            app_module._result_cache_set("b", {"value": "B"})
            cached = app_module._result_cache_get("a")
            app_module._result_cache_set("c", {"value": "C"})

        self.assertEqual(cached, {"value": "A"})
        self.assertIn("a", app_module.CALC_RESULT_CACHE)
        self.assertIn("c", app_module.CALC_RESULT_CACHE)
        self.assertNotIn("b", app_module.CALC_RESULT_CACHE)

    def test_local_result_cache_upsert_deduplicates_order_queue(self) -> None:
        with patch.object(app_module, "_redis_client", return_value=None):
            app_module._result_cache_set("same-key", {"value": 1})
            app_module._result_cache_set("same-key", {"value": 2})

        self.assertEqual(list(app_module.CALC_RESULT_CACHE_ORDER), ["same-key"])
        self.assertEqual(app_module._result_cache_get("same-key"), {"value": 2})

    def test_health_endpoint_reports_runtime_summary(self) -> None:
        with app_module.CALCULATOR_JOB_LOCK:
            app_module.CALCULATOR_JOBS.clear()
            app_module.CALCULATOR_JOBS["job-queued"] = {"status": "queued", "job_id": "job-queued"}
            app_module.CALCULATOR_JOBS["job-failed"] = {"status": "failed", "job_id": "job-failed"}
        with app_module.CALC_RESULT_CACHE_LOCK:
            app_module.CALC_RESULT_CACHE.clear()
            app_module.CALC_RESULT_CACHE_ORDER.clear()
            app_module.CALC_RESULT_CACHE["cache-key"] = (time.time() + 60.0, {"value": 1})
            app_module.CALC_RESULT_CACHE_ORDER.append("cache-key")

        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(
            app_module,
            "BAT_DATA",
            [{"Player": "Hitter"}],
        ), patch.object(
            app_module,
            "PIT_DATA",
            [{"Player": "Pitcher 1"}, {"Player": "Pitcher 2"}],
        ):
            response = self.client.get("/api/health")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload.get("status"), "ok")
        self.assertEqual(payload.get("projection_rows", {}).get("bat"), 1)
        self.assertEqual(payload.get("projection_rows", {}).get("pitch"), 2)
        self.assertEqual(payload.get("jobs", {}).get("total"), 2)
        self.assertEqual(payload.get("jobs", {}).get("queued"), 1)
        self.assertEqual(payload.get("jobs", {}).get("failed"), 1)
        self.assertIn("queue_pressure", payload)
        self.assertEqual(payload.get("queue_pressure", {}).get("active_jobs"), 1)
        self.assertEqual(
            payload.get("queue_pressure", {}).get("max_active_jobs_total"),
            app_module.CALCULATOR_MAX_ACTIVE_JOBS_TOTAL,
        )
        self.assertFalse(payload.get("queue_pressure", {}).get("at_capacity"))
        self.assertIn("dynasty_lookup_cache", payload)
        self.assertIn("status", payload.get("dynasty_lookup_cache", {}))
        self.assertIn("version_expected", payload.get("dynasty_lookup_cache", {}))
        self.assertIn("version_found", payload.get("dynasty_lookup_cache", {}))
        self.assertEqual(payload.get("result_cache", {}).get("local_entries"), 1)
        self.assertIn("timestamp", payload)

    def test_calculate_request_default_horizon_is_twenty(self) -> None:
        req = app_module.CalculateRequest()
        self.assertEqual(req.horizon, 20)
        self.assertEqual(req.minors, app_module.COMMON_DEFAULT_MINOR_SLOTS)
        self.assertEqual(req.sgp_denominator_mode, "classic")
        self.assertFalse(req.enable_playing_time_reliability)
        self.assertFalse(req.enable_age_risk_adjustment)
        self.assertFalse(req.enable_replacement_blend)

    def test_mode_must_be_common_or_league(self) -> None:
        response = self.client.post("/api/calculate", json={"mode": "invalid"})
        self.assertEqual(response.status_code, 422)

    def test_league_mode_with_points_scoring_uses_points_calculator(self) -> None:
        # League mode value is accepted but ignored — scoring_mode drives routing.
        response = self.client.post("/api/calculate", json={"mode": "league", "scoring_mode": "points"})
        self.assertEqual(response.status_code, 200)

    def test_rejects_invalid_ip_bounds(self) -> None:
        response = self.client.post("/api/calculate", json={"ip_min": 1200, "ip_max": 1000})
        self.assertEqual(response.status_code, 422)

    def test_rejects_invalid_sgp_winsor_bounds(self) -> None:
        response = self.client.post(
            "/api/calculate",
            json={"sgp_winsor_low_pct": 0.95, "sgp_winsor_high_pct": 0.90},
        )
        self.assertEqual(response.status_code, 422)

    def test_rejects_zero_total_hitter_slots(self) -> None:
        response = self.client.post(
            "/api/calculate",
            json={
                "hit_c": 0,
                "hit_1b": 0,
                "hit_2b": 0,
                "hit_3b": 0,
                "hit_ss": 0,
                "hit_ci": 0,
                "hit_mi": 0,
                "hit_of": 0,
                "hit_ut": 0,
            },
        )
        self.assertEqual(response.status_code, 422)

    def test_rejects_zero_total_pitcher_slots(self) -> None:
        response = self.client.post(
            "/api/calculate",
            json={
                "pit_p": 0,
                "pit_sp": 0,
                "pit_rp": 0,
            },
        )
        self.assertEqual(response.status_code, 422)

    def test_points_mode_requires_non_zero_scoring_rule(self) -> None:
        response = self.client.post(
            "/api/calculate",
            json={
                "scoring_mode": "points",
                "pts_hit_1b": 0,
                "pts_hit_2b": 0,
                "pts_hit_3b": 0,
                "pts_hit_hr": 0,
                "pts_hit_r": 0,
                "pts_hit_rbi": 0,
                "pts_hit_sb": 0,
                "pts_hit_bb": 0,
                "pts_hit_so": 0,
                "pts_pit_ip": 0,
                "pts_pit_w": 0,
                "pts_pit_l": 0,
                "pts_pit_k": 0,
                "pts_pit_sv": 0,
                "pts_pit_svh": 0,
                "pts_pit_h": 0,
                "pts_pit_er": 0,
                "pts_pit_bb": 0,
            },
        )
        self.assertEqual(response.status_code, 422)

    def test_calculate_response_includes_identity_fields(self) -> None:
        fake_out = pd.DataFrame(
            [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 5.0,
                    "RawDynastyValue": 6.0,
                    "minor_eligible": False,
                    "Value_2026": 5.0,
                }
            ]
        )

        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                self.kwargs = kwargs

        def fake_calculate(*args, **kwargs):
            return fake_out

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Jane Roe": ("jane-roe", "jane-roe")},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post("/api/calculate", json={})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["total"], 1)
        row = payload["data"][0]
        self.assertEqual(row["PlayerKey"], "jane-roe")
        self.assertEqual(row["PlayerEntityKey"], "jane-roe")
        self.assertEqual(row["DynastyMatchStatus"], "matched")

    def test_rejects_unknown_start_year(self) -> None:
        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(app_module, "META", {"years": [2026, 2027, 2028]}):
            response = self.client.post("/api/calculate", json={"start_year": 2031})

        self.assertEqual(response.status_code, 422)
        self.assertIn("start_year", response.json()["detail"])

    def test_returns_422_for_unfillable_roster_settings(self) -> None:
        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                self.kwargs = kwargs

        def fake_calculate(*args, **kwargs):
            raise ValueError("Not enough players (10) to fill required slots (15).")

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post(
                "/api/calculate",
                json={"teams": 20, "bench": 40, "minors": 60, "start_year": 2026},
            )

        self.assertEqual(response.status_code, 422)
        self.assertIn("Not enough players", response.json()["detail"])

    def test_returns_422_for_slot_eligibility_shortage(self) -> None:
        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                self.kwargs = kwargs

        def fake_calculate(*args, **kwargs):
            raise ValueError("Year 2045: Cannot fill slot 'RP': need 36 eligible pitchers (IP > 0) but only found 21.")

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post(
                "/api/calculate",
                json={"teams": 12, "pit_rp": 3, "start_year": 2026},
            )

        self.assertEqual(response.status_code, 422)
        self.assertIn("Cannot fill slot 'RP'", response.json()["detail"])

    def test_calculate_passes_slot_overrides_and_ir_to_settings(self) -> None:
        fake_out = pd.DataFrame(
            [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 5.0,
                    "RawDynastyValue": 6.0,
                    "minor_eligible": False,
                    "Value_2026": 5.0,
                }
            ]
        )
        captured_kwargs: dict = {}

        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                captured_kwargs.update(kwargs)

        def fake_calculate(*args, **kwargs):
            return fake_out

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Jane Roe": ("jane-roe", "jane-roe")},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post(
                "/api/calculate",
                json={
                    "hit_c": 2,
                    "hit_of": 3,
                    "pit_p": 7,
                    "pit_sp": 1,
                    "pit_rp": 1,
                    "ir": 4,
                    "sgp_denominator_mode": "robust",
                    "enable_playing_time_reliability": True,
                    "enable_age_risk_adjustment": True,
                    "enable_replacement_blend": True,
                    "replacement_blend_alpha": 0.55,
                },
            )

        self.assertEqual(response.status_code, 200)
        self.assertEqual(captured_kwargs.get("hitter_slots", {}).get("C"), 2)
        self.assertEqual(captured_kwargs.get("hitter_slots", {}).get("OF"), 3)
        self.assertEqual(captured_kwargs.get("pitcher_slots", {}).get("P"), 7)
        self.assertEqual(captured_kwargs.get("pitcher_slots", {}).get("SP"), 1)
        self.assertEqual(captured_kwargs.get("pitcher_slots", {}).get("RP"), 1)
        self.assertEqual(captured_kwargs.get("ir_slots"), 4)
        self.assertEqual(captured_kwargs.get("sgp_denominator_mode"), "robust")
        self.assertTrue(captured_kwargs.get("enable_playing_time_reliability"))
        self.assertTrue(captured_kwargs.get("enable_age_risk_adjustment"))
        self.assertTrue(captured_kwargs.get("enable_replacement_blend"))
        self.assertEqual(captured_kwargs.get("replacement_blend_alpha"), 0.55)

    def test_points_mode_respects_custom_scoring_weights(self) -> None:
        bat_rows = [
            {
                "Player": "Dual Threat",
                "Team": "SEA",
                "Year": 2026,
                "Pos": "OF",
                "Age": 28,
                "AB": 100,
                "H": 30,
                "2B": 5,
                "3B": 0,
                "HR": 10,
                "R": 0,
                "RBI": 0,
                "SB": 0,
                "BB": 0,
                "SO": 0,
                "PlayerKey": "dual-threat",
                "PlayerEntityKey": "dual-threat",
            }
        ]
        pit_rows = [
            {
                "Player": "Dual Threat",
                "Team": "SEA",
                "Year": 2026,
                "Pos": "SP",
                "Age": 28,
                "IP": 100,
                "W": 10,
                "L": 5,
                "K": 120,
                "SV": 0,
                "SVH": 0,
                "H": 80,
                "ER": 30,
                "BB": 25,
                "PlayerKey": "dual-threat",
                "PlayerEntityKey": "dual-threat",
            }
        ]

        base_payload = {
            "scoring_mode": "points",
            "start_year": 2026,
            "horizon": 1,
            "teams": 2,
            "hit_c": 0,
            "hit_1b": 0,
            "hit_2b": 0,
            "hit_3b": 0,
            "hit_ss": 0,
            "hit_ci": 0,
            "hit_mi": 0,
            "hit_of": 1,
            "hit_ut": 0,
            "pit_p": 1,
            "pit_sp": 0,
            "pit_rp": 0,
            "bench": 0,
            "minors": 0,
            "ir": 0,
            "pts_hit_1b": 0,
            "pts_hit_2b": 0,
            "pts_hit_3b": 0,
            "pts_hit_r": 0,
            "pts_hit_rbi": 0,
            "pts_hit_sb": 0,
            "pts_hit_bb": 0,
            "pts_hit_so": 0,
            "pts_pit_ip": 0,
            "pts_pit_w": 0,
            "pts_pit_l": 0,
            "pts_pit_k": 0,
            "pts_pit_sv": 0,
            "pts_pit_svh": 0,
            "pts_pit_h": 0,
            "pts_pit_er": 0,
            "pts_pit_bb": 0,
        }

        with patch.object(app_module, "_refresh_data_if_needed", return_value=None), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "BAT_DATA",
            bat_rows,
        ), patch.object(
            app_module,
            "PIT_DATA",
            pit_rows,
        ), patch.object(
            app_module,
            "BAT_DATA_RAW",
            bat_rows,
        ), patch.object(
            app_module,
            "PIT_DATA_RAW",
            pit_rows,
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Dual Threat": ("dual-threat", "dual-threat")},
        ):
            low_hr = self.client.post("/api/calculate", json={**base_payload, "pts_hit_hr": 2})
            high_hr = self.client.post("/api/calculate", json={**base_payload, "pts_hit_hr": 4})

        self.assertEqual(low_hr.status_code, 200)
        self.assertEqual(high_hr.status_code, 200)
        low_row = low_hr.json()["data"][0]
        high_row = high_hr.json()["data"][0]
        low_raw = low_row["RawDynastyValue"]
        high_raw = high_row["RawDynastyValue"]
        self.assertEqual(low_raw, 20.0)
        self.assertEqual(high_raw, 40.0)
        self.assertGreater(high_raw, low_raw)
        self.assertEqual(high_row["HittingPoints"], 40.0)
        self.assertEqual(high_row["PitchingPoints"], 0.0)
        self.assertEqual(high_row["SelectedPoints"], 40.0)
        self.assertEqual(high_row["HittingBestSlot"], "OF")
        self.assertEqual(high_row["PitchingBestSlot"], "P")
        self.assertEqual(high_row["HittingValue"], 40.0)
        self.assertEqual(high_row["PitchingValue"], 0.0)
        self.assertEqual(high_row["HittingAssignmentSlot"], "OF")
        self.assertIsNone(high_row["PitchingAssignmentSlot"])
        self.assertEqual(high_row["HittingAssignmentValue"], 40.0)
        self.assertEqual(high_row["PitchingAssignmentValue"], 0.0)
        self.assertEqual(high_row["KeepDropValue"], 40.0)

    def test_points_mode_applies_position_eligibility_for_year_values(self) -> None:
        bat_rows = [
            {
                "Player": "Dual Threat",
                "Team": "SEA",
                "Year": 2026,
                "Pos": "OF",
                "Age": 28,
                "AB": 100,
                "H": 30,
                "2B": 5,
                "3B": 0,
                "HR": 10,
                "R": 0,
                "RBI": 0,
                "SB": 0,
                "BB": 0,
                "SO": 0,
                "PlayerKey": "dual-threat",
                "PlayerEntityKey": "dual-threat",
            }
        ]
        pit_rows = [
            {
                "Player": "Dual Threat",
                "Team": "SEA",
                "Year": 2026,
                "Pos": "SP",
                "Age": 28,
                "IP": 100,
                "W": 10,
                "L": 5,
                "K": 120,
                "SV": 0,
                "SVH": 0,
                "H": 80,
                "ER": 30,
                "BB": 25,
                "PlayerKey": "dual-threat",
                "PlayerEntityKey": "dual-threat",
            }
        ]

        payload = {
            "scoring_mode": "points",
            "start_year": 2026,
            "horizon": 1,
            "teams": 2,
            "hit_c": 1,
            "hit_1b": 0,
            "hit_2b": 0,
            "hit_3b": 0,
            "hit_ss": 0,
            "hit_ci": 0,
            "hit_mi": 0,
            "hit_of": 0,
            "hit_ut": 0,
            "pit_p": 1,
            "pit_sp": 0,
            "pit_rp": 0,
            "bench": 0,
            "minors": 0,
            "ir": 0,
            "pts_hit_1b": 0,
            "pts_hit_2b": 0,
            "pts_hit_3b": 0,
            "pts_hit_hr": 4,
            "pts_hit_r": 0,
            "pts_hit_rbi": 0,
            "pts_hit_sb": 0,
            "pts_hit_bb": 0,
            "pts_hit_so": 0,
            "pts_pit_ip": 0,
            "pts_pit_w": 0,
            "pts_pit_l": 0,
            "pts_pit_k": 0,
            "pts_pit_sv": 0,
            "pts_pit_svh": 0,
            "pts_pit_h": 0,
            "pts_pit_er": 0,
            "pts_pit_bb": 0,
        }

        with patch.object(app_module, "_refresh_data_if_needed", return_value=None), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "BAT_DATA",
            bat_rows,
        ), patch.object(
            app_module,
            "PIT_DATA",
            pit_rows,
        ), patch.object(
            app_module,
            "BAT_DATA_RAW",
            bat_rows,
        ), patch.object(
            app_module,
            "PIT_DATA_RAW",
            pit_rows,
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Dual Threat": ("dual-threat", "dual-threat")},
        ):
            response = self.client.post("/api/calculate", json=payload)

        self.assertEqual(response.status_code, 200)
        row = response.json()["data"][0]
        self.assertEqual(row["Value_2026"], 0.0)
        self.assertEqual(row["RawDynastyValue"], 0.0)
        self.assertEqual(row["HittingPoints"], 40.0)
        self.assertEqual(row["SelectedPoints"], 0.0)
        self.assertIsNone(row["HittingBestSlot"])

    def test_meta_includes_calculator_guardrails_payload(self) -> None:
        with patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ):
            response = self.client.get("/api/meta")

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        guardrails = payload.get("calculator_guardrails", {})
        self.assertEqual(guardrails.get("hitters_per_team"), 13)
        self.assertEqual(guardrails.get("pitchers_per_team"), 9)
        self.assertEqual(guardrails.get("default_minors_slots"), app_module.COMMON_DEFAULT_MINOR_SLOTS)
        self.assertIn("default_points_scoring", guardrails)
        self.assertIn("playable_by_year", guardrails)
        self.assertIn("rate_limit_sync_authenticated_per_minute", guardrails)
        self.assertIn("rate_limit_job_create_authenticated_per_minute", guardrails)
        self.assertIn("rate_limit_job_status_authenticated_per_minute", guardrails)
        self.assertIn("max_active_jobs_total", guardrails)
        default_pitcher_categories = guardrails.get("default_roto_pitcher_categories", [])
        self.assertIn("QS", default_pitcher_categories)
        self.assertIn("QA3", default_pitcher_categories)
        freshness = payload.get("projection_freshness", {})
        self.assertIn("newest_projection_date", freshness)
        self.assertIn("date_coverage_pct", freshness)
        self.assertIn("last_projection_update", payload)
        self.assertIn("projection_window_start", payload)
        self.assertIn("projection_window_end", payload)

    def test_calculate_response_includes_explanations_payload(self) -> None:
        fake_out = pd.DataFrame(
            [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 5.0,
                    "RawDynastyValue": 6.0,
                    "minor_eligible": False,
                    "Value_2026": 5.0,
                }
            ]
        )

        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                self.kwargs = kwargs

        def fake_calculate(*args, **kwargs):
            return fake_out

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(app_module, "_refresh_data_if_needed", return_value=None), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Jane Roe": ("jane-roe", "jane-roe")},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post("/api/calculate", json={})

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertIn("explanations", payload)
        explanation = payload["explanations"]["jane-roe"]
        self.assertEqual(explanation["player"], "Jane Roe")
        self.assertEqual(explanation["mode"], "roto")
        self.assertEqual(explanation["per_year"][0]["year"], 2026)

    def test_projection_export_csv_endpoint(self) -> None:
        sample_rows = [{"Player": "Jane Roe", "Team": "SEA", "Year": 2026, "Pos": "OF"}]
        with patch.object(app_module, "BAT_DATA", sample_rows), patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ):
            response = self.client.get("/api/projections/export/bat?format=csv&include_dynasty=false")
        self.assertEqual(response.status_code, 200)
        self.assertIn("text/csv", response.headers.get("content-type", ""))
        self.assertIn("Player", response.text)

    def test_projection_export_csv_respects_columns_and_drops_internal_fields(self) -> None:
        sample_rows = [
            {
                "Player": "Jane Roe",
                "Team": "SEA",
                "Year": 2026,
                "Pos": "OF",
                "DynastyValue": 5.126,
                "AVG": 0.2894,
                "PlayerKey": "jane-roe",
                "PlayerEntityKey": "jane-roe",
                "DynastyMatchStatus": "matched",
            }
        ]
        with patch.object(app_module, "BAT_DATA", sample_rows), patch.object(
            app_module,
            "_refresh_data_if_needed",
            return_value=None,
        ):
            app_module.PROJECTION_SERVICE.clear_caches()
            response = self.client.get(
                "/api/projections/export/bat",
                params={
                    "format": "csv",
                    "include_dynasty": "false",
                    "columns": "Player,DynastyValue,AVG,PlayerKey,DynastyMatchStatus",
                },
            )

        self.assertEqual(response.status_code, 200)
        lines = response.text.splitlines()
        self.assertGreaterEqual(len(lines), 2)
        self.assertEqual(lines[0], "Player,Dynasty Value,AVG")
        self.assertNotIn("PlayerKey", lines[0])
        self.assertNotIn("DynastyMatchStatus", lines[0])
        self.assertIn("5.13", lines[1])
        self.assertIn("0.289", lines[1])

    def test_calculate_export_csv_respects_export_columns_and_drops_internal_fields(self) -> None:
        fake_result = {
            "total": 1,
            "settings": {},
            "data": [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 7.126,
                    "RawDynastyValue": 8.555,
                    "PlayerKey": "jane-roe",
                    "PlayerEntityKey": "jane-roe",
                    "DynastyMatchStatus": "matched",
                    "minor_eligible": False,
                    "Value_2026": 2.345,
                }
            ],
            "explanations": {},
        }
        with patch.object(app_module, "_run_calculate_request", return_value=fake_result):
            response = self.client.post(
                "/api/calculate/export",
                json={
                    "format": "csv",
                    "export_columns": [
                        "Player",
                        "DynastyValue",
                        "Team",
                        "RawDynastyValue",
                        "PlayerKey",
                        "Value_2026",
                    ],
                },
            )

        self.assertEqual(response.status_code, 200)
        lines = response.text.splitlines()
        self.assertGreaterEqual(len(lines), 2)
        self.assertEqual(lines[0], "Player,Dynasty Value,Team,2026 Dyn Value")
        self.assertNotIn("RawDynastyValue", lines[0])
        self.assertNotIn("PlayerKey", lines[0])
        self.assertIn("7.13", lines[1])
        self.assertIn("2.35", lines[1])

    def test_calculate_export_csv_points_defaults_include_points_summary_columns(self) -> None:
        fake_result = {
            "total": 1,
            "settings": {"scoring_mode": "points"},
            "data": [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 7.126,
                    "RawDynastyValue": 8.555,
                    "HittingPoints": 22.4,
                    "PitchingPoints": 0.0,
                    "SelectedPoints": 19.2,
                    "HittingBestSlot": "OF",
                    "PitchingBestSlot": "P",
                    "HittingValue": 19.2,
                    "PitchingValue": 0.0,
                    "HittingAssignmentSlot": "OF",
                    "PitchingAssignmentSlot": None,
                    "HittingAssignmentValue": 19.2,
                    "PitchingAssignmentValue": 0.0,
                    "KeepDropValue": 17.9,
                    "Value_2026": 2.345,
                }
            ],
            "explanations": {},
        }
        with patch.object(app_module, "_run_calculate_request", return_value=fake_result):
            response = self.client.post("/api/calculate/export", json={"format": "csv", "scoring_mode": "points"})

        self.assertEqual(response.status_code, 200)
        lines = response.text.splitlines()
        self.assertGreaterEqual(len(lines), 2)
        self.assertEqual(
            lines[0],
            (
                "Player,Dynasty Value,Age,Team,Pos,Hitting Points,Pitching Points,"
                "Selected Points,Hitting Best Slot,Pitching Best Slot,Hitting Value,"
                "Pitching Value,Hitting Assignment Slot,Pitching Assignment Slot,"
                "Hitting Assignment Value,Pitching Assignment Value,Keep/Drop Value,"
                "2026 Dyn Value"
            ),
        )
        self.assertIn("22.4", lines[1])
        self.assertIn("19.2", lines[1])

    def test_calculate_export_xlsx_endpoint(self) -> None:
        fake_out = pd.DataFrame(
            [
                {
                    "Player": "Jane Roe",
                    "Team": "SEA",
                    "Pos": "OF",
                    "Age": 26,
                    "DynastyValue": 5.0,
                    "RawDynastyValue": 6.0,
                    "minor_eligible": False,
                    "Value_2026": 5.0,
                }
            ]
        )

        class FakeSettings:
            def __init__(self, **kwargs) -> None:
                self.kwargs = kwargs

        def fake_calculate(*args, **kwargs):
            return fake_out

        fake_module = types.SimpleNamespace(
            CommonDynastyRotoSettings=FakeSettings,
            calculate_common_dynasty_values=fake_calculate,
        )

        with patch.object(app_module, "_refresh_data_if_needed", return_value=None), patch.object(
            app_module,
            "META",
            {"years": [2026]},
        ), patch.object(
            app_module,
            "_player_identity_by_name",
            return_value={"Jane Roe": ("jane-roe", "jane-roe")},
        ), patch.dict(
            sys.modules,
            {"dynasty_roto_values": fake_module},
        ):
            response = self.client.post("/api/calculate/export", json={"format": "xlsx", "include_explanations": True})

        self.assertEqual(response.status_code, 200)
        self.assertIn(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            response.headers.get("content-type", ""),
        )
        self.assertTrue(response.content.startswith(b"PK"))
        workbook = load_workbook(io.BytesIO(response.content))
        self.assertIn("Data", workbook.sheetnames)
        self.assertIn("Explainability", workbook.sheetnames)
        data_headers = [cell.value for cell in next(workbook["Data"].iter_rows(min_row=1, max_row=1))]
        self.assertIn("Dynasty Value", data_headers)
        self.assertNotIn("PlayerKey", data_headers)
        self.assertNotIn("PlayerEntityKey", data_headers)
        self.assertNotIn("DynastyMatchStatus", data_headers)
        self.assertNotIn("RawDynastyValue", data_headers)
        self.assertNotIn("minor_eligible", data_headers)

    def test_rate_limit_enforced_for_sync_calculate(self) -> None:
        with patch.object(app_module, "CALCULATOR_SYNC_RATE_LIMIT_PER_MINUTE", 1), patch.object(
            app_module,
            "_run_calculate_request",
            return_value={"total": 0, "settings": {}, "data": [], "explanations": {}},
        ):
            first = self.client.post("/api/calculate", json={})
            second = self.client.post("/api/calculate", json={})
        self.assertEqual(first.status_code, 200)
        self.assertEqual(second.status_code, 429)

    def test_calculate_auth_blocks_missing_key_when_required(self) -> None:
        with patch.object(app_module, "REQUIRE_CALCULATE_AUTH", True), patch.object(
            app_module,
            "CALCULATE_API_KEY_IDENTITIES",
            {"secret-key": "api_key:test-secret"},
        ), patch.object(
            app_module,
            "_run_calculate_request",
            return_value={"total": 0, "settings": {}, "data": [], "explanations": {}},
        ):
            missing = self.client.post("/api/calculate", json={})
            authorized = self.client.post("/api/calculate", json={}, headers={"x-api-key": "secret-key"})

        self.assertEqual(missing.status_code, 401)
        self.assertEqual(authorized.status_code, 200)

    def test_calculate_auth_returns_503_when_enabled_without_keys(self) -> None:
        with patch.object(app_module, "REQUIRE_CALCULATE_AUTH", True), patch.object(
            app_module,
            "CALCULATE_API_KEY_IDENTITIES",
            {},
        ):
            response = self.client.post("/api/calculate", json={})
        self.assertEqual(response.status_code, 503)

    def test_sync_rate_limit_applies_per_api_key_identity(self) -> None:
        with patch.object(app_module, "REQUIRE_CALCULATE_AUTH", True), patch.object(
            app_module,
            "CALCULATE_API_KEY_IDENTITIES",
            {"key-a": "api_key:a", "key-b": "api_key:b"},
        ), patch.object(app_module, "CALCULATOR_SYNC_RATE_LIMIT_PER_MINUTE", 1), patch.object(
            app_module,
            "CALCULATOR_SYNC_AUTH_RATE_LIMIT_PER_MINUTE",
            1,
        ), patch.object(
            app_module,
            "_run_calculate_request",
            return_value={"total": 0, "settings": {}, "data": [], "explanations": {}},
        ):
            first_key_a = self.client.post("/api/calculate", json={}, headers={"x-api-key": "key-a"})
            second_key_a = self.client.post("/api/calculate", json={}, headers={"x-api-key": "key-a"})
            first_key_b = self.client.post("/api/calculate", json={}, headers={"x-api-key": "key-b"})

        self.assertEqual(first_key_a.status_code, 200)
        self.assertEqual(second_key_a.status_code, 429)
        self.assertEqual(first_key_b.status_code, 200)

    def test_client_ip_ignores_untrusted_forwarded_chain(self) -> None:
        request = types.SimpleNamespace(
            headers={"x-forwarded-for": "203.0.113.9, 10.9.1.2"},
            client=types.SimpleNamespace(host="198.51.100.21"),
        )
        with patch.object(app_module, "TRUST_X_FORWARDED_FOR", False), patch.object(
            app_module,
            "TRUSTED_PROXY_NETWORKS",
            (ipaddress.ip_network("10.0.0.0/8"),),
        ):
            resolved = app_module._client_ip(request)
        self.assertEqual(resolved, "198.51.100.21")

    def test_client_ip_uses_first_untrusted_forwarded_hop_for_trusted_proxy(self) -> None:
        request = types.SimpleNamespace(
            headers={"x-forwarded-for": "198.51.100.55, 172.16.10.9, 10.1.2.3"},
            client=types.SimpleNamespace(host="10.200.0.10"),
        )
        with patch.object(app_module, "TRUST_X_FORWARDED_FOR", False), patch.object(
            app_module,
            "TRUSTED_PROXY_NETWORKS",
            (ipaddress.ip_network("10.0.0.0/8"), ipaddress.ip_network("172.16.0.0/12")),
        ):
            resolved = app_module._client_ip(request)
        self.assertEqual(resolved, "198.51.100.55")

    def test_client_ip_prefers_cf_connecting_ip_when_trusted(self) -> None:
        request = types.SimpleNamespace(
            headers={
                "cf-connecting-ip": "203.0.113.50",
                "x-forwarded-for": "198.51.100.99",
            },
            client=types.SimpleNamespace(host="10.0.0.1"),
        )
        with patch.object(app_module, "TRUST_X_FORWARDED_FOR", True), patch.object(
            app_module,
            "TRUSTED_PROXY_NETWORKS",
            (),
        ):
            resolved = app_module._client_ip(request)
        self.assertEqual(resolved, "203.0.113.50")

    def test_client_ip_ignores_cf_connecting_ip_when_untrusted(self) -> None:
        request = types.SimpleNamespace(
            headers={
                "cf-connecting-ip": "203.0.113.50",
                "x-forwarded-for": "198.51.100.99",
            },
            client=types.SimpleNamespace(host="192.0.2.1"),
        )
        with patch.object(app_module, "TRUST_X_FORWARDED_FOR", False), patch.object(
            app_module,
            "TRUSTED_PROXY_NETWORKS",
            (),
        ):
            resolved = app_module._client_ip(request)
        self.assertEqual(resolved, "192.0.2.1")

    def test_rate_limit_bucket_cleanup_evicts_stale_keys(self) -> None:
        with app_module.REQUEST_RATE_LIMIT_LOCK:
            app_module.REQUEST_RATE_LIMIT_BUCKETS.clear()
            app_module.REQUEST_RATE_LIMIT_BUCKETS[("calc-sync", "old")] = deque([100.0])
            app_module.REQUEST_RATE_LIMIT_BUCKETS[("calc-sync", "mixed")] = deque([939.0, 945.0])
            app_module._REQUEST_RATE_LIMIT_LAST_SWEEP_TS = 0.0
            app_module._cleanup_rate_limit_buckets_locked(now=1000.0, window_start=940.0)
            snapshot = {key: list(value) for key, value in app_module.REQUEST_RATE_LIMIT_BUCKETS.items()}
            last_sweep = app_module._REQUEST_RATE_LIMIT_LAST_SWEEP_TS

        self.assertNotIn(("calc-sync", "old"), snapshot)
        self.assertEqual(snapshot.get(("calc-sync", "mixed")), [945.0])
        self.assertEqual(last_sweep, 1000.0)

    def test_active_job_cap_enforced_per_ip(self) -> None:
        with patch.object(app_module, "CALCULATOR_MAX_ACTIVE_JOBS_PER_IP", 1), patch.object(
            app_module.CALCULATOR_JOB_EXECUTOR,
            "submit",
            return_value=None,
        ):
            first = self.client.post("/api/calculate/jobs", json={})
            second = self.client.post("/api/calculate/jobs", json={})

        self.assertEqual(first.status_code, 202)
        first_payload = first.json()
        self.assertEqual(first_payload.get("status"), "queued")
        self.assertEqual(first_payload.get("queue_position"), 1)
        self.assertEqual(first_payload.get("queued_jobs"), 1)
        self.assertEqual(first_payload.get("running_jobs"), 0)
        self.assertEqual(second.status_code, 429)

    def test_cancel_queued_job_marks_job_cancelled(self) -> None:
        fake_future = Mock()
        fake_future.cancel.return_value = True

        with patch.object(
            app_module.CALCULATOR_JOB_EXECUTOR,
            "submit",
            return_value=fake_future,
        ):
            create = self.client.post("/api/calculate/jobs", json={})

        self.assertEqual(create.status_code, 202)
        job_id = create.json()["job_id"]

        cancel_resp = self.client.delete(f"/api/calculate/jobs/{job_id}")
        self.assertEqual(cancel_resp.status_code, 200)
        payload = cancel_resp.json()
        self.assertEqual(payload.get("status"), app_module.CALC_JOB_CANCELLED_STATUS)
        self.assertEqual(payload.get("error", {}).get("status_code"), 499)
        fake_future.cancel.assert_called_once()

        status_resp = self.client.get(f"/api/calculate/jobs/{job_id}")
        self.assertEqual(status_resp.status_code, 200)
        self.assertEqual(status_resp.json().get("status"), app_module.CALC_JOB_CANCELLED_STATUS)

    def test_cancel_running_job_marks_job_cancelled(self) -> None:
        with patch.object(
            app_module.CALCULATOR_JOB_EXECUTOR,
            "submit",
            return_value=None,
        ):
            create = self.client.post("/api/calculate/jobs", json={})

        self.assertEqual(create.status_code, 202)
        job_id = create.json()["job_id"]
        with app_module.CALCULATOR_JOB_LOCK:
            job = app_module.CALCULATOR_JOBS[job_id]
            job["status"] = "running"
            job["started_at"] = app_module._iso_now()
            client_ip = str(job.get("client_ip") or "")

        cancel_resp = self.client.delete(f"/api/calculate/jobs/{job_id}")
        self.assertEqual(cancel_resp.status_code, 200)
        payload = cancel_resp.json()
        self.assertEqual(payload.get("status"), app_module.CALC_JOB_CANCELLED_STATUS)
        self.assertEqual(payload.get("error", {}).get("status_code"), 499)

        with app_module.CALCULATOR_JOB_LOCK:
            self.assertEqual(app_module._active_jobs_for_ip(client_ip), 0)

    def test_cleanup_expires_cancelled_jobs(self) -> None:
        created_ts = time.time() - app_module.CALCULATOR_JOB_TTL_SECONDS - 5.0
        with app_module.CALCULATOR_JOB_LOCK:
            app_module.CALCULATOR_JOBS.clear()
            app_module.CALCULATOR_JOBS["cancelled-job"] = {
                "job_id": "cancelled-job",
                "status": app_module.CALC_JOB_CANCELLED_STATUS,
                "created_ts": created_ts,
            }
            app_module._cleanup_calculation_jobs(time.time())
            self.assertNotIn("cancelled-job", app_module.CALCULATOR_JOBS)

    def test_calculation_job_lifecycle_success(self) -> None:
        fake_result = {"total": 1, "settings": {}, "data": [{"Player": "Jane Roe"}]}

        with patch.object(
            app_module,
            "_run_calculate_request",
            return_value=fake_result,
        ):
            create = self.client.post("/api/calculate/jobs", json={})
            self.assertEqual(create.status_code, 202)
            job_id = create.json()["job_id"]

            deadline = time.time() + 2.0
            while True:
                status = self.client.get(f"/api/calculate/jobs/{job_id}")
                self.assertEqual(status.status_code, 200)
                payload = status.json()
                if payload["status"] == "completed":
                    break
                if payload["status"] == "failed":
                    self.fail(f"Expected completed job, got failed payload: {payload}")
                if time.time() > deadline:
                    self.fail(f"Timed out waiting for job completion: {payload}")
                time.sleep(0.01)

            self.assertEqual(payload["result"]["total"], 1)

    def test_calculation_job_lifecycle_failure(self) -> None:
        def fake_failure(*args, **kwargs):
            raise HTTPException(status_code=422, detail="Not enough players for selected settings.")

        with patch.object(
            app_module,
            "_run_calculate_request",
            side_effect=fake_failure,
        ):
            create = self.client.post("/api/calculate/jobs", json={})
            self.assertEqual(create.status_code, 202)
            job_id = create.json()["job_id"]

            deadline = time.time() + 2.0
            while True:
                status = self.client.get(f"/api/calculate/jobs/{job_id}")
                self.assertEqual(status.status_code, 200)
                payload = status.json()
                if payload["status"] == "failed":
                    break
                if payload["status"] == "completed":
                    self.fail(f"Expected failed job, got completed payload: {payload}")
                if time.time() > deadline:
                    self.fail(f"Timed out waiting for failed status: {payload}")
                time.sleep(0.01)

            self.assertEqual(payload["error"]["status_code"], 422)
            self.assertIn("Not enough players", payload["error"]["detail"])
