"""CSV/XLSX export and record-serialization helpers."""

from __future__ import annotations

import io
import json
import math
from datetime import date, datetime
from typing import Literal

import pandas as pd
from fastapi.responses import StreamingResponse
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def as_float(value: object) -> float | None:
    try:
        parsed = float(value)
    except (TypeError, ValueError):
        return None
    if math.isfinite(parsed):
        return parsed
    return None


def clean_records_for_json(records: list[dict]) -> list[dict]:
    for row in records:
        for key, value in row.items():
            if value is None:
                continue

            if isinstance(value, (int, float)) and not isinstance(value, bool):
                if not math.isfinite(float(value)):
                    row[key] = None
                continue

            try:
                if pd.isna(value):
                    row[key] = None
            except (TypeError, ValueError):
                continue
    return records


def flatten_explanations_for_export(explanations: dict[str, dict]) -> list[dict]:
    rows: list[dict] = []
    for player_id, detail in explanations.items():
        if not isinstance(detail, dict):
            continue
        per_year = detail.get("per_year")
        if not isinstance(per_year, list):
            continue
        for entry in per_year:
            if not isinstance(entry, dict):
                continue
            points = entry.get("points")
            points = points if isinstance(points, dict) else {}
            rows.append(
                {
                    "PlayerEntityKey": player_id,
                    "Player": detail.get("player"),
                    "Team": detail.get("team"),
                    "Pos": detail.get("pos"),
                    "Mode": detail.get("mode"),
                    "Year": entry.get("year"),
                    "YearValue": entry.get("year_value"),
                    "DiscountFactor": entry.get("discount_factor"),
                    "DiscountedContribution": entry.get("discounted_contribution"),
                    "HittingPoints": points.get("hitting_points"),
                    "PitchingPoints": points.get("pitching_points"),
                    "SelectedPoints": points.get("selected_points"),
                    "HittingRulePoints": json.dumps((points.get("hitting") or {}).get("rule_points", {}), sort_keys=True),
                    "PitchingRulePoints": json.dumps((points.get("pitching") or {}).get("rule_points", {}), sort_keys=True),
                }
            )
    return rows


def coerce_export_column_tokens(value: list[str] | tuple[str, ...] | set[str] | str | None) -> list[str]:
    if value is None:
        return []

    raw_tokens: list[object]
    if isinstance(value, str):
        raw_tokens = [value]
    elif isinstance(value, (list, tuple, set)):
        raw_tokens = list(value)
    else:
        return []

    seen: set[str] = set()
    out: list[str] = []
    for raw in raw_tokens:
        for token in str(raw or "").split(","):
            col = token.strip()
            if not col or col in seen:
                continue
            seen.add(col)
            out.append(col)
    return out


def ordered_columns_from_rows(rows: list[dict]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        for raw_key in row.keys():
            col = str(raw_key or "").strip()
            if not col or col in seen:
                continue
            seen.add(col)
            ordered.append(col)
    return ordered


def resolve_export_columns(
    *,
    available_columns: list[str],
    requested_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    default_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    required_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    disallowed_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
) -> list[str]:
    available = coerce_export_column_tokens(available_columns)
    available_set = set(available)
    disallowed_set = set(coerce_export_column_tokens(disallowed_columns))

    def select(source: list[str] | tuple[str, ...] | set[str] | str | None) -> list[str]:
        selected: list[str] = []
        seen_local: set[str] = set()
        for col in coerce_export_column_tokens(source):
            if col in seen_local or col in disallowed_set or col not in available_set:
                continue
            seen_local.add(col)
            selected.append(col)
        return selected

    requested = select(requested_columns)
    if not requested:
        requested = select(default_columns)
    if not requested:
        requested = select(available)

    required = select(required_columns)
    required_set = set(required)
    resolved = required + [col for col in requested if col not in required_set]
    if resolved:
        return resolved

    return [col for col in available if col not in disallowed_set]


def export_column_label(col: str, *, export_header_label_overrides: dict[str, str]) -> str:
    text = str(col or "").strip()
    if text.startswith("Value_"):
        suffix = text.split("_", 1)[1].strip() if "_" in text else ""
        return f"{suffix} Dyn Value" if suffix else text
    return export_header_label_overrides.get(text, text)


def is_missing_export_value(value: object) -> bool:
    if value is None:
        return True
    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def export_rounding_decimals(
    col: str,
    *,
    export_three_decimal_cols: set[str],
    export_two_decimal_cols: set[str],
    export_whole_number_cols: set[str],
    export_integer_cols: set[str],
) -> int | None:
    if col.startswith("Value_"):
        return 2
    if col in export_three_decimal_cols:
        return 3
    if col in export_two_decimal_cols:
        return 2
    if col in export_whole_number_cols or col in export_integer_cols:
        return 0
    return None


def round_export_value(value: object, decimals: int) -> object:
    if isinstance(value, bool):
        return value
    if is_missing_export_value(value):
        return None
    parsed = as_float(value)
    if parsed is None:
        return value
    if decimals <= 0:
        return int(round(parsed))
    return round(parsed, decimals)


def coerce_export_date_value(value: object) -> object:
    if is_missing_export_value(value):
        return None
    parsed = pd.to_datetime(value, errors="coerce")
    if pd.isna(parsed):
        return value
    return parsed.date()


def xlsx_number_format_for_decimals(decimals: int) -> str:
    if decimals <= 0:
        return "#,##0"
    if decimals == 1:
        return "#,##0.0"
    if decimals == 2:
        return "#,##0.00"
    return "#,##0.000"


def xlsx_apply_table_formatting(
    ws,
    df: pd.DataFrame,
    *,
    decimals_by_col: dict[str, int] | None = None,
    date_cols: set[str] | None = None,
) -> None:
    if ws.max_row < 1 or ws.max_column < 1:
        return

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    header_fill = PatternFill(fill_type="solid", fgColor="EEF3F8")
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    decimals_by_col = decimals_by_col or {}
    date_cols = date_cols or set()

    for col_idx, col_name in enumerate(df.columns, start=1):
        col_letter = get_column_letter(col_idx)
        sample_values = [str(col_name)]
        for value in df[col_name].head(200).tolist():
            if is_missing_export_value(value):
                continue
            sample_values.append(str(value))
        width = min(48, max(8, max((len(text) for text in sample_values), default=8) + 2))
        ws.column_dimensions[col_letter].width = width

        decimals = decimals_by_col.get(col_name)
        if decimals is not None:
            number_format = xlsx_number_format_for_decimals(decimals)
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                value = cell.value
                if value is None or isinstance(value, bool):
                    continue
                parsed = as_float(value)
                if parsed is None:
                    continue
                cell.value = int(round(parsed)) if decimals <= 0 else round(parsed, decimals)
                cell.number_format = number_format
            continue

        if col_name in date_cols:
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = "yyyy-mm-dd"


def prepare_data_export_frame(
    rows: list[dict],
    *,
    selected_columns: list[str],
    export_date_cols: set[str],
    export_header_label_overrides: dict[str, str],
    export_three_decimal_cols: set[str],
    export_two_decimal_cols: set[str],
    export_whole_number_cols: set[str],
    export_integer_cols: set[str],
) -> tuple[pd.DataFrame, dict[str, int], set[str]]:
    frame = pd.DataFrame.from_records(rows)
    for col in selected_columns:
        if col not in frame.columns:
            frame[col] = None
    frame = frame.reindex(columns=selected_columns)

    decimals_by_display: dict[str, int] = {}
    display_date_cols: set[str] = set()
    rename_map: dict[str, str] = {}

    for col in selected_columns:
        display_name = export_column_label(col, export_header_label_overrides=export_header_label_overrides)

        if col in export_date_cols and col in frame.columns:
            frame[col] = frame[col].map(coerce_export_date_value)
            display_date_cols.add(display_name)

        decimals = export_rounding_decimals(
            col,
            export_three_decimal_cols=export_three_decimal_cols,
            export_two_decimal_cols=export_two_decimal_cols,
            export_whole_number_cols=export_whole_number_cols,
            export_integer_cols=export_integer_cols,
        )
        if decimals is not None and col in frame.columns:
            frame[col] = frame[col].map(lambda value, d=decimals: round_export_value(value, d))
            decimals_by_display[display_name] = decimals

        rename_map[col] = display_name

    frame = frame.rename(columns=rename_map)
    return frame, decimals_by_display, display_date_cols


def prepare_explainability_export_frame(
    rows: list[dict],
    *,
    export_date_cols: set[str],
    export_three_decimal_cols: set[str],
    export_two_decimal_cols: set[str],
    export_whole_number_cols: set[str],
    export_integer_cols: set[str],
) -> tuple[pd.DataFrame, dict[str, int], set[str]]:
    frame = pd.DataFrame.from_records(rows)
    decimals_by_col: dict[str, int] = {}
    date_cols: set[str] = set()
    for col in list(frame.columns):
        if col in export_date_cols:
            frame[col] = frame[col].map(coerce_export_date_value)
            date_cols.add(col)
        decimals = export_rounding_decimals(
            col,
            export_three_decimal_cols=export_three_decimal_cols,
            export_two_decimal_cols=export_two_decimal_cols,
            export_whole_number_cols=export_whole_number_cols,
            export_integer_cols=export_integer_cols,
        )
        if decimals is not None:
            frame[col] = frame[col].map(lambda value, d=decimals: round_export_value(value, d))
            decimals_by_col[col] = decimals
    return frame, decimals_by_col, date_cols


def default_calculator_export_columns(
    rows: list[dict],
    *,
    calculator_result_stat_export_order: tuple[str, ...],
    calculator_result_points_export_order: tuple[str, ...],
    value_col_sort_key,
) -> list[str]:
    available = ordered_columns_from_rows(rows)
    if not available:
        return ["Player", "DynastyValue", "Age", "Team", "Pos"]

    available_set = set(available)
    year_cols = sorted(
        [col for col in available if col.startswith("Value_")],
        key=value_col_sort_key,
    )
    stat_cols = [col for col in calculator_result_stat_export_order if col in available_set]
    points_cols = [col for col in calculator_result_points_export_order if col in available_set]

    ordered: list[str] = []
    for col in ["Player", "DynastyValue", "Age", "Team", "Pos", *points_cols, *stat_cols, *year_cols]:
        if col in available_set and col not in ordered:
            ordered.append(col)
    return ordered


def tabular_export_response(
    rows: list[dict],
    *,
    filename_base: str,
    file_format: Literal["csv", "xlsx"],
    explain_rows: list[dict] | None = None,
    selected_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    default_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    required_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    disallowed_columns: list[str] | tuple[str, ...] | set[str] | str | None = None,
    export_date_cols: set[str],
    export_header_label_overrides: dict[str, str],
    export_three_decimal_cols: set[str],
    export_two_decimal_cols: set[str],
    export_whole_number_cols: set[str],
    export_integer_cols: set[str],
) -> StreamingResponse:
    available_columns = ordered_columns_from_rows(rows)
    resolved_columns = resolve_export_columns(
        available_columns=available_columns,
        requested_columns=selected_columns,
        default_columns=default_columns,
        required_columns=required_columns,
        disallowed_columns=disallowed_columns,
    )
    data_df, data_decimals, data_date_cols = prepare_data_export_frame(
        rows,
        selected_columns=resolved_columns,
        export_date_cols=export_date_cols,
        export_header_label_overrides=export_header_label_overrides,
        export_three_decimal_cols=export_three_decimal_cols,
        export_two_decimal_cols=export_two_decimal_cols,
        export_whole_number_cols=export_whole_number_cols,
        export_integer_cols=export_integer_cols,
    )

    if file_format == "csv":
        payload = data_df.to_csv(index=False).encode("utf-8")
        content_type = "text/csv; charset=utf-8"
        extension = "csv"
    else:
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            data_df.to_excel(writer, index=False, sheet_name="Data")
            xlsx_apply_table_formatting(
                writer.sheets["Data"],
                data_df,
                decimals_by_col=data_decimals,
                date_cols=data_date_cols,
            )
            if explain_rows:
                explain_df, explain_decimals, explain_date_cols = prepare_explainability_export_frame(
                    explain_rows,
                    export_date_cols=export_date_cols,
                    export_three_decimal_cols=export_three_decimal_cols,
                    export_two_decimal_cols=export_two_decimal_cols,
                    export_whole_number_cols=export_whole_number_cols,
                    export_integer_cols=export_integer_cols,
                )
                explain_df.to_excel(
                    writer,
                    index=False,
                    sheet_name="Explainability",
                )
                xlsx_apply_table_formatting(
                    writer.sheets["Explainability"],
                    explain_df,
                    decimals_by_col=explain_decimals,
                    date_cols=explain_date_cols,
                )
        payload = output.getvalue()
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        extension = "xlsx"

    response = StreamingResponse(io.BytesIO(payload), media_type=content_type)
    response.headers["Content-Disposition"] = f'attachment; filename="{filename_base}.{extension}"'
    return response
