"""Excel workbook formatting helpers used by dynasty export flows."""

from __future__ import annotations

from typing import Dict, Optional

import pandas as pd
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


def _xlsx_apply_header_style(ws) -> None:
    """Apply a consistent header style to row 1."""
    max_col = ws.max_column
    if max_col < 1:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    header_border = Border(bottom=thin)

    ws.row_dimensions[1].height = 22
    for col_num in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border


def _xlsx_set_freeze_filters_and_view(ws, freeze_panes: str, add_autofilter: bool = False) -> None:
    """Freeze panes, optionally add AutoFilter, and hide gridlines."""
    ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = False

    if add_autofilter:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = ref
    else:
        ws.auto_filter.ref = None


def _xlsx_add_table(ws, table_name: str, style_name: str = "TableStyleMedium9") -> None:
    """Wrap the used range in an Excel table."""
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    ws.auto_filter.ref = None
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName=table_name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _xlsx_set_column_widths(
    ws,
    df: pd.DataFrame,
    overrides: Optional[Dict[str, float]] = None,
    sample_rows: int = 1000,
    min_width: float = 8.0,
    max_width: float = 45.0,
) -> None:
    """Best-effort auto-fit with readability caps."""
    if df is None or df.empty:
        return

    resolved_overrides = dict(overrides or {})
    for col in df.columns:
        if isinstance(col, str) and col.startswith("Value_"):
            resolved_overrides.setdefault(col, 10.0)

    for i, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(i)
        if col_name in resolved_overrides:
            ws.column_dimensions[letter].width = float(resolved_overrides[col_name])
            continue

        col_series = df[col_name]
        if str(col_name).lower().endswith("date"):
            ws.column_dimensions[letter].width = 14.0
            continue
        if pd.api.types.is_numeric_dtype(col_series):
            base = max(len(str(col_name)) + 2, 10)
            ws.column_dimensions[letter].width = float(min(max(base, min_width), 16.0))
            continue
        if pd.api.types.is_bool_dtype(col_series):
            ws.column_dimensions[letter].width = float(max(12.0, len(str(col_name)) + 2))
            continue

        sample = col_series.dropna().astype(str).head(sample_rows)
        max_len = int(sample.str.len().max()) if not sample.empty else 0
        width = min(max(max_len, len(str(col_name))) + 2, int(max_width))
        ws.column_dimensions[letter].width = max(float(width), float(min_width))


def _xlsx_apply_number_formats(ws, df: pd.DataFrame, formats_by_col: Dict[str, str]) -> None:
    """Apply number formats to column cells for data rows only."""
    if df is None or df.empty:
        return

    max_row = ws.max_row
    if max_row < 2:
        return

    columns = list(df.columns)
    for col_name, fmt in formats_by_col.items():
        if col_name not in columns:
            continue
        col_idx = columns.index(col_name) + 1
        for row_idx in range(2, max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt


def _xlsx_add_value_color_scale(ws, df: pd.DataFrame, col_name: str) -> None:
    """Apply a red-yellow-green color scale to a value column."""
    if df is None or df.empty or col_name not in df.columns:
        return
    max_row = ws.max_row
    if max_row < 3:
        return

    col_idx = list(df.columns).index(col_name) + 1
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{max_row}"
    rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",
        end_type="max",
        end_color="63BE7B",
    )
    ws.conditional_formatting.add(cell_range, rule)


def _xlsx_format_player_values(ws, df: pd.DataFrame, table_name: str = "PlayerValuesTbl") -> None:
    """Format summary-tab workbook sheet."""
    _xlsx_apply_header_style(ws)
    _xlsx_set_freeze_filters_and_view(ws, freeze_panes="B2")
    _xlsx_add_table(ws, table_name=table_name)

    overrides = {
        "Player": 24.0,
        "Team": 8.0,
        "MLBTeam": 8.0,
        "Pos": 10.0,
        "OldestProjectionDate": 14.0,
        "DynastyValue": 12.0,
        "RawDynastyValue": 14.0,
        "CenteringBaselineMean": 16.0,
    }
    _xlsx_set_column_widths(ws, df, overrides=overrides)

    formats = {
        "Age": "0",
        "OldestProjectionDate": "yyyy-mm-dd",
        "DynastyValue": "0.00",
        "RawDynastyValue": "0.00",
        "CenteringBaselineMean": "0.00",
    }
    for col in df.columns:
        if isinstance(col, str) and col.startswith("Value_"):
            formats[col] = "0.00"
    _xlsx_apply_number_formats(ws, df, formats)
    _xlsx_add_value_color_scale(ws, df, "DynastyValue")


def _xlsx_format_detail_sheet(
    ws,
    df: pd.DataFrame,
    *,
    table_name: str,
    is_pitch: bool,
) -> None:
    """Format Bat_Aggregated / Pitch_Aggregated workbook sheets."""
    _ = is_pitch
    _xlsx_apply_header_style(ws)
    _xlsx_set_freeze_filters_and_view(ws, freeze_panes="C2")
    _xlsx_add_table(ws, table_name=table_name)

    overrides = {
        "Player": 24.0,
        "Team": 8.0,
        "MLBTeam": 8.0,
        "Pos": 10.0,
        "BestSlot": 10.0,
        "OldestProjectionDate": 14.0,
        "DynastyValue": 12.0,
        "RawDynastyValue": 14.0,
        "YearValue": 10.0,
    }
    _xlsx_set_column_widths(ws, df, overrides=overrides)

    formats: Dict[str, str] = {
        "Year": "0",
        "Age": "0",
        "OldestProjectionDate": "yyyy-mm-dd",
        "YearValue": "0.00",
        "DynastyValue": "0.00",
        "RawDynastyValue": "0.00",
        "AVG": "0.000",
        "OBP": "0.000",
        "SLG": "0.000",
        "OPS": "0.000",
        "ERA": "0.00",
        "WHIP": "0.00",
        "IP": "0.0",
    }
    _xlsx_apply_number_formats(ws, df, formats)
    _xlsx_add_value_color_scale(ws, df, "YearValue")
    _xlsx_add_value_color_scale(ws, df, "DynastyValue")
