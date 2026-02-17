"""
dynasty_roto_values.py

Unified script for dynasty roto player values.

Modes:
- common: 5x5 roto defaults, typical lineup, no IP cap by default.
- league: custom league settings (SP/RP/P slots, IP min/max, OPS/SVH/QA3 cats).

Default dynasty roto configuration:
- 12-team roto
- 5x5 categories:
    Hitters: R, RBI, HR, SB, AVG
    Pitchers: W, K, SV, ERA, WHIP
- No IP cap (by default)
- Typical roto lineup:
    Hitters: C(1), 1B, 2B, 3B, SS, CI, MI, OF(5), UT
    Pitchers: P(9)
- Roster (common defaults): 28 total = 22 starters + 6 bench + 0 minors + 0 IL.
- Dynasty value: discounted multi-year marginal roto points (SGP),
  then centered so that ~0 corresponds to the replacement-level roster
  cutoff (active + bench).

Input format (same as your file):
Excel workbook with sheets:
- "Bat" with columns at least:
  Player, Year, Team, Age, Pos, AB, H, R, HR, RBI, SB
- "Pitch" with columns at least:
  Player, Year, Team, Age, Pos, IP, W, K, ER, H, BB
  (SV is recommended; if missing, we attempt fallbacks)
Optional:
  - A minors eligibility column in either sheet named like MinorEligible, Minor, or minor_eligible.
If multiple projections exist for the same Player/Year, the script averages the
three most recent rows (by projection date column if present, otherwise file order)
over counting stats, then recomputes rate stats (AVG/OBP/OPS/ERA/WHIP).

Outputs:
- common_player_values.csv
- common_player_values.xlsx

Run (common mode):
  python dynasty_roto_values.py common --input "Dynasty Baseball Projections.xlsx" --start-year 2026

Optional (common):
  python dynasty_roto_values.py common --input "Dynasty Baseball Projections.xlsx" --teams 12 --sims 200 --horizon 10

Run (league mode):
  python dynasty_roto_values.py league --input "Dynasty Baseball Projections.xlsx" --start-year 2026

Dependencies:
  pip install pandas numpy openpyxl
Optional (faster/better slot assignment):
  pip install scipy
"""

from __future__ import annotations

import argparse
import re
from typing import Callable, Dict, List, Optional, Set, Tuple, Union

import numpy as np
import pandas as pd

try:
    from backend.valuation.assignment import (
        HAVE_SCIPY,
        assign_players_to_slots,
        assign_players_to_slots_with_vacancy_fill,
        build_slot_list,
        build_team_slot_template,
        expand_slot_counts,
        league_assign_players_to_slots,
        league_build_slot_list,
        league_build_team_slot_template,
        league_expand_slot_counts,
        validate_assigned_slots,
    )
    from backend.valuation.models import (
        CommonDynastyRotoSettings,
        HIT_CATS,
        HIT_COMPONENT_COLS,
        LEAGUE_HIT_STAT_COLS,
        LeagueSettings,
        PIT_CATS,
        PIT_COMPONENT_COLS,
    )
    from backend.valuation.positions import (
        eligible_hit_slots,
        eligible_pit_slots,
        league_eligible_hit_slots,
        league_eligible_pit_slots,
        league_parse_hit_positions,
        league_parse_pit_positions,
        parse_hit_positions,
        parse_pit_positions,
    )
except ImportError:
    # Support direct execution/import when /backend is added to sys.path.
    from valuation.assignment import (
        HAVE_SCIPY,
        assign_players_to_slots,
        assign_players_to_slots_with_vacancy_fill,
        build_slot_list,
        build_team_slot_template,
        expand_slot_counts,
        league_assign_players_to_slots,
        league_build_slot_list,
        league_build_team_slot_template,
        league_expand_slot_counts,
        validate_assigned_slots,
    )
    from valuation.models import (
        CommonDynastyRotoSettings,
        HIT_CATS,
        HIT_COMPONENT_COLS,
        LEAGUE_HIT_STAT_COLS,
        LeagueSettings,
        PIT_CATS,
        PIT_COMPONENT_COLS,
    )
    from valuation.positions import (
        eligible_hit_slots,
        eligible_pit_slots,
        league_eligible_hit_slots,
        league_eligible_pit_slots,
        league_parse_hit_positions,
        league_parse_pit_positions,
        parse_hit_positions,
        parse_pit_positions,
    )

# Excel formatting (output workbook)
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

# Projection de-duplication helpers
PROJECTION_DATE_COLS = ["ProjectionDate", "Date", "Updated", "LastUpdated", "Timestamp", "Created", "AsOf"]
PLAYER_KEY_COL = "PlayerKey"
PLAYER_ENTITY_KEY_COL = "PlayerEntityKey"
PLAYER_KEY_PATTERN = re.compile(r"[^a-z0-9]+")

# Bench-stash penalty curve defaults:
# - first stash round per team should still carry a small cost
# - later rounds are progressively more punitive
BENCH_STASH_MIN_PENALTY = 0.10
BENCH_STASH_MAX_PENALTY = 0.85
BENCH_STASH_PENALTY_GAMMA = 1.35


def positive_int_arg(value: Union[str, int]) -> int:
    """argparse type: integer >= 1."""
    try:
        ivalue = int(value)
    except (TypeError, ValueError) as exc:
        raise argparse.ArgumentTypeError(f"Expected an integer, got: {value!r}") from exc
    if ivalue < 1:
        raise argparse.ArgumentTypeError(f"Expected an integer >= 1, got: {ivalue}")
    return ivalue


def non_negative_int_arg(value: Union[str, int]) -> int:
    """argparse type: integer >= 0."""
    try:
        ivalue = int(value)
    except (TypeError, ValueError) as exc:
        raise argparse.ArgumentTypeError(f"Expected an integer, got: {value!r}") from exc
    if ivalue < 0:
        raise argparse.ArgumentTypeError(f"Expected an integer >= 0, got: {ivalue}")
    return ivalue


def non_negative_float_arg(value: Union[str, float]) -> float:
    """argparse type: float >= 0."""
    try:
        fvalue = float(value)
    except (TypeError, ValueError) as exc:
        raise argparse.ArgumentTypeError(f"Expected a float, got: {value!r}") from exc
    if fvalue < 0.0:
        raise argparse.ArgumentTypeError(f"Expected a float >= 0, got: {fvalue}")
    return fvalue


def discount_arg(value: Union[str, float]) -> float:
    """argparse type: annual discount factor in the interval (0, 1]."""
    try:
        fvalue = float(value)
    except (TypeError, ValueError) as exc:
        raise argparse.ArgumentTypeError(f"Expected a float, got: {value!r}") from exc
    if not (0.0 < fvalue <= 1.0):
        raise argparse.ArgumentTypeError(f"Expected discount in (0, 1], got: {fvalue}")
    return fvalue


def optional_non_negative_float_arg(value: Union[str, float]) -> Optional[float]:
    """argparse type: float >= 0, or None for disabled limits."""
    if value is None:
        return None

    if isinstance(value, str):
        lowered = value.strip().lower()
        if lowered in {"none", "null", "off", "no", "disabled"}:
            return None

    try:
        fvalue = float(value)
    except (TypeError, ValueError) as exc:
        raise argparse.ArgumentTypeError(f"Expected a float or 'none', got: {value!r}") from exc

    if fvalue < 0.0:
        raise argparse.ArgumentTypeError(f"Expected a float >= 0 or 'none', got: {fvalue}")
    return fvalue


def validate_ip_bounds(ip_min: float, ip_max: Optional[float]) -> None:
    """Ensure optional IP bounds are internally consistent."""
    if ip_min < 0:
        raise ValueError(f"ip_min must be >= 0, got {ip_min}")
    if ip_max is not None and ip_max < ip_min:
        raise ValueError(f"ip_max ({ip_max}) must be >= ip_min ({ip_min})")


def _find_projection_date_col(df: pd.DataFrame) -> Optional[str]:
    for col in PROJECTION_DATE_COLS:
        if col in df.columns:
            return col
    return None


def _normalize_player_key(value: object) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return "unknown-player"
    key = PLAYER_KEY_PATTERN.sub("-", text).strip("-")
    return key or "unknown-player"


def _normalize_team_key(value: object) -> str:
    return str(value or "").strip().upper()


def _normalize_year_key(value: object) -> str:
    if value is None or value == "":
        return ""
    try:
        numeric = float(value)
        if pd.notna(numeric) and numeric.is_integer():
            return str(int(numeric))
    except (TypeError, ValueError):
        pass
    return str(value).strip()


def _team_column_for_dataframe(df: pd.DataFrame) -> Optional[str]:
    for col in ("Team", "MLBTeam"):
        if col in df.columns:
            return col
    return None


def _add_player_identity_keys(
    bat: pd.DataFrame,
    pit: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    bat_out = bat.copy()
    pit_out = pit.copy()

    def _prepare(df: pd.DataFrame, *, team_col: Optional[str]) -> pd.DataFrame:
        out = df.copy()

        if PLAYER_KEY_COL in out.columns:
            existing_keys = out[PLAYER_KEY_COL].astype("string").fillna("").str.strip()
        else:
            existing_keys = pd.Series("", index=out.index, dtype="string")
        normalized_keys = out.get("Player", pd.Series("", index=out.index)).map(_normalize_player_key)
        out["_player_key"] = existing_keys.where(existing_keys != "", normalized_keys)

        if "Year" in out.columns:
            out["_year_key"] = out["Year"].map(_normalize_year_key)
        else:
            out["_year_key"] = ""

        if team_col and team_col in out.columns:
            out["_team_key"] = out[team_col].map(_normalize_team_key)
        else:
            out["_team_key"] = ""

        return out

    bat_prepared = _prepare(bat_out, team_col=_team_column_for_dataframe(bat_out))
    pit_prepared = _prepare(pit_out, team_col=_team_column_for_dataframe(pit_out))
    combined = pd.concat([bat_prepared, pit_prepared], ignore_index=True, sort=False)

    teams_by_player_year: dict[tuple[str, str], set[str]] = {}
    for _, row in combined.iterrows():
        player_key = str(row.get("_player_key", "")).strip()
        year_key = str(row.get("_year_key", "")).strip()
        team_key = str(row.get("_team_key", "")).strip()
        if not player_key or not team_key:
            continue
        teams_by_player_year.setdefault((player_key, year_key), set()).add(team_key)

    ambiguous_players = {
        player_key
        for (player_key, _), teams in teams_by_player_year.items()
        if len(teams) > 1
    }

    def _finalize(df: pd.DataFrame) -> pd.DataFrame:
        out = df.copy()
        out[PLAYER_KEY_COL] = out["_player_key"].map(lambda value: str(value).strip() or "unknown-player")

        if PLAYER_ENTITY_KEY_COL in out.columns:
            existing_entities = out[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()
        else:
            existing_entities = pd.Series("", index=out.index, dtype="string")

        entity_values: list[str] = []
        for idx, row in out.iterrows():
            existing_entity = str(existing_entities.loc[idx]).strip()
            if existing_entity:
                entity_values.append(existing_entity)
                continue

            player_key = str(row.get(PLAYER_KEY_COL, "")).strip() or "unknown-player"
            if player_key in ambiguous_players:
                team_key = str(row.get("_team_key", "")).strip().lower() or "unknown"
                entity_values.append(f"{player_key}__{team_key}")
            else:
                entity_values.append(player_key)

        out[PLAYER_ENTITY_KEY_COL] = entity_values
        return out.drop(columns=["_player_key", "_year_key", "_team_key"], errors="ignore")

    return _finalize(bat_prepared), _finalize(pit_prepared)


def _build_player_identity_lookup(bat: pd.DataFrame, pit: pd.DataFrame) -> pd.DataFrame:
    parts: list[pd.DataFrame] = []
    for frame in (bat, pit):
        if frame.empty or PLAYER_ENTITY_KEY_COL not in frame.columns:
            continue
        subset = frame[[PLAYER_ENTITY_KEY_COL, PLAYER_KEY_COL, "Player"]].copy()
        subset[PLAYER_ENTITY_KEY_COL] = subset[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()
        subset[PLAYER_KEY_COL] = subset[PLAYER_KEY_COL].astype("string").fillna("").str.strip()
        subset["Player"] = subset["Player"].astype("string").fillna("").str.strip()
        parts.append(subset)

    if not parts:
        return pd.DataFrame(columns=[PLAYER_ENTITY_KEY_COL, PLAYER_KEY_COL, "Player"])

    merged = pd.concat(parts, ignore_index=True).dropna(subset=[PLAYER_ENTITY_KEY_COL])
    merged = merged[merged[PLAYER_ENTITY_KEY_COL] != ""]
    merged = merged.drop_duplicates(subset=[PLAYER_ENTITY_KEY_COL], keep="first").reset_index(drop=True)
    merged[PLAYER_KEY_COL] = merged.apply(
        lambda row: (
            str(row.get(PLAYER_KEY_COL) or "").strip()
            or str(row.get(PLAYER_ENTITY_KEY_COL) or "").split("__", 1)[0]
            or _normalize_player_key(row.get("Player"))
        ),
        axis=1,
    )
    return merged[[PLAYER_ENTITY_KEY_COL, PLAYER_KEY_COL, "Player"]]


def _attach_identity_columns_to_output(out: pd.DataFrame, identity_lookup: pd.DataFrame) -> pd.DataFrame:
    if "Player" not in out.columns:
        return out

    result = out.rename(columns={"Player": PLAYER_ENTITY_KEY_COL}).copy()
    if identity_lookup.empty:
        result["Player"] = result[PLAYER_ENTITY_KEY_COL]
        result[PLAYER_KEY_COL] = result[PLAYER_ENTITY_KEY_COL].astype("string").str.split("__").str[0]
    else:
        result = result.merge(identity_lookup, on=PLAYER_ENTITY_KEY_COL, how="left")
        result["Player"] = result["Player"].fillna(result[PLAYER_ENTITY_KEY_COL])
        result[PLAYER_KEY_COL] = result[PLAYER_KEY_COL].fillna(
            result[PLAYER_ENTITY_KEY_COL].astype("string").str.split("__").str[0]
        )

    front = ["Player", PLAYER_KEY_COL, PLAYER_ENTITY_KEY_COL]
    rest = [c for c in result.columns if c not in front]
    return result[front + rest]


def average_recent_projections(
    df: pd.DataFrame,
    stat_cols: List[str],
    group_cols: Optional[List[str]] = None,
    max_entries: int = 3,
) -> pd.DataFrame:
    """
    If multiple projections exist for the same (Player, Year), average the most recent
    `max_entries` rows (by projection date column if present, otherwise file order).

    Adds two columns to the averaged output:
      - ProjectionsUsed: number of rows actually averaged (<= max_entries)
      - OldestProjectionDate: the oldest projection date among the rows averaged
        (i.e., the "third-oldest" among the selected rows when max_entries=3).
        If no projection date column exists, this will be NaT.
    """
    if max_entries < 1:
        raise ValueError("max_entries must be >= 1")

    df = df.copy()
    group_cols = group_cols or ["Player", "Year"]
    effective_group_cols = list(group_cols)

    missing_group_cols = [c for c in group_cols if c not in df.columns]
    if missing_group_cols:
        raise ValueError(f"average_recent_projections missing required group columns: {missing_group_cols}")

    team_col = _team_column_for_dataframe(df)
    if (
        team_col
        and "Player" in group_cols
        and "Year" in group_cols
        and "_entity_team" not in effective_group_cols
    ):
        team_values = df[team_col].astype("string").fillna("").str.strip()
        team_nonempty = team_values.where(team_values != "", pd.NA)
        team_counts = team_nonempty.groupby([df[c] for c in group_cols], dropna=False).transform("nunique")
        if team_counts.gt(1).any():
            # Split only ambiguous same-name/same-year groups where teams differ.
            df["_entity_team"] = team_values.where(team_counts > 1, "")
            effective_group_cols.append("_entity_team")

    date_col = _find_projection_date_col(df)

    df["_projection_order"] = np.arange(len(df))

    # Always create _projection_date so downstream logic can rely on it.
    if date_col:
        df["_projection_date"] = pd.to_datetime(df[date_col], errors="coerce")
        df["_sort_key"] = df["_projection_date"].fillna(pd.Timestamp.min)
    else:
        df["_projection_date"] = pd.NaT
        df["_sort_key"] = df["_projection_order"]

    # Keep up to `max_entries` most-recent rows per (Player, Year)
    recent = (
        df.sort_values(["_sort_key", "_projection_order"], ascending=False)
        .groupby(effective_group_cols, as_index=False, sort=False)
        .head(max_entries)
    )

    # Per-row markers so we can aggregate to group-level metadata
    recent["ProjectionsUsed"] = 1
    recent["OldestProjectionDate"] = recent["_projection_date"]

    stat_cols_present = [c for c in stat_cols if c in recent.columns]

    # Meta cols are carried forward from the most recent row (same behavior as before)
    meta_cols = [
        c for c in recent.columns
        if c not in stat_cols_present
        and c not in effective_group_cols
        and c
        not in {
            "_projection_order",
            "_projection_date",
            "_sort_key",
            "ProjectionsUsed",
            "OldestProjectionDate",
        }
    ]

    agg: Dict[str, str] = {c: "mean" for c in stat_cols_present}
    agg["ProjectionsUsed"] = "sum"
    agg["OldestProjectionDate"] = "min"
    for c in meta_cols:
        agg[c] = "first"

    out = (
        recent.sort_values(["_sort_key", "_projection_order"], ascending=False)
        .groupby(effective_group_cols, as_index=False, sort=False)
        .agg(agg)
    )
    out = out.drop(columns=["_entity_team"], errors="ignore")

    # Nice column order: group keys, then the new metadata columns, then everything else
    front = list(group_cols) + ["ProjectionsUsed", "OldestProjectionDate"]
    rest = [c for c in out.columns if c not in front]
    return out[front + rest]


def projection_meta_for_start_year(
    bat_df: pd.DataFrame,
    pit_df: pd.DataFrame,
    start_year: int,
) -> pd.DataFrame:
    """
    Produce one row per Player with:
      - ProjectionsUsed: max of Bat vs Pitch (handles hitter-only / pitcher-only / two-way)
      - OldestProjectionDate: min (oldest) of Bat vs Pitch dates

    Uses the already-averaged frames produced by average_recent_projections().
    """
    cols = ["Player", "ProjectionsUsed", "OldestProjectionDate"]

    def _subset_projection_meta(df: pd.DataFrame) -> pd.DataFrame:
        """Return a safe start-year projection metadata slice.

        Handles empty frames and cases where upstream processing provided no rows,
        ensuring we still return the expected metadata columns.
        """
        if df.empty:
            return pd.DataFrame(columns=cols)

        missing = [c for c in ["Year"] + cols if c not in df.columns]
        if missing:
            raise ValueError(f"projection metadata is missing required columns: {missing}")

        return df.loc[df["Year"] == start_year, cols].copy()

    b = _subset_projection_meta(bat_df)
    p = _subset_projection_meta(pit_df)

    m = b.merge(p, on="Player", how="outer", suffixes=("_bat", "_pit"))

    # How many projections were used (cap is enforced upstream by max_entries)
    m["ProjectionsUsed"] = m[["ProjectionsUsed_bat", "ProjectionsUsed_pit"]].max(axis=1, skipna=True)
    m["ProjectionsUsed"] = m["ProjectionsUsed"].round().astype("Int64")

    # Oldest date among whichever side exists (and the min if both exist)
    m["OldestProjectionDate"] = m[["OldestProjectionDate_bat", "OldestProjectionDate_pit"]].min(axis=1, skipna=True)

    # Store as date-only (no time) for cleaner Excel display
    m["OldestProjectionDate"] = pd.to_datetime(m["OldestProjectionDate"], errors="coerce").dt.date

    return m[["Player", "ProjectionsUsed", "OldestProjectionDate"]]

# ----------------------------
# Helpers: recent-projection averaging + detail sheet formatting
# ----------------------------

DERIVED_HIT_RATE_COLS: Set[str] = {"AVG", "OBP", "OPS"}
DERIVED_PIT_RATE_COLS: Set[str] = {"ERA", "WHIP"}


def numeric_stat_cols_for_recent_avg(
    df: pd.DataFrame,
    group_cols: Optional[List[str]] = None,
    exclude_cols: Optional[Set[str]] = None,
) -> List[str]:
    """Return numeric columns that should be averaged when collapsing projections.

    This is used both for valuation (so that any categories you care about get averaged)
    and for building "detail" output tabs that closely match the input sheets.
    """
    group_cols = group_cols or ["Player", "Year"]
    exclude_cols = set(exclude_cols or set())

    cols: List[str] = []
    for c in df.columns:
        if c in group_cols or c in exclude_cols:
            continue
        if pd.api.types.is_numeric_dtype(df[c]):
            cols.append(c)
    return cols


def reorder_detail_columns(
    df: pd.DataFrame,
    input_cols: List[str],
    add_after: Optional[str] = None,
    extra_cols: Optional[List[str]] = None,
) -> pd.DataFrame:
    """Reorder a detail DataFrame to resemble the original input sheet.

    - Start with input_cols order (only those present).
    - Insert extra_cols immediately after `add_after` if provided and present.
    - Append any remaining columns at the end.
    """
    df = df.copy()

    base = [c for c in input_cols if c in df.columns]
    extras = [c for c in (extra_cols or []) if c in df.columns and c not in base]

    if add_after and add_after in base and extras:
        idx = base.index(add_after) + 1
        ordered = base[:idx] + extras + base[idx:]
    else:
        ordered = base + extras

    remaining = [c for c in df.columns if c not in ordered]
    return df[ordered + remaining]





def recompute_common_rates_hit(df: pd.DataFrame) -> pd.DataFrame:
    """Recompute rate stats after averaging counting components.

    Common mode uses OBP/OPS as categories, while AVG remains useful for
    display. Recompute all three from counting components when possible so
    aggregated rows stay internally consistent.
    """
    df = df.copy()

    # AVG = H / AB
    if "H" in df.columns and "AB" in df.columns:
        h = df["H"].to_numpy(dtype=float)
        ab = df["AB"].to_numpy(dtype=float)
        df["AVG"] = np.divide(h, ab, out=np.zeros_like(h), where=ab > 0)

    # OBP + OPS (OPS = OBP + SLG)
    needed = {"H", "2B", "3B", "HR", "BB", "HBP", "AB", "SF"}
    if needed.issubset(df.columns):
        h = df["H"].to_numpy(dtype=float)
        b2 = df["2B"].to_numpy(dtype=float)
        b3 = df["3B"].to_numpy(dtype=float)
        hr = df["HR"].to_numpy(dtype=float)
        bb = df["BB"].to_numpy(dtype=float)
        hbp = df["HBP"].to_numpy(dtype=float)
        ab = df["AB"].to_numpy(dtype=float)
        sf = df["SF"].to_numpy(dtype=float)

        # TB = 1B + 2*2B + 3*3B + 4*HR, and 1B = H - 2B - 3B - HR
        # => TB = H + 2B + 2*3B + 3*HR
        tb = h + b2 + 2.0 * b3 + 3.0 * hr

        obp_den = ab + bb + hbp + sf
        obp = np.divide(h + bb + hbp, obp_den, out=np.zeros_like(obp_den), where=obp_den > 0)
        slg = np.divide(tb, ab, out=np.zeros_like(ab), where=ab > 0)

        df["OBP"] = obp
        df["OPS"] = obp + slg

    return df

def recompute_common_rates_pit(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "ER" in df.columns and "IP" in df.columns:
        er = df["ER"].to_numpy(dtype=float)
        ip = df["IP"].to_numpy(dtype=float)
        df["ERA"] = np.divide(9.0 * er, ip, out=np.full_like(ip, np.nan), where=ip > 0)
    if "H" in df.columns and "BB" in df.columns and "IP" in df.columns:
        h = df["H"].to_numpy(dtype=float)
        bb = df["BB"].to_numpy(dtype=float)
        ip = df["IP"].to_numpy(dtype=float)
        df["WHIP"] = np.divide(h + bb, ip, out=np.full_like(ip, np.nan), where=ip > 0)
    return df


def recompute_league_rates_hit(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if {"H", "2B", "3B", "HR"}.issubset(df.columns):
        df["TB"] = df["H"] + df["2B"] + 2 * df["3B"] + 3 * df["HR"]
    if {"H", "BB", "HBP", "AB", "SF"}.issubset(df.columns):
        df["OBP_num"] = df["H"] + df["BB"] + df["HBP"]
        df["OBP_den"] = df["AB"] + df["BB"] + df["HBP"] + df["SF"]
    if "H" in df.columns and "AB" in df.columns:
        h = df["H"].to_numpy(dtype=float)
        ab = df["AB"].to_numpy(dtype=float)
        df["AVG"] = np.divide(h, ab, out=np.zeros_like(h), where=ab > 0)
    if {"OBP_num", "OBP_den", "TB", "AB"}.issubset(df.columns):
        obp_den = df["OBP_den"].to_numpy(dtype=float)
        ab = df["AB"].to_numpy(dtype=float)
        obp = np.divide(df["OBP_num"].to_numpy(dtype=float), obp_den, out=np.zeros_like(obp_den), where=obp_den > 0)
        slg = np.divide(df["TB"].to_numpy(dtype=float), ab, out=np.zeros_like(ab), where=ab > 0)
        df["OPS"] = obp + slg
    return df


def recompute_league_rates_pit(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    if "ER" in df.columns and "IP" in df.columns:
        er = df["ER"].to_numpy(dtype=float)
        ip = df["IP"].to_numpy(dtype=float)
        df["ERA"] = np.divide(9.0 * er, ip, out=np.full_like(ip, np.nan), where=ip > 0)
    if "H" in df.columns and "BB" in df.columns and "IP" in df.columns:
        h = df["H"].to_numpy(dtype=float)
        bb = df["BB"].to_numpy(dtype=float)
        ip = df["IP"].to_numpy(dtype=float)
        df["WHIP"] = np.divide(h + bb, ip, out=np.full_like(ip, np.nan), where=ip > 0)
    return df

# ----------------------------
# Column aliases and requirements
# ----------------------------

COMMON_COLUMN_ALIASES = {
    "mlbteam": "Team",
    "team": "Team",
    "player_name": "Player",
    "name": "Player",
}

LEAGUE_COLUMN_ALIASES = {
    "team": "MLBTeam",
    "mlb_team": "MLBTeam",
    "player_name": "Player",
    "name": "Player",
}


def require_cols(df: pd.DataFrame, cols: List[str], sheet_name: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise ValueError(f"Sheet '{sheet_name}' is missing required columns: {missing}")


def normalize_input_schema(df: pd.DataFrame, aliases: Dict[str, str]) -> pd.DataFrame:
    """Normalize incoming sheet columns (trim + alias mapping) while preserving existing names."""
    out = df.copy()
    out.columns = [str(c).strip() for c in out.columns]

    lower_to_actual = {c.lower(): c for c in out.columns}
    rename_map: Dict[str, str] = {}
    for alias, canonical in aliases.items():
        actual = lower_to_actual.get(alias.lower())
        if actual and canonical not in out.columns:
            rename_map[actual] = canonical

    if rename_map:
        out = out.rename(columns=rename_map)
    return out


# ----------------------------
# Utility: z-scores for initial starter-pool weights
# (only used to construct baseline + starter pool; not the final valuation)
# ----------------------------

def zscore(s: pd.Series) -> pd.Series:
    x = s.astype(float)
    mu = float(x.mean())
    sd = float(x.std(ddof=0))
    if sd == 0.0 or np.isnan(sd):
        return x * 0.0
    return (x - mu) / sd

def initial_hitter_weight(df: pd.DataFrame) -> pd.Series:
    """
    Rough first-pass weight to select/assign starters with positional scarcity.
    Uses 5x5 categories with AVG translated into hit surplus over average.
    """
    df = df.copy()
    h = df["H"].astype(float)
    ab = df["AB"].astype(float)
    avg = np.divide(h, ab, out=np.zeros_like(ab, dtype=float), where=ab > 0)
    mean_avg = float(np.nanmean(avg)) if len(avg) else 0.0
    # Convert AVG into counting impact so playing time is represented.
    df["AVG_surplus_H"] = (avg - mean_avg) * ab

    w = (
        zscore(df["R"]) +
        zscore(df["RBI"]) +
        zscore(df["HR"]) +
        zscore(df["SB"]) +
        zscore(df["AVG_surplus_H"])
    )
    return w

def initial_pitcher_weight(df: pd.DataFrame) -> pd.Series:
    """
    Rough first-pass weight for pitchers:
    counting stats + "runs prevented" (ERA) + "baserunners prevented" (WHIP),
    both scaled by IP to reflect volume.
    """
    df = df.copy()
    ip_sum = float(df["IP"].sum())
    mean_era = float(9.0 * df["ER"].sum() / ip_sum) if ip_sum > 0 else float(df["ERA"].mean())
    mean_whip = float((df["H"].sum() + df["BB"].sum()) / ip_sum) if ip_sum > 0 else float(df["WHIP"].mean())

    df["ERA_surplus_ER"] = (mean_era - df["ERA"]) * df["IP"] / 9.0
    df["WHIP_surplus"] = (mean_whip - df["WHIP"]) * df["IP"]

    w = (
        zscore(df["W"]) +
        zscore(df["K"]) +
        zscore(df["SV"]) +
        zscore(df["ERA_surplus_ER"]) +
        zscore(df["WHIP_surplus"])
    )
    return w


# ----------------------------
# Team stat calculations (default 5x5)
# ----------------------------

def team_avg(H: float, AB: float) -> float:
    return float(H / AB) if AB > 0 else 0.0

def team_obp(H: float, BB: float, HBP: float, AB: float, SF: float) -> float:
    den = AB + BB + HBP + SF
    return float((H + BB + HBP) / den) if den > 0 else 0.0

def team_ops(H: float, BB: float, HBP: float, AB: float, SF: float, b2: float, b3: float, HR: float) -> float:
    obp = team_obp(H, BB, HBP, AB, SF)
    slg = float((H + b2 + 2.0 * b3 + 3.0 * HR) / AB) if AB > 0 else 0.0
    return float(obp + slg)

def team_era(ER: float, IP: float) -> float:
    return float(9.0 * ER / IP) if IP > 0 else float("nan")

def team_whip(H: float, BB: float, IP: float) -> float:
    return float((H + BB) / IP) if IP > 0 else float("nan")


def common_replacement_pitcher_rates(
    all_pit_df: pd.DataFrame,
    assigned_pit_df: pd.DataFrame,
    n_rep: int,
) -> Dict[str, float]:
    """Per-inning replacement rates from the best available non-starter pitchers."""
    assigned_players = set(assigned_pit_df["Player"])
    rep = all_pit_df[~all_pit_df["Player"].isin(assigned_players)].copy()
    rep = rep.sort_values("weight", ascending=False).head(max(int(n_rep), 1))

    ip = float(rep["IP"].sum()) if not rep.empty else 0.0
    if ip <= 0:
        return {k: 0.0 for k in ["W", "QS", "K", "SV", "SVH", "ER", "H", "BB"]}

    return {
        "W": float(rep["W"].sum() / ip),
        "QS": float(rep["QS"].sum() / ip),
        "K": float(rep["K"].sum() / ip),
        "SV": float(rep["SV"].sum() / ip),
        "SVH": float(rep["SVH"].sum() / ip),
        "ER": float(rep["ER"].sum() / ip),
        "H": float(rep["H"].sum() / ip),
        "BB": float(rep["BB"].sum() / ip),
    }


def common_apply_pitching_bounds(
    totals: Dict[str, float],
    lg: CommonDynastyRotoSettings,
    rep_rates: Optional[Dict[str, float]],
) -> Dict[str, float]:
    """Apply optional IP cap/fill and IP-min qualification to common-mode pitching totals."""
    out = {k: float(totals.get(k, 0.0)) for k in PIT_COMPONENT_COLS}
    ip = float(out["IP"])

    if lg.ip_max is not None:
        ip_cap = float(lg.ip_max)

        # If over cap, scale all counting components down to cap.
        if ip > ip_cap and ip > 0:
            factor = ip_cap / ip
            for col in PIT_COMPONENT_COLS:
                out[col] = float(out[col]) * factor
            ip = ip_cap

        # If under cap, assume streamable replacement innings.
        if ip < ip_cap and rep_rates is not None:
            add = ip_cap - ip
            out["IP"] = ip_cap
            for col in ["W", "QS", "K", "SV", "SVH", "ER", "H", "BB"]:
                out[col] = float(out[col]) + add * float(rep_rates.get(col, 0.0))
            ip = ip_cap

    out["ERA"] = team_era(out["ER"], ip)
    out["WHIP"] = team_whip(out["H"], out["BB"], ip)

    # Optional IP minimum qualification rule (default OFF)
    if lg.ip_min and lg.ip_min > 0 and ip < lg.ip_min:
        out["ERA"] = 99.0
        out["WHIP"] = 5.0

    return out


# ----------------------------
# Monte Carlo SGP denominators
# ----------------------------

def _mean_adjacent_rank_gap(values: np.ndarray, *, ascending: bool) -> float:
    """Mean absolute adjacent difference after rank-order sorting.

    Returns 0.0 for degenerate inputs (<2 finite values), which can occur
    in single-team test leagues and avoids NumPy empty-slice warnings.
    """
    arr = np.asarray(values, dtype=float)
    arr = arr[np.isfinite(arr)]
    if arr.size < 2:
        return 0.0

    sorted_arr = np.sort(arr)
    if not ascending:
        sorted_arr = sorted_arr[::-1]
    return float(np.mean(np.abs(np.diff(sorted_arr))))


def simulate_sgp_hit(assigned_hit: pd.DataFrame, lg: CommonDynastyRotoSettings, rng: np.random.Generator) -> Dict[str, float]:
    """
    Estimates how many of each stat ~= 1 roto point (SGP denominator),
    by simulating random allocations of the "starter pool" to 12 teams.
    """
    per_team = lg.hitter_slots
    diffs = {c: [] for c in HIT_CATS}

    groups = {slot: assigned_hit[assigned_hit["AssignedSlot"] == slot] for slot in per_team.keys()}
    idx_ab = HIT_COMPONENT_COLS.index("AB")
    idx_h = HIT_COMPONENT_COLS.index("H")
    idx_r = HIT_COMPONENT_COLS.index("R")
    idx_hr = HIT_COMPONENT_COLS.index("HR")
    idx_rbi = HIT_COMPONENT_COLS.index("RBI")
    idx_sb = HIT_COMPONENT_COLS.index("SB")

    for _ in range(lg.sims_for_sgp):
        AB = np.zeros(lg.n_teams)
        H = np.zeros(lg.n_teams)
        R = np.zeros(lg.n_teams)
        HR = np.zeros(lg.n_teams)
        RBI = np.zeros(lg.n_teams)
        SB = np.zeros(lg.n_teams)

        for slot, cnt in per_team.items():
            df_slot = groups[slot]
            idx = rng.permutation(len(df_slot))
            arr = df_slot.iloc[idx][HIT_COMPONENT_COLS].to_numpy(dtype=float)
            arr = arr.reshape(lg.n_teams, cnt, len(HIT_COMPONENT_COLS))
            sums = arr.sum(axis=1)

            AB += sums[:, idx_ab]
            H += sums[:, idx_h]
            R += sums[:, idx_r]
            HR += sums[:, idx_hr]
            RBI += sums[:, idx_rbi]
            SB += sums[:, idx_sb]

        AVG = np.divide(H, AB, out=np.zeros_like(H), where=AB > 0)

        vals = {"R": R, "RBI": RBI, "HR": HR, "SB": SB, "AVG": AVG}
        for c in HIT_CATS:
            x = vals[c].astype(float)
            diffs[c].append(_mean_adjacent_rank_gap(x, ascending=False))

    return {c: (float(np.mean(diffs[c])) if diffs[c] else 0.0) for c in HIT_CATS}

def simulate_sgp_pit(
    assigned_pit: pd.DataFrame,
    lg: CommonDynastyRotoSettings,
    rng: np.random.Generator,
    rep_rates: Optional[Dict[str, float]] = None,
) -> Dict[str, float]:
    diffs = {c: [] for c in PIT_CATS}
    per_team = lg.pitcher_slots
    groups = {slot: assigned_pit[assigned_pit["AssignedSlot"] == slot] for slot in per_team.keys()}
    idx_ip = PIT_COMPONENT_COLS.index("IP")
    idx_w = PIT_COMPONENT_COLS.index("W")
    idx_k = PIT_COMPONENT_COLS.index("K")
    idx_sv = PIT_COMPONENT_COLS.index("SV")
    idx_er = PIT_COMPONENT_COLS.index("ER")
    idx_h = PIT_COMPONENT_COLS.index("H")
    idx_bb = PIT_COMPONENT_COLS.index("BB")

    for _ in range(lg.sims_for_sgp):
        IP = np.zeros(lg.n_teams)
        W = np.zeros(lg.n_teams)
        K = np.zeros(lg.n_teams)
        SV = np.zeros(lg.n_teams)
        ER = np.zeros(lg.n_teams)
        H = np.zeros(lg.n_teams)
        BB = np.zeros(lg.n_teams)

        for slot, cnt in per_team.items():
            df_slot = groups[slot]
            idx = rng.permutation(len(df_slot))
            arr = df_slot.iloc[idx][PIT_COMPONENT_COLS].to_numpy(dtype=float)
            arr = arr.reshape(lg.n_teams, cnt, len(PIT_COMPONENT_COLS))
            sums = arr.sum(axis=1)

            IP += sums[:, idx_ip]
            W += sums[:, idx_w]
            K += sums[:, idx_k]
            SV += sums[:, idx_sv]
            ER += sums[:, idx_er]
            H += sums[:, idx_h]
            BB += sums[:, idx_bb]

        vals = {c: [] for c in PIT_CATS}
        for t in range(lg.n_teams):
            bounded = common_apply_pitching_bounds(
                {
                    "IP": float(IP[t]),
                    "W": float(W[t]),
                    "K": float(K[t]),
                    "SV": float(SV[t]),
                    "ER": float(ER[t]),
                    "H": float(H[t]),
                    "BB": float(BB[t]),
                },
                lg,
                rep_rates,
            )
            vals["W"].append(float(bounded["W"]))
            vals["K"].append(float(bounded["K"]))
            vals["SV"].append(float(bounded["SV"]))
            vals["ERA"].append(float(bounded["ERA"]))
            vals["WHIP"].append(float(bounded["WHIP"]))

        for c in PIT_CATS:
            x = np.array(vals[c], dtype=float)
            diffs[c].append(_mean_adjacent_rank_gap(x, ascending=(c in {"ERA", "WHIP"})))

    return {c: (float(np.mean(diffs[c])) if diffs[c] else 0.0) for c in PIT_CATS}


# ----------------------------
# Year context + player year values
# ----------------------------

def compute_year_context(year: int, bat: pd.DataFrame, pit: pd.DataFrame, lg: CommonDynastyRotoSettings, rng_seed: Optional[int] = None) -> dict:
    bat_y = bat[bat["Year"] == year].copy()
    pit_y = pit[pit["Year"] == year].copy()

    # Clean numeric NaNs
    for c in HIT_COMPONENT_COLS:
        if c not in bat_y.columns:
            bat_y[c] = 0.0
        bat_y[c] = bat_y[c].fillna(0.0)
    for c in PIT_COMPONENT_COLS:
        if c not in pit_y.columns:
            pit_y[c] = 0.0
        pit_y[c] = pit_y[c].fillna(0.0)

    # Starter-pool candidates (must have playing time)
    bat_play = bat_y[bat_y["AB"] > 0].copy()
    pit_play = pit_y[pit_y["IP"] > 0].copy()

    if bat_play.empty:
        raise ValueError(
            f"Year {year}: no hitters with AB > 0 after filtering. Check Year values and AB projections."
        )
    if pit_play.empty:
        raise ValueError(
            f"Year {year}: no pitchers with IP > 0 after filtering. Check Year values and IP projections."
        )

    # Initial weights to define the league baseline pool/positional scarcity
    bat_play["weight"] = initial_hitter_weight(bat_play)
    pit_play["weight"] = initial_pitcher_weight(pit_play)

    league_hit_slots = expand_slot_counts(lg.hitter_slots, lg.n_teams)
    league_pit_slots = expand_slot_counts(lg.pitcher_slots, lg.n_teams)

    assigned_hit = assign_players_to_slots_with_vacancy_fill(
        bat_play,
        league_hit_slots,
        eligible_hit_slots,
        stat_cols=HIT_COMPONENT_COLS,
        year=year,
        side_label="hitter",
        weight_col="weight",
    )
    assigned_pit = assign_players_to_slots_with_vacancy_fill(
        pit_play,
        league_pit_slots,
        eligible_pit_slots,
        stat_cols=PIT_COMPONENT_COLS,
        year=year,
        side_label="pitcher",
        weight_col="weight",
    )

    baseline_hit = assigned_hit.groupby("AssignedSlot")[HIT_COMPONENT_COLS].mean()
    baseline_pit = assigned_pit.groupby("AssignedSlot")[PIT_COMPONENT_COLS].mean()

    # Baseline "average team" totals
    team_hit_slots = build_team_slot_template(lg.hitter_slots)
    team_pit_slots = build_team_slot_template(lg.pitcher_slots)

    base_hit_tot = baseline_hit.loc[team_hit_slots].sum()
    base_avg = team_avg(
        float(base_hit_tot["H"]),
        float(base_hit_tot["AB"]),
    )

    base_pit_tot = baseline_pit.loc[team_pit_slots].sum()
    rep_rates = common_replacement_pitcher_rates(
        pit_play,
        assigned_pit,
        n_rep=lg.replacement_pitchers_n,
    )
    base_pit_bounded = common_apply_pitching_bounds(
        {col: float(base_pit_tot[col]) for col in PIT_COMPONENT_COLS},
        lg,
        rep_rates,
    )

    # SGP denominators by simulation
    seed = year if rng_seed is None else int(rng_seed)
    rng_hit = np.random.default_rng(seed)
    rng_pit = np.random.default_rng(seed + 1)
    sgp_hit = simulate_sgp_hit(assigned_hit, lg, rng_hit)
    sgp_pit = simulate_sgp_pit(assigned_pit, lg, rng_pit, rep_rates=rep_rates)

    return {
        "year": year,
        "bat_y": bat_y,
        "pit_y": pit_y,
        "assigned_hit": assigned_hit,
        "assigned_pit": assigned_pit,
        "baseline_hit": baseline_hit,
        "baseline_pit": baseline_pit,
        "base_hit_tot": base_hit_tot,
        "base_avg": base_avg,
        "base_pit_tot": base_pit_tot,
        "base_pit_bounded": base_pit_bounded,
        "rep_rates": rep_rates,
        "sgp_hit": sgp_hit,
        "sgp_pit": sgp_pit,
    }

def compute_year_player_values(ctx: dict, lg: CommonDynastyRotoSettings) -> Tuple[pd.DataFrame, pd.DataFrame]:
    year = int(ctx["year"])
    bat_y = ctx["bat_y"]
    pit_y = ctx["pit_y"]

    baseline_hit = ctx["baseline_hit"]
    baseline_pit = ctx["baseline_pit"]
    base_hit_tot = ctx["base_hit_tot"]
    base_avg = float(ctx["base_avg"])

    base_pit_tot = ctx["base_pit_tot"]
    base_pit_bounded = dict(ctx["base_pit_bounded"])
    rep_rates = ctx.get("rep_rates")

    sgp_hit = ctx["sgp_hit"]
    sgp_pit = ctx["sgp_pit"]

    # --- Hitters: best eligible slot vs average starter at that slot ---
    hit_rows = []
    for row in bat_y.to_dict(orient="records"):
        pos_set = parse_hit_positions(row.get("Pos", ""))
        slots = eligible_hit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_hit.index:
                continue
            b = baseline_hit.loc[slot]

            new_tot = base_hit_tot.copy()
            for col in HIT_COMPONENT_COLS:
                new_tot[col] = new_tot[col] - b[col] + float(row.get(col, 0.0))

            new_avg = team_avg(
                float(new_tot["H"]),
                float(new_tot["AB"]),
            )

            delta = {
                "R": float(new_tot["R"] - base_hit_tot["R"]),
                "RBI": float(new_tot["RBI"] - base_hit_tot["RBI"]),
                "HR": float(new_tot["HR"] - base_hit_tot["HR"]),
                "SB": float(new_tot["SB"] - base_hit_tot["SB"]),
                "AVG": float(new_avg - base_avg),
            }

            val = 0.0
            for c in HIT_CATS:
                denom = float(sgp_hit[c])
                val += (delta[c] / denom) if denom else 0.0

            if val > best_val:
                best_val = val
                best_slot = slot

        hit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "H",
            "Team": row.get("Team", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    hit_vals = pd.DataFrame(hit_rows)

    # --- Pitchers: best eligible slot (usually just P) vs average starter at that slot ---
    pit_rows = []
    for row in pit_y.to_dict(orient="records"):
        pos_set = parse_pit_positions(row.get("Pos", ""))
        slots = eligible_pit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_pit.index:
                continue
            b = baseline_pit.loc[slot]

            new_tot = base_pit_tot.copy()
            for col in PIT_COMPONENT_COLS:
                new_tot[col] = new_tot[col] - b[col] + float(row.get(col, 0.0))

            new_tot_bounded = common_apply_pitching_bounds(
                {col: float(new_tot[col]) for col in PIT_COMPONENT_COLS},
                lg,
                rep_rates,
            )

            delta = {
                "W": float(new_tot_bounded["W"] - base_pit_bounded["W"]),
                "K": float(new_tot_bounded["K"] - base_pit_bounded["K"]),
                "SV": float(new_tot_bounded["SV"] - base_pit_bounded["SV"]),
                "ERA": float(base_pit_bounded["ERA"] - new_tot_bounded["ERA"]),       # lower is better
                "WHIP": float(base_pit_bounded["WHIP"] - new_tot_bounded["WHIP"]),    # lower is better
            }

            val = 0.0
            for c in PIT_CATS:
                denom = float(sgp_pit[c])
                val += (delta[c] / denom) if denom else 0.0

            if val > best_val:
                best_val = val
                best_slot = slot

        pit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "P",
            "Team": row.get("Team", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    pit_vals = pd.DataFrame(pit_rows)
    return hit_vals, pit_vals


def compute_replacement_baselines(
    ctx: dict,
    lg: CommonDynastyRotoSettings,
    rostered_players: Set[str],
    n_repl: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Build per-slot replacement baselines from the unrostered pool."""
    n_repl = int(n_repl or lg.n_teams)

    bat_y = ctx["bat_y"].copy()
    pit_y = ctx["pit_y"].copy()

    for c in HIT_COMPONENT_COLS:
        bat_y[c] = bat_y[c].fillna(0.0)
    for c in PIT_COMPONENT_COLS:
        pit_y[c] = pit_y[c].fillna(0.0)

    bat_y["weight"] = initial_hitter_weight(bat_y)
    pit_y["weight"] = initial_pitcher_weight(pit_y)

    fa_hit = bat_y[(~bat_y["Player"].isin(rostered_players)) & (bat_y["AB"] > 0)].copy()
    fa_pit = pit_y[(~pit_y["Player"].isin(rostered_players)) & (pit_y["IP"] > 0)].copy()

    fa_hit["elig"] = fa_hit["Pos"].apply(lambda p: eligible_hit_slots(parse_hit_positions(p)))
    fa_pit["elig"] = fa_pit["Pos"].apply(lambda p: eligible_pit_slots(parse_pit_positions(p)))

    baseline_hit_avg = ctx["baseline_hit"]
    baseline_pit_avg = ctx["baseline_pit"]

    repl_hit_rows: List[dict] = []
    for slot in baseline_hit_avg.index:
        cand = (
            fa_hit[fa_hit["elig"].apply(lambda s: slot in s)]
            .sort_values("weight", ascending=False)
            .head(n_repl)
        )
        repl = baseline_hit_avg.loc[slot] if len(cand) == 0 else cand[HIT_COMPONENT_COLS].mean()
        row = {c: float(repl.get(c, 0.0)) for c in HIT_COMPONENT_COLS}
        row["AssignedSlot"] = slot
        repl_hit_rows.append(row)

    repl_hit = pd.DataFrame(repl_hit_rows).set_index("AssignedSlot")

    repl_pit_rows: List[dict] = []
    for slot in baseline_pit_avg.index:
        cand = (
            fa_pit[fa_pit["elig"].apply(lambda s: slot in s)]
            .sort_values("weight", ascending=False)
            .head(n_repl)
        )
        repl = baseline_pit_avg.loc[slot] if len(cand) == 0 else cand[PIT_COMPONENT_COLS].mean()
        row = {c: float(repl.get(c, 0.0)) for c in PIT_COMPONENT_COLS}
        row["AssignedSlot"] = slot
        repl_pit_rows.append(row)

    repl_pit = pd.DataFrame(repl_pit_rows).set_index("AssignedSlot")
    return repl_hit, repl_pit


def compute_year_player_values_vs_replacement(
    ctx: dict,
    lg: CommonDynastyRotoSettings,
    repl_hit: pd.DataFrame,
    repl_pit: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """Compute per-year values as marginal roto points above replacement."""
    year = int(ctx["year"])
    bat_y = ctx["bat_y"]
    pit_y = ctx["pit_y"]

    baseline_hit_avg = ctx["baseline_hit"]
    baseline_pit_avg = ctx["baseline_pit"]
    base_hit_tot_avg = ctx["base_hit_tot"]
    base_pit_tot_avg = ctx["base_pit_tot"]
    rep_rates = ctx.get("rep_rates")

    sgp_hit = ctx["sgp_hit"]
    sgp_pit = ctx["sgp_pit"]

    hit_rows = []
    for row in bat_y.to_dict(orient="records"):
        pos_set = parse_hit_positions(row.get("Pos", ""))
        slots = eligible_hit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_hit_avg.index or slot not in repl_hit.index:
                continue

            b_avg = baseline_hit_avg.loc[slot]
            b_rep = repl_hit.loc[slot]

            base_tot = base_hit_tot_avg.copy()
            new_tot = base_hit_tot_avg.copy()
            for col in HIT_COMPONENT_COLS:
                base_tot[col] = base_tot[col] - b_avg[col] + b_rep[col]
                new_tot[col] = new_tot[col] - b_avg[col] + float(row.get(col, 0.0))

            base_avg = team_avg(
                float(base_tot["H"]),
                float(base_tot["AB"]),
            )
            new_avg = team_avg(
                float(new_tot["H"]),
                float(new_tot["AB"]),
            )

            delta = {
                "R": float(new_tot["R"] - base_tot["R"]),
                "RBI": float(new_tot["RBI"] - base_tot["RBI"]),
                "HR": float(new_tot["HR"] - base_tot["HR"]),
                "SB": float(new_tot["SB"] - base_tot["SB"]),
                "AVG": float(new_avg - base_avg),
            }

            val = 0.0
            for c in HIT_CATS:
                denom = float(sgp_hit[c])
                val += (delta[c] / denom) if denom else 0.0

            if val > best_val:
                best_val = val
                best_slot = slot

        hit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "H",
            "Team": row.get("Team", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    hit_vals = pd.DataFrame(hit_rows)

    pit_rows = []
    for row in pit_y.to_dict(orient="records"):
        pos_set = parse_pit_positions(row.get("Pos", ""))
        slots = eligible_pit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_pit_avg.index or slot not in repl_pit.index:
                continue

            b_avg = baseline_pit_avg.loc[slot]
            b_rep = repl_pit.loc[slot]

            base_raw = {c: float(base_pit_tot_avg[c]) for c in PIT_COMPONENT_COLS}
            new_raw = {c: float(base_pit_tot_avg[c]) for c in PIT_COMPONENT_COLS}
            for col in PIT_COMPONENT_COLS:
                base_raw[col] = base_raw[col] - float(b_avg[col]) + float(b_rep[col])
                new_raw[col] = new_raw[col] - float(b_avg[col]) + float(row.get(col, 0.0))

            base_bounded = common_apply_pitching_bounds(base_raw, lg, rep_rates)
            new_bounded = common_apply_pitching_bounds(new_raw, lg, rep_rates)

            delta = {
                "W": float(new_bounded["W"] - base_bounded["W"]),
                "K": float(new_bounded["K"] - base_bounded["K"]),
                "SV": float(new_bounded["SV"] - base_bounded["SV"]),
                "ERA": float(base_bounded["ERA"] - new_bounded["ERA"]),
                "WHIP": float(base_bounded["WHIP"] - new_bounded["WHIP"]),
            }

            val = 0.0
            for c in PIT_CATS:
                denom = float(sgp_pit[c])
                val += (delta[c] / denom) if denom else 0.0

            if val > best_val:
                best_val = val
                best_slot = slot

        pit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "P",
            "Team": row.get("Team", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    pit_vals = pd.DataFrame(pit_rows)
    return hit_vals, pit_vals


def combine_two_way(hit_vals: pd.DataFrame, pit_vals: pd.DataFrame, two_way: str) -> pd.DataFrame:
    merged = pd.merge(
        hit_vals[["Player", "Year", "YearValue", "BestSlot", "Team", "Age", "Pos"]],
        pit_vals[["Player", "Year", "YearValue", "BestSlot", "Team", "Age", "Pos"]],
        on=["Player", "Year"],
        how="outer",
        suffixes=("_hit", "_pit"),
    )

    out_vals = []
    out_slots = []

    for _, r in merged.iterrows():
        hv = r.get("YearValue_hit")
        pv = r.get("YearValue_pit")

        if pd.isna(hv) and pd.isna(pv):
            out_vals.append(np.nan)
            out_slots.append(None)
            continue
        if pd.isna(hv):
            out_vals.append(float(pv))
            out_slots.append(r.get("BestSlot_pit"))
            continue
        if pd.isna(pv):
            out_vals.append(float(hv))
            out_slots.append(r.get("BestSlot_hit"))
            continue

        hv = float(hv)
        pv = float(pv)

        if two_way == "sum":
            out_vals.append(hv + pv)
            out_slots.append(f"{r.get('BestSlot_hit')}+{r.get('BestSlot_pit')}")
        else:  # "max"
            if hv >= pv:
                out_vals.append(hv)
                out_slots.append(r.get("BestSlot_hit"))
            else:
                out_vals.append(pv)
                out_slots.append(r.get("BestSlot_pit"))

    merged["YearValue"] = out_vals
    merged["BestSlot"] = out_slots
    merged["Team"] = merged["Team_hit"].combine_first(merged["Team_pit"])
    merged["Pos"] = merged["Pos_hit"].combine_first(merged["Pos_pit"])
    merged["Age"] = merged["Age_hit"].combine_first(merged["Age_pit"])

    return merged[["Player", "Year", "YearValue", "BestSlot", "Team", "Pos", "Age"]]


# ----------------------------
# Dynasty aggregation + centering
# ----------------------------


# ----------------------------
# Dynasty aggregation utilities
# ----------------------------

def dynasty_keep_or_drop_value(values: List[float], years: List[int], discount: float) -> float:
    """Compute the optimal discounted value of owning a player with a drop option.

    At the start of each season you either:
      - **Keep** the player for that season (receiving that season's `values[i]`, which may be negative), or
      - **Drop** the player permanently and receive 0 from that season onward.

    Discounting is applied between seasons using `discount ** year_gap`, where
    `year_gap = years[i+1] - years[i]`.

    This implements the one-dimensional dynamic program:

        F[i] = max(0, values[i] + discount**(gap) * F[i+1])

    Returns the optimal value in "start-year" units (i.e., relative to `years[0]`).
    """
    if not years or not values:
        return 0.0
    if len(values) != len(years):
        raise ValueError("values and years must have the same length")
    if len(years) == 1:
        return float(max(values[0], 0.0))

    f_next = 0.0
    for i in range(len(years) - 1, -1, -1):
        v = float(values[i])
        if i == len(years) - 1:
            hold = v
        else:
            gap = int(years[i + 1]) - int(years[i])
            if gap < 0:
                raise ValueError("years must be increasing")
            hold = v + (discount ** gap) * f_next
        f_next = max(0.0, hold)

    return float(f_next)

def _infer_minor_eligibility_by_year(
    bat_df: pd.DataFrame,
    pit_df: pd.DataFrame,
    *,
    years: Optional[List[int]],
    hitter_usage_max: int,
    pitcher_usage_max: int,
    hitter_age_max: int,
    pitcher_age_max: int,
) -> pd.DataFrame:
    """Infer year-level minor eligibility from projected MLB usage and age.

    The rule is intentionally monotonic: once a player is inferred to lose minor
    eligibility, they cannot regain it in later years.
    """

    def _per_side(df: pd.DataFrame, usage_col: str) -> pd.DataFrame:
        if df.empty or "Player" not in df.columns or "Year" not in df.columns:
            return pd.DataFrame(columns=["Player", "Year", usage_col, "Age"])

        cols = ["Player", "Year", usage_col]
        if "Age" in df.columns:
            cols.append("Age")
        side = df[cols].copy()
        side["Year"] = pd.to_numeric(side["Year"], errors="coerce")
        side = side.dropna(subset=["Player", "Year"])
        if side.empty:
            return pd.DataFrame(columns=["Player", "Year", usage_col, "Age"])

        side["Year"] = side["Year"].astype(int)
        side[usage_col] = pd.to_numeric(side[usage_col], errors="coerce").fillna(0.0).clip(lower=0.0)
        if "Age" not in side.columns:
            side["Age"] = np.nan
        else:
            side["Age"] = pd.to_numeric(side["Age"], errors="coerce")

        return side.groupby(["Player", "Year"], as_index=False).agg({usage_col: "max", "Age": "min"})

    bat_year = _per_side(bat_df, "AB")
    pit_year = _per_side(pit_df, "IP")
    merged = bat_year.merge(pit_year, on=["Player", "Year"], how="outer", suffixes=("_hit", "_pit"))
    if merged.empty:
        return pd.DataFrame(columns=["Player", "Year", "minor_eligible"])

    years_set = {int(y) for y in years} if years else None
    if years_set is not None:
        merged = merged[merged["Year"].isin(years_set)].copy()
    if merged.empty:
        return pd.DataFrame(columns=["Player", "Year", "minor_eligible"])

    merged["AB"] = pd.to_numeric(merged["AB"], errors="coerce").fillna(0.0).clip(lower=0.0)
    merged["IP"] = pd.to_numeric(merged["IP"], errors="coerce").fillna(0.0).clip(lower=0.0)
    age_hit = pd.to_numeric(merged["Age_hit"], errors="coerce")
    age_pit = pd.to_numeric(merged["Age_pit"], errors="coerce")
    merged["Age"] = age_hit.combine_first(age_pit)
    merged = merged.sort_values(["Player", "Year"]).reset_index(drop=True)

    merged["cum_AB"] = merged.groupby("Player", sort=False)["AB"].cumsum()
    merged["cum_IP"] = merged.groupby("Player", sort=False)["IP"].cumsum()

    raw_minor = (
        ((merged["cum_AB"] > 0.0) & (merged["cum_AB"] <= float(hitter_usage_max)) & (merged["Age"] <= float(hitter_age_max)))
        | ((merged["cum_IP"] > 0.0) & (merged["cum_IP"] <= float(pitcher_usage_max)) & (merged["Age"] <= float(pitcher_age_max)))
    )
    merged["minor_eligible_raw"] = _fillna_bool(raw_minor.astype("boolean"), default=False)

    def _enforce_once_lost(series: pd.Series) -> pd.Series:
        had_eligibility = False
        lost_eligibility = False
        out: List[bool] = []
        for value in series.tolist():
            eligible_now = bool(value)
            if lost_eligibility:
                out.append(False)
                continue
            if had_eligibility and not eligible_now:
                lost_eligibility = True
                out.append(False)
                continue
            if eligible_now:
                had_eligibility = True
            out.append(eligible_now)
        return pd.Series(out, index=series.index, dtype=bool)

    merged["minor_eligible"] = (
        merged.groupby("Player", sort=False)["minor_eligible_raw"].apply(_enforce_once_lost).reset_index(level=0, drop=True)
    )
    return merged[["Player", "Year", "minor_eligible"]]


def infer_minor_eligible(bat: pd.DataFrame, pit: pd.DataFrame, lg: CommonDynastyRotoSettings, start_year: int) -> pd.DataFrame:
    """Best-effort start-year minor eligibility inference from projections."""
    inferred = _infer_minor_eligibility_by_year(
        bat,
        pit,
        years=[start_year],
        hitter_usage_max=lg.minor_ab_max,
        pitcher_usage_max=lg.minor_ip_max,
        hitter_age_max=lg.minor_age_max_hit,
        pitcher_age_max=lg.minor_age_max_pit,
    )
    out = inferred[inferred["Year"] == int(start_year)][["Player", "minor_eligible"]].copy()
    if out.empty:
        return pd.DataFrame(columns=["Player", "minor_eligible"])
    return out.groupby("Player", as_index=False)["minor_eligible"].max()


def _non_vacant_player_names(df: Optional[pd.DataFrame]) -> Set[str]:
    """Collect non-placeholder player names from an assignment table."""
    if df is None or df.empty or "Player" not in df.columns:
        return set()
    names = df["Player"].dropna().astype(str)
    return {name for name in names if name and not name.startswith("__VACANT_")}


def _players_with_playing_time(bat_df: pd.DataFrame, pit_df: pd.DataFrame, years: List[int]) -> Set[str]:
    """Return players with projected MLB playing time in the valuation window."""
    years_set = {int(y) for y in years}
    players: Set[str] = set()

    if {"Player", "Year", "AB"}.issubset(bat_df.columns):
        hitters = bat_df.loc[(bat_df["Year"].isin(years_set)) & (bat_df["AB"] > 0), "Player"]
        players.update(hitters.dropna().astype(str))

    if {"Player", "Year", "IP"}.issubset(pit_df.columns):
        pitchers = pit_df.loc[(pit_df["Year"].isin(years_set)) & (pit_df["IP"] > 0), "Player"]
        players.update(pitchers.dropna().astype(str))

    return players


def _select_mlb_roster_with_active_floor(
    stash_sorted: pd.DataFrame,
    *,
    excluded_players: Set[str],
    total_mlb_slots: int,
    active_floor_names: Set[str],
) -> pd.DataFrame:
    """Pick MLB rostered players while guaranteeing active-floor names when possible."""
    remaining = stash_sorted[~stash_sorted["Player"].isin(excluded_players)].copy()
    if total_mlb_slots <= 0 or remaining.empty:
        return remaining.iloc[0:0].copy()

    floor = remaining[remaining["Player"].isin(active_floor_names)].copy()
    floor = floor.sort_values("StashScore", ascending=False)
    if len(floor) > total_mlb_slots:
        floor = floor.head(total_mlb_slots).copy()

    floor_names = set(floor["Player"]) if not floor.empty else set()
    fill_needed = max(total_mlb_slots - len(floor), 0)
    if fill_needed == 0:
        return floor.reset_index(drop=True)

    fill = remaining[~remaining["Player"].isin(floor_names)].head(fill_needed).copy()
    return pd.concat([floor, fill], ignore_index=True)


def _estimate_bench_negative_penalty(start_ctx: dict, lg: object) -> float:
    """Estimate marginal active-slot opportunity cost for one bench stash slot.

    Returns a factor in [0, 1] used to scale negative year values for players
    that can be stashed on the bench instead of occupying an active lineup spot.
    The openness heuristic is derived from hitter usage when available.
    """
    bench_slots = int(getattr(lg, "bench_slots", 0) or 0)
    if bench_slots <= 0:
        return 1.0

    hitter_slots = getattr(lg, "hitter_slots", {}) or {}
    active_hit_slots_per_team = int(sum(max(int(v), 0) for v in hitter_slots.values()))
    if active_hit_slots_per_team <= 0:
        return 1.0

    default_open_fraction = 0.15
    open_fraction = default_open_fraction

    assigned_hit = start_ctx.get("assigned_hit")
    if isinstance(assigned_hit, pd.DataFrame) and not assigned_hit.empty and "G" in assigned_hit.columns:
        non_vacant = assigned_hit[~assigned_hit["Player"].astype(str).str.startswith("__VACANT_")].copy()
        if not non_vacant.empty:
            g_total = float(non_vacant["G"].fillna(0.0).clip(lower=0.0).sum())
            max_games = float(max(len(non_vacant) * 162, 1))
            modeled_open = (max_games - g_total) / max_games
            open_fraction = float(np.clip(modeled_open, 0.0, 1.0))

    # Opportunity cost of one stash slot:
    # 1) Estimate total open hitter slot-seasons across a team.
    # 2) Assume remaining bench slots can absorb those open starts first.
    # 3) Only uncovered open starts create a real stash penalty.
    open_slot_seasons = open_fraction * float(active_hit_slots_per_team)
    remaining_bench_slots = float(max(bench_slots - 1, 0))
    uncovered_open_slots = max(open_slot_seasons - remaining_bench_slots, 0.0)
    return float(np.clip(uncovered_open_slots, 0.0, 1.0))


def _bench_stash_round_penalty(
    round_number: int,
    *,
    bench_slots: int,
    min_penalty: float = BENCH_STASH_MIN_PENALTY,
    max_penalty: float = BENCH_STASH_MAX_PENALTY,
    gamma: float = BENCH_STASH_PENALTY_GAMMA,
) -> float:
    """Penalty factor for a stash round (1-based), clipped to [0, 1]."""
    total_bench_slots = int(max(bench_slots, 0))
    if total_bench_slots <= 0:
        return 1.0

    round_num = max(int(round_number), 1)
    if round_num > total_bench_slots:
        return 1.0

    lo = float(np.clip(min_penalty, 0.0, 1.0))
    hi = float(np.clip(max_penalty, lo, 1.0))
    shape = float(max(gamma, 1e-9))

    if total_bench_slots == 1:
        return hi

    x = float(round_num - 1) / float(total_bench_slots - 1)
    penalty = lo + (hi - lo) * (x ** shape)
    return float(np.clip(penalty, 0.0, 1.0))


def _build_bench_stash_penalty_map(
    stash_sorted: pd.DataFrame,
    *,
    bench_stash_players: Set[str],
    n_teams: int,
    bench_slots: int,
) -> Dict[str, float]:
    """Assign player-specific bench penalties by stash round across teams."""
    if stash_sorted.empty or not bench_stash_players:
        return {}

    team_count = int(max(n_teams, 1))
    penalty_map: Dict[str, float] = {}
    stash_rank = 0

    for player in stash_sorted.get("Player", pd.Series(dtype=object)).dropna().astype(str).tolist():
        if player not in bench_stash_players or player in penalty_map:
            continue
        round_number = 1 + (stash_rank // team_count)
        penalty_map[player] = _bench_stash_round_penalty(round_number, bench_slots=bench_slots)
        stash_rank += 1

    return penalty_map


def _apply_negative_value_stash_rules(
    value: float,
    *,
    can_minor_stash: bool,
    can_bench_stash: bool,
    bench_negative_penalty: float,
) -> float:
    """Apply stash rules to negative year values before keep/drop aggregation."""
    if value >= 0.0:
        return float(value)
    if can_minor_stash:
        return 0.0
    if can_bench_stash:
        return float(value) * float(np.clip(bench_negative_penalty, 0.0, 1.0))
    return float(value)



def _fillna_bool(series: pd.Series, default: bool = False) -> pd.Series:
    """
    Coerce a Series to boolean and fill missing values without relying on pandas'
    deprecated silent downcasting behavior (avoids FutureWarning on .fillna/.ffill/.bfill).
    """
    # Use pandas' nullable BooleanDtype to handle NA safely, then convert to plain bool.
    return series.astype("boolean").fillna(default).astype(bool)

def _normalize_minor_eligibility(series: pd.Series) -> pd.Series:
    def _coerce(value: object) -> Optional[bool]:
        if pd.isna(value):
            return None
        if isinstance(value, (bool, np.bool_)):
            return bool(value)
        if isinstance(value, (int, float, np.integer, np.floating)):
            return bool(value)
        if isinstance(value, str):
            cleaned = value.strip().lower()
            if cleaned in {"y", "yes", "true", "t", "1"}:
                return True
            if cleaned in {"n", "no", "false", "f", "0", ""}:
                return False
            coerced = pd.to_numeric(cleaned, errors="coerce")
            if not pd.isna(coerced):
                return bool(coerced)
        return None

    return series.apply(_coerce)


def minor_eligibility_by_year_from_input(
    bat: pd.DataFrame,
    pit: pd.DataFrame,
) -> Optional[pd.DataFrame]:
    """Parse explicit minor-eligibility flags from input at Player/Year granularity."""
    candidates = {"minor", "minor_eligible", "minors_eligible", "minor_eligibility", "minors_eligibility", "minoreligible"}

    def _extract(df: pd.DataFrame) -> Optional[pd.DataFrame]:
        if df.empty or "Player" not in df.columns or "Year" not in df.columns:
            return None

        col_map = {c: c.strip().lower().replace(" ", "_") for c in df.columns}
        matched = [c for c, norm in col_map.items() if norm in candidates or ("minor" in norm and "elig" in norm)]
        if not matched:
            return None

        col = matched[0]
        subset = df[["Player", "Year", col]].copy()
        subset["Year"] = pd.to_numeric(subset["Year"], errors="coerce")
        subset["minor_eligible"] = _normalize_minor_eligibility(subset[col])
        subset = subset.drop(columns=[col]).dropna(subset=["Player", "Year", "minor_eligible"])
        if subset.empty:
            return None

        subset["Year"] = subset["Year"].astype(int)
        subset["minor_score"] = subset["minor_eligible"].map({True: 2, False: 1}).astype(int)
        grouped = subset.groupby(["Player", "Year"], as_index=False)["minor_score"].max()
        grouped["minor_eligible"] = grouped["minor_score"] >= 2
        return grouped[["Player", "Year", "minor_eligible"]]

    parts = [part for part in (_extract(bat), _extract(pit)) if part is not None]
    if not parts:
        return None

    merged = pd.concat(parts, ignore_index=True)
    merged["minor_score"] = merged["minor_eligible"].map({True: 2, False: 1}).astype(int)
    merged = merged.groupby(["Player", "Year"], as_index=False)["minor_score"].max()
    merged["minor_eligible"] = merged["minor_score"] >= 2
    return merged[["Player", "Year", "minor_eligible"]]


def minor_eligibility_from_input(
    bat: pd.DataFrame,
    pit: pd.DataFrame,
    start_year: int,
) -> Optional[pd.DataFrame]:
    """Backward-compatible start-year view of explicit minor eligibility."""
    by_year = minor_eligibility_by_year_from_input(bat, pit)
    if by_year is None or by_year.empty:
        return None

    subset = by_year[by_year["Year"] == int(start_year)][["Player", "minor_eligible"]].copy()
    if subset.empty:
        return None
    return subset.groupby("Player", as_index=False)["minor_eligible"].max()


def _resolve_minor_eligibility_by_year(
    bat_df: pd.DataFrame,
    pit_df: pd.DataFrame,
    *,
    years: List[int],
    hitter_usage_max: int,
    pitcher_usage_max: int,
    hitter_age_max: int,
    pitcher_age_max: int,
) -> pd.DataFrame:
    """Build Player/Year minor eligibility using input flags first, inference fallback."""
    years_set = {int(y) for y in years}

    inferred = _infer_minor_eligibility_by_year(
        bat_df,
        pit_df,
        years=years,
        hitter_usage_max=hitter_usage_max,
        pitcher_usage_max=pitcher_usage_max,
        hitter_age_max=hitter_age_max,
        pitcher_age_max=pitcher_age_max,
    )
    explicit = minor_eligibility_by_year_from_input(bat_df, pit_df)

    if explicit is None or explicit.empty:
        out = inferred.copy()
    else:
        merged = inferred.merge(explicit, on=["Player", "Year"], how="outer", suffixes=("_infer", "_input"))
        merged["minor_eligible"] = merged["minor_eligible_input"].combine_first(merged["minor_eligible_infer"])
        out = merged[["Player", "Year", "minor_eligible"]].copy()

    if out.empty:
        return pd.DataFrame(columns=["Player", "Year", "minor_eligible"])

    out = out[out["Year"].isin(years_set)].copy()
    if out.empty:
        return pd.DataFrame(columns=["Player", "Year", "minor_eligible"])
    out["minor_eligible"] = _fillna_bool(out["minor_eligible"])
    return out.groupby(["Player", "Year"], as_index=False)["minor_eligible"].max()

def calculate_common_dynasty_values(
    excel_path: str,
    lg: CommonDynastyRotoSettings,
    start_year: Optional[int] = None,
    years: Optional[List[int]] = None,
    verbose: bool = True,
    return_details: bool = False,
    seed: int = 0,
    recent_projections: int = 3,
):
    """Compute common-mode dynasty values.

    If return_details=True, also returns (bat_detail, pit_detail) tables that:
      - collapse duplicate (Player, Year) rows by averaging the most-recent N projections
      - keep the original input columns in roughly the same order
      - attach YearValue/BestSlot (per side) and DynastyValue to each Player/Year row
    """
    bat_raw = normalize_input_schema(pd.read_excel(excel_path, sheet_name="Bat"), COMMON_COLUMN_ALIASES)
    pit_raw = normalize_input_schema(pd.read_excel(excel_path, sheet_name="Pitch"), COMMON_COLUMN_ALIASES)

    bat_input_cols = list(bat_raw.columns)
    pit_input_cols = list(pit_raw.columns)
    bat_date_col = _find_projection_date_col(bat_raw)
    pit_date_col = _find_projection_date_col(pit_raw)

    bat_raw, pit_raw = _add_player_identity_keys(bat_raw, pit_raw)
    identity_lookup = _build_player_identity_lookup(bat_raw, pit_raw)
    bat_raw["Player"] = bat_raw[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()
    pit_raw["Player"] = pit_raw[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()

    # Average *all numeric stat columns* (except derived rates and Age) so the
    # aggregated detail tabs reflect the true averaged projections.
    bat_stat_cols = numeric_stat_cols_for_recent_avg(
        bat_raw,
        group_cols=["Player", "Year"],
        exclude_cols={"Age"} | DERIVED_HIT_RATE_COLS,
    )
    pit_stat_cols = numeric_stat_cols_for_recent_avg(
        pit_raw,
        group_cols=["Player", "Year"],
        exclude_cols={"Age"} | DERIVED_PIT_RATE_COLS,
    )

    bat = average_recent_projections(bat_raw, bat_stat_cols, max_entries=recent_projections)
    pit = average_recent_projections(pit_raw, pit_stat_cols, max_entries=recent_projections)

    # Recompute rates after averaging components
    bat = recompute_common_rates_hit(bat)
    pit = recompute_common_rates_pit(pit)

    # Backfill optional hitter rate components when source files omit them.
    for missing_col in ("BB", "HBP", "SF", "2B", "3B"):
        if missing_col not in bat.columns:
            bat[missing_col] = 0.0

    # Required fields
    require_cols(bat, ["Player", "Year", "Team", "Age", "Pos"] + HIT_COMPONENT_COLS, "Bat")
    require_cols(pit, ["Player", "Year", "Team", "Age", "Pos"], "Pitch")
    require_cols(pit, ["IP", "W", "K", "ER", "H", "BB"], "Pitch")

    # Ensure SV exists (fallback from legacy combined save/hold columns when needed).
    if "SV" not in pit.columns:
        if {"SVH", "HLD"}.issubset(pit.columns):
            pit["SV"] = (pit["SVH"] - pit["HLD"]).clip(lower=0.0).fillna(0.0)
        elif "SVH" in pit.columns:
            pit["SV"] = pit["SVH"].fillna(0.0)
        else:
            pit["SV"] = 0.0
    pit["SV"] = pit["SV"].fillna(0.0)

    # Keep SVH available for exports/other modes.
    if "SVH" not in pit.columns:
        if {"SV", "HLD"}.issubset(pit.columns):
            pit["SVH"] = pit["SV"].fillna(0.0) + pit["HLD"].fillna(0.0)
        else:
            pit["SVH"] = pit["SV"].fillna(0.0)
    pit["SVH"] = pit["SVH"].fillna(0.0)

    # Ensure QS exists (fallback to QA3 when provided by source).
    if "QS" not in pit.columns:
        if "QA3" in pit.columns:
            pit["QS"] = pit["QA3"].fillna(0.0)
        else:
            pit["QS"] = 0.0
    pit["QS"] = pit["QS"].fillna(0.0)

    if start_year is None:
        start_year = int(min(bat["Year"].min(), pit["Year"].min()))

    if years is None:
        max_year = int(max(bat["Year"].max(), pit["Year"].max()))
        years = [y for y in range(start_year, start_year + lg.horizon_years) if y <= max_year]

    if not years:
        raise ValueError("No valuation years available after applying start year / horizon to projection file years.")

    # Projection metadata: how many projections were averaged (<= recent_projections) and the oldest date used
    proj_meta = projection_meta_for_start_year(bat, pit, start_year)

    years_set = {int(y) for y in years}
    if lg.minor_slots and lg.minor_slots > 0:
        elig_year_df = _resolve_minor_eligibility_by_year(
            bat,
            pit,
            years=years,
            hitter_usage_max=lg.minor_ab_max,
            pitcher_usage_max=lg.minor_ip_max,
            hitter_age_max=lg.minor_age_max_hit,
            pitcher_age_max=lg.minor_age_max_pit,
        )
        start_minor = elig_year_df[elig_year_df["Year"] == int(start_year)][["Player", "minor_eligible"]].copy()
        if start_minor.empty:
            elig_df = pd.DataFrame(columns=["Player", "minor_eligible"])
        else:
            elig_df = start_minor.groupby("Player", as_index=False)["minor_eligible"].max()
    else:
        elig_year_df = pd.DataFrame(columns=["Player", "Year", "minor_eligible"])
        elig_df = pd.DataFrame(columns=["Player", "minor_eligible"])

    active_per_team = sum(lg.hitter_slots.values()) + sum(lg.pitcher_slots.values())
    total_minor_slots = lg.n_teams * lg.minor_slots
    total_mlb_slots = lg.n_teams * (active_per_team + lg.bench_slots + lg.ir_slots)

    # PASS 1: average-starter values to estimate who is rostered in a deep league.
    year_contexts: Dict[int, dict] = {}
    year_tables_avg: List[pd.DataFrame] = []

    for y in years:
        if verbose:
            print(f"Year {y}: baseline + SGP + player values (avg-starter pass) ...")
        ctx = compute_year_context(y, bat, pit, lg, rng_seed=seed + y)
        year_contexts[y] = ctx
        hit_vals, pit_vals = compute_year_player_values(ctx, lg)
        combined = combine_two_way(hit_vals, pit_vals, two_way=lg.two_way)
        year_tables_avg.append(combined)

    start_ctx = year_contexts.get(start_year, {})
    active_floor_names = (
        _non_vacant_player_names(start_ctx.get("assigned_hit"))
        | _non_vacant_player_names(start_ctx.get("assigned_pit"))
    )

    all_year_avg = pd.concat(year_tables_avg, ignore_index=True)
    wide_avg = all_year_avg.pivot_table(index="Player", columns="Year", values="YearValue", aggfunc="max").reset_index()
    for y in years:
        if y not in wide_avg.columns:
            wide_avg[y] = 0.0

    minor_eligibility_by_year = (
        {
            (str(row.Player), int(row.Year)): bool(row.minor_eligible)
            for row in elig_year_df.itertuples(index=False)
            if int(row.Year) in years_set
        }
        if not elig_year_df.empty
        else {}
    )
    bench_stash_players = _players_with_playing_time(bat, pit, years)

    def _stash_row(row: pd.Series, bench_penalty_by_player: Dict[str, float]) -> float:
        player = str(row["Player"])
        can_bench_stash = bool(lg.bench_slots and lg.bench_slots > 0 and player in bench_stash_players)
        bench_penalty = float(bench_penalty_by_player.get(player, 1.0))
        vals: List[float] = []
        for y in years:
            v = row.get(y)
            if pd.isna(v):
                v = 0.0
            v = float(v)
            v = _apply_negative_value_stash_rules(
                v,
                can_minor_stash=bool(
                    lg.minor_slots
                    and lg.minor_slots > 0
                    and bool(minor_eligibility_by_year.get((player, int(y)), False))
                ),
                can_bench_stash=can_bench_stash,
                bench_negative_penalty=bench_penalty,
            )
            vals.append(v)
        return dynasty_keep_or_drop_value(vals, years, lg.discount)

    provisional_bench_penalty = {player: 0.0 for player in bench_stash_players}
    wide_avg["StashScore"] = wide_avg.apply(lambda row: _stash_row(row, provisional_bench_penalty), axis=1)
    provisional_stash_sorted = wide_avg[["Player", "StashScore"]].sort_values("StashScore", ascending=False).reset_index(drop=True)
    bench_penalty_by_player = _build_bench_stash_penalty_map(
        provisional_stash_sorted,
        bench_stash_players=bench_stash_players,
        n_teams=lg.n_teams,
        bench_slots=lg.bench_slots,
    )
    wide_avg["StashScore"] = wide_avg.apply(lambda row: _stash_row(row, bench_penalty_by_player), axis=1)
    stash = wide_avg[["Player", "StashScore"]].copy()
    stash = stash.merge(elig_df, on="Player", how="left")
    stash["minor_eligible"] = _fillna_bool(stash["minor_eligible"])

    stash_sorted = stash.sort_values("StashScore", ascending=False).reset_index(drop=True)
    minors_pool = stash_sorted[stash_sorted["minor_eligible"]]
    minors_sel = minors_pool.head(total_minor_slots)
    minor_names = set(minors_sel["Player"])

    remaining = stash_sorted[~stash_sorted["Player"].isin(minor_names)]
    extra_minor_needed = max(total_minor_slots - len(minors_sel), 0)
    extra_minors = remaining.head(extra_minor_needed)
    extra_minor_names = set(extra_minors["Player"])

    mlb_sel = _select_mlb_roster_with_active_floor(
        stash_sorted,
        excluded_players=minor_names | extra_minor_names,
        total_mlb_slots=total_mlb_slots,
        active_floor_names=active_floor_names,
    )
    rostered_names: Set[str] = set(mlb_sel["Player"]) | minor_names | extra_minor_names

    # PASS 2: replacement-level per-year values from the unrostered pool.
    # By default, replacement baselines are frozen from start_year.
    year_tables: List[pd.DataFrame] = []
    hit_year_tables: List[pd.DataFrame] = []
    pit_year_tables: List[pd.DataFrame] = []

    frozen_repl_hit: Optional[pd.DataFrame] = None
    frozen_repl_pit: Optional[pd.DataFrame] = None
    if lg.freeze_replacement_baselines:
        start_ctx_for_replacement = year_contexts.get(start_year)
        if start_ctx_for_replacement is None:
            raise ValueError(
                f"Start year {start_year} context is unavailable for replacement baseline calculation."
            )
        frozen_repl_hit, frozen_repl_pit = compute_replacement_baselines(
            start_ctx_for_replacement,
            lg,
            rostered_names,
            n_repl=lg.n_teams,
        )

    for y in years:
        if verbose:
            print(f"Year {y}: replacement baselines + player values (replacement pass) ...")
        ctx = year_contexts[y]
        if lg.freeze_replacement_baselines:
            # Reuse a fixed replacement baseline from the start year.
            repl_hit = frozen_repl_hit
            repl_pit = frozen_repl_pit
        else:
            repl_hit, repl_pit = compute_replacement_baselines(
                ctx,
                lg,
                rostered_names,
                n_repl=lg.n_teams,
            )
        if repl_hit is None or repl_pit is None:
            raise ValueError("Replacement baselines were not initialized.")
        hit_vals, pit_vals = compute_year_player_values_vs_replacement(ctx, lg, repl_hit, repl_pit)

        if not hit_vals.empty:
            hit_year_tables.append(hit_vals[["Player", "Year", "BestSlot", "YearValue"]].copy())
        if not pit_vals.empty:
            pit_year_tables.append(pit_vals[["Player", "Year", "BestSlot", "YearValue"]].copy())

        combined = combine_two_way(hit_vals, pit_vals, two_way=lg.two_way)
        year_tables.append(combined)

    all_year = pd.concat(year_tables, ignore_index=True) if year_tables else pd.DataFrame()

    # Wide format: one row per player with Value_YEAR columns
    wide = all_year.pivot_table(index="Player", columns="Year", values="YearValue", aggfunc="max").reset_index()
    wide.columns = ["Player"] + [f"Value_{int(c)}" for c in wide.columns[1:]]

    # Metadata from start year
    meta = (
        all_year[all_year["Year"] == start_year][["Player", "Team", "Pos", "Age"]]
        .drop_duplicates("Player")
    )
    out = meta.merge(wide, on="Player", how="right")

    # Attach projection metadata (based on the start-year averaged projections)
    out = out.merge(proj_meta, on="Player", how="left")
    out = out.merge(elig_df, on="Player", how="left")
    out["minor_eligible"] = _fillna_bool(out["minor_eligible"])

    # Raw dynasty value: optimal keep/drop value.
    #
    # Old behavior: sum of positive years only (i.e., negatives were always free to ignore).
    # New behavior:
    #   - If the player can be stashed in a minors slot (league has minors slots AND player is minors-eligible),
    #     we still treat negative years as 0 (you keep them in minors, so no "holding" penalty).
    #   - Otherwise, negative years *do* count as a cost if you keep the player, but you can always drop
    #     the player permanently for 0 (so truly droppable players won't go negative overall).
    raw_vals: List[float] = []
    for _, r in out.iterrows():
        player = str(r.get("Player") or "")
        can_bench_stash = bool(lg.bench_slots and lg.bench_slots > 0 and player in bench_stash_players)
        bench_penalty = float(bench_penalty_by_player.get(player, 1.0))

        vals: List[float] = []
        for y in years:
            v = r.get(f"Value_{y}")
            if pd.isna(v):
                v = 0.0
            v = float(v)
            v = _apply_negative_value_stash_rules(
                v,
                can_minor_stash=bool(
                    lg.minor_slots
                    and lg.minor_slots > 0
                    and bool(minor_eligibility_by_year.get((player, int(y)), False))
                ),
                can_bench_stash=can_bench_stash,
                bench_negative_penalty=bench_penalty,
            )
            vals.append(v)

        raw_vals.append(dynasty_keep_or_drop_value(vals, years, lg.discount))

    out["RawDynastyValue"] = raw_vals

    # Centering: replacement-level roster cutoff with minors reserved first.
    out_sorted = out.sort_values("RawDynastyValue", ascending=False).reset_index(drop=True)
    minors_pool = out_sorted[out_sorted["minor_eligible"]]
    minors_sel = minors_pool.head(total_minor_slots)
    minor_names = set(minors_sel["Player"])

    remaining = out_sorted[~out_sorted["Player"].isin(minor_names)]
    extra_minor_needed = max(total_minor_slots - len(minors_sel), 0)
    extra_minors = remaining.head(extra_minor_needed)
    extra_minor_names = set(extra_minors["Player"])

    remaining = remaining[~remaining["Player"].isin(extra_minor_names)]
    mlb_sel = remaining.head(total_mlb_slots)

    rostered = pd.concat([minors_sel, extra_minors, mlb_sel], ignore_index=True)
    baseline_value = float(rostered["RawDynastyValue"].iloc[-1]) if len(rostered) else 0.0

    out["DynastyValue"] = out["RawDynastyValue"] - baseline_value
    out["CenteringBaselineValue"] = baseline_value
    out["CenteringBaselineMean"] = baseline_value

    out = out.sort_values("DynastyValue", ascending=False).reset_index(drop=True)
    out = _attach_identity_columns_to_output(out, identity_lookup)

    if not return_details:
        return out

    # ----------------------------
    # Detail tabs (aggregated projections + value columns)
    # ----------------------------
    hit_year = pd.concat(hit_year_tables, ignore_index=True) if hit_year_tables else pd.DataFrame(columns=["Player", "Year", "BestSlot", "YearValue"])
    pit_year = pd.concat(pit_year_tables, ignore_index=True) if pit_year_tables else pd.DataFrame(columns=["Player", "Year", "BestSlot", "YearValue"])

    player_vals = out[[PLAYER_ENTITY_KEY_COL, "DynastyValue", "RawDynastyValue", "minor_eligible"]].copy()

    bat_detail = bat.merge(hit_year, on=["Player", "Year"], how="left")
    bat_detail = bat_detail.merge(
        player_vals,
        left_on="Player",
        right_on=PLAYER_ENTITY_KEY_COL,
        how="left",
    ).drop(columns=[PLAYER_ENTITY_KEY_COL], errors="ignore")

    pit_detail = pit.merge(pit_year, on=["Player", "Year"], how="left")
    pit_detail = pit_detail.merge(
        player_vals,
        left_on="Player",
        right_on=PLAYER_ENTITY_KEY_COL,
        how="left",
    ).drop(columns=[PLAYER_ENTITY_KEY_COL], errors="ignore")

    display_by_entity = (
        dict(zip(identity_lookup[PLAYER_ENTITY_KEY_COL], identity_lookup["Player"]))
        if not identity_lookup.empty
        else {}
    )
    if display_by_entity:
        bat_detail["Player"] = bat_detail["Player"].map(display_by_entity).fillna(bat_detail["Player"])
        pit_detail["Player"] = pit_detail["Player"].map(display_by_entity).fillna(pit_detail["Player"])

    extra = ["ProjectionsUsed", "OldestProjectionDate", "BestSlot", "YearValue", "DynastyValue", "RawDynastyValue", "minor_eligible"]
    bat_detail = reorder_detail_columns(bat_detail, bat_input_cols, add_after=bat_date_col, extra_cols=extra)
    pit_detail = reorder_detail_columns(pit_detail, pit_input_cols, add_after=pit_date_col, extra_cols=extra)

    return out, bat_detail, pit_detail

# ----------------------------
# Custom league version (renamed to avoid collisions)
# ----------------------------

"""
Dynasty roto player values for your league settings.

What this script does (high level):
1) Reads your Excel file (sheets: "Bat" and "Pitch")
2) Builds an "average team" baseline using a league-wide optimal slot assignment
   (positional scarcity baked in).
3) Estimates SGP denominators (how much of each stat ~= 1 roto point) by Monte Carlo
   simulation of 12-team leagues built from that starter pool.
4) Computes each player's per-year marginal roto points vs the average starter in their
   best eligible slot, with the pitching IP cap/min/max accounted for:
   - Pitching totals are capped at 1500 IP (stats scale down if over).
   - If under 1500 IP, the script fills the gap with "replacement innings" from the
     best available non-starters (so high-IP SP only help insofar as they replace
     worse innings; "too many SP" gets diminishing returns under the cap).
5) Produces a single DynastyValue per player by discounting future years and then
   centering so that ~0 is the replacement-level rostered cutoff in this league format.

Dependencies:
  pip install pandas numpy openpyxl scipy
"""


# ----------------------------
# Helpers: stat components
# ----------------------------

def league_hitter_components(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Total Bases: TB = H + 2B + 2*3B + 3*HR
    df["TB"] = df["H"] + df["2B"] + 2 * df["3B"] + 3 * df["HR"]

    # OBP numerator/denominator (standard OBP)
    df["OBP_num"] = df["H"] + df["BB"] + df["HBP"]
    df["OBP_den"] = df["AB"] + df["BB"] + df["HBP"] + df["SF"]

    return df

def league_ensure_pitch_cols(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()

    # Allow alternative source columns if needed
    if "SVH" not in df.columns:
        if "SV" in df.columns and "HLD" in df.columns:
            df["SVH"] = df["SV"].fillna(0) + df["HLD"].fillna(0)
        else:
            df["SVH"] = 0.0

    return df


# ----------------------------
# Core math: baseline, assignment, SGP
# ----------------------------

def league_zscore(s: pd.Series) -> pd.Series:
    s = s.astype(float)
    mu = s.mean()
    sd = s.std(ddof=0)
    if sd == 0 or np.isnan(sd):
        return (s - mu) * 0.0
    return (s - mu) / sd

def league_initial_hitter_weight(df: pd.DataFrame) -> pd.Series:
    """
    Rough first-pass weight used only to determine the starter pool and average slot baselines.
    """
    df = df.copy()
    mean_hit_rate = df["H"].sum() / df["AB"].sum() if df["AB"].sum() > 0 else 0.0
    mean_ops = df["OPS"].mean() if "OPS" in df.columns else 0.0

    df["H_surplus"] = df["H"] - mean_hit_rate * df["AB"]
    df["OPS_surplus"] = (df.get("OPS", 0.0) - mean_ops) * df["AB"]

    cols = ["R", "HR", "RBI", "SB", "H_surplus", "OPS_surplus"]
    zsum = 0.0
    for c in cols:
        zsum += league_zscore(df[c])
    return zsum

def league_initial_pitcher_weight(df: pd.DataFrame) -> pd.Series:
    """
    Rough first-pass weight used only to determine the starter pool and average slot baselines.
    """
    df = df.copy()
    ip_sum = df["IP"].sum()
    mean_era = (df["ER"].sum() * 9 / ip_sum) if ip_sum > 0 else df["ERA"].mean()
    mean_whip = ((df["H"].sum() + df["BB"].sum()) / ip_sum) if ip_sum > 0 else df["WHIP"].mean()

    # Convert ratios into "runs prevented" / "baserunners prevented" relative to mean
    df["ERA_surplus_ER"] = (mean_era - df["ERA"]) * df["IP"] / 9
    df["WHIP_surplus"] = (mean_whip - df["WHIP"]) * df["IP"]

    cols = ["W", "K", "SVH", "QA3", "ERA_surplus_ER", "WHIP_surplus"]
    zsum = 0.0
    for c in cols:
        zsum += league_zscore(df[c])
    return zsum

def league_team_avg_ops(hit_tot: pd.Series) -> Tuple[float, float]:
    ab = float(hit_tot["AB"])
    avg = float(hit_tot["H"] / ab) if ab > 0 else 0.0
    obp_den = float(hit_tot["OBP_den"])
    obp = float(hit_tot["OBP_num"] / obp_den) if obp_den > 0 else 0.0
    slg = float(hit_tot["TB"] / ab) if ab > 0 else 0.0
    ops = obp + slg
    return avg, ops

def league_replacement_pitcher_rates(all_pit_df: pd.DataFrame, assigned_pit_df: pd.DataFrame, n_rep: int = 100) -> Dict[str, float]:
    """
    Compute per-inning replacement rates from the best available non-starter pitchers.
    """
    assigned_players = set(assigned_pit_df["Player"])
    rep = all_pit_df[~all_pit_df["Player"].isin(assigned_players)].copy()
    rep = rep.sort_values("weight", ascending=False).head(n_rep)

    ip = rep["IP"].sum()
    if ip <= 0:
        return {k: 0.0 for k in ["W", "K", "SVH", "QA3", "ER", "H", "BB"]}

    return {
        "W": rep["W"].sum() / ip,
        "K": rep["K"].sum() / ip,
        "SVH": rep["SVH"].sum() / ip,
        "QA3": rep["QA3"].sum() / ip,
        "ER": rep["ER"].sum() / ip,
        "H": rep["H"].sum() / ip,
        "BB": rep["BB"].sum() / ip,
    }

def league_apply_ip_cap(t: Dict[str, float], ip_cap: float, rep_rates: Optional[Dict[str, float]]) -> Dict[str, float]:
    """
    Enforce the 1500 IP cap and fill missing innings with replacement to reach the cap.
    """
    out = dict(t)
    ip = float(out.get("IP", 0.0))

    # If over cap: scale everything down proportionally (exactly matches "stats stop accruing at 1500")
    if ip > ip_cap and ip > 0:
        f = ip_cap / ip
        for k in ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]:
            out[k] = float(out.get(k, 0.0)) * f
        ip = ip_cap

    # If under cap: fill with replacement innings
    if ip < ip_cap and rep_rates is not None:
        add = ip_cap - ip
        out["IP"] = ip_cap
        for k in ["W", "K", "SVH", "QA3", "ER", "H", "BB"]:
            out[k] = float(out.get(k, 0.0)) + add * float(rep_rates.get(k, 0.0))
        ip = ip_cap

    # Ratios on capped totals
    out["ERA"] = 9.0 * out["ER"] / ip if ip > 0 else np.nan
    out["WHIP"] = (out["H"] + out["BB"]) / ip if ip > 0 else np.nan
    return out

def league_simulate_sgp_hit(assigned_hit_df: pd.DataFrame, lg: LeagueSettings, rng: np.random.Generator) -> Dict[str, float]:
    """
    Monte Carlo estimate of the average adjacent gap between roto ranks ("stat per roto point").
    """
    # Group players by the slot they were assigned to in the league-wide optimal assignment
    groups = {slot: assigned_hit_df[assigned_hit_df["AssignedSlot"] == slot] for slot in assigned_hit_df["AssignedSlot"].unique()}
    per_team = lg.hitter_slots

    cats = ["R", "HR", "RBI", "SB", "AVG", "OPS"]
    diffs = {c: [] for c in cats}

    for _ in range(lg.sims_for_sgp):
        # Team totals for each simulation
        team_tot = [{col: 0.0 for col in LEAGUE_HIT_STAT_COLS} for _ in range(lg.n_teams)]

        for slot, df_slot in groups.items():
            cnt = per_team[slot]
            idx = rng.permutation(len(df_slot))
            arr = df_slot.iloc[idx][LEAGUE_HIT_STAT_COLS].to_numpy()
            arr = arr.reshape(lg.n_teams, cnt, len(LEAGUE_HIT_STAT_COLS))

            # Vector sums per team, then add
            for t in range(lg.n_teams):
                sums = arr[t].sum(axis=0)
                for k, col in enumerate(LEAGUE_HIT_STAT_COLS):
                    team_tot[t][col] += float(sums[k])

        # Compute category totals
        vals = {c: [] for c in cats}
        for t in range(lg.n_teams):
            tot = team_tot[t]
            avg, ops = league_team_avg_ops(pd.Series(tot))
            vals["R"].append(tot["R"])
            vals["HR"].append(tot["HR"])
            vals["RBI"].append(tot["RBI"])
            vals["SB"].append(tot["SB"])
            vals["AVG"].append(avg)
            vals["OPS"].append(ops)

        for c in cats:
            arr = np.array(vals[c], dtype=float)
            diffs[c].append(_mean_adjacent_rank_gap(arr, ascending=False))

    return {c: (float(np.mean(diffs[c])) if diffs[c] else 0.0) for c in cats}

def league_simulate_sgp_pit(assigned_pit_df: pd.DataFrame, lg: LeagueSettings, rep_rates: Dict[str, float], rng: np.random.Generator) -> Dict[str, float]:
    """
    Monte Carlo estimate of the average adjacent gap between roto ranks ("stat per roto point") for pitching.
    """
    groups = {slot: assigned_pit_df[assigned_pit_df["AssignedSlot"] == slot] for slot in assigned_pit_df["AssignedSlot"].unique()}
    per_team = lg.pitcher_slots

    cats = ["W", "K", "SVH", "QA3", "ERA", "WHIP"]
    diffs = {c: [] for c in cats}

    base_cols = ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]

    for _ in range(lg.sims_for_sgp):
        team_raw = [{col: 0.0 for col in base_cols} for _ in range(lg.n_teams)]

        for slot, df_slot in groups.items():
            cnt = per_team[slot]
            idx = rng.permutation(len(df_slot))
            arr = df_slot.iloc[idx][base_cols].to_numpy()
            arr = arr.reshape(lg.n_teams, cnt, len(base_cols))

            for t in range(lg.n_teams):
                sums = arr[t].sum(axis=0)
                for k, col in enumerate(base_cols):
                    team_raw[t][col] += float(sums[k])

        vals = {c: [] for c in cats}
        for t in range(lg.n_teams):
            capped = league_apply_ip_cap(team_raw[t], ip_cap=lg.ip_max, rep_rates=rep_rates)
            vals["W"].append(capped["W"])
            vals["K"].append(capped["K"])
            vals["SVH"].append(capped["SVH"])
            vals["QA3"].append(capped["QA3"])
            vals["ERA"].append(capped["ERA"])
            vals["WHIP"].append(capped["WHIP"])

        for c in cats:
            arr = np.array(vals[c], dtype=float)
            diffs[c].append(_mean_adjacent_rank_gap(arr, ascending=(c in {"ERA", "WHIP"})))

    return {c: (float(np.mean(diffs[c])) if diffs[c] else 0.0) for c in cats}


# ----------------------------
# Year context + player year-values
# ----------------------------

def league_sum_slots(baseline_df: pd.DataFrame, slot_list: List[str]) -> pd.Series:
    return baseline_df.loc[slot_list].sum()

def league_compute_year_context(year: int, bat_df: pd.DataFrame, pit_df: pd.DataFrame, lg: LeagueSettings, rng_seed: int) -> dict:
    bat_y = league_hitter_components(bat_df[bat_df["Year"] == year].copy())
    pit_y = league_ensure_pitch_cols(pit_df[pit_df["Year"] == year].copy())

    # Use only playing-time > 0 rows to build the "starter pool" baselines
    bat_play = bat_y[bat_y["AB"] > 0].copy()
    pit_play = pit_y[pit_y["IP"] > 0].copy()

    if bat_play.empty:
        raise ValueError(
            f"Year {year}: no hitters with AB > 0 after filtering. Check Year values and AB projections."
        )
    if pit_play.empty:
        raise ValueError(
            f"Year {year}: no pitchers with IP > 0 after filtering. Check Year values and IP projections."
        )

    bat_play["weight"] = league_initial_hitter_weight(bat_play)
    pit_play["weight"] = league_initial_pitcher_weight(pit_play)

    league_hit_slots = league_expand_slot_counts(lg.hitter_slots, lg.n_teams)
    league_pit_slots = league_expand_slot_counts(lg.pitcher_slots, lg.n_teams)

    assigned_hit = league_assign_players_to_slots(bat_play, league_hit_slots, league_eligible_hit_slots, weight_col="weight")
    assigned_pit = league_assign_players_to_slots(pit_play, league_pit_slots, league_eligible_pit_slots, weight_col="weight")

    baseline_hit = assigned_hit.groupby("AssignedSlot")[LEAGUE_HIT_STAT_COLS].mean()
    baseline_pit = assigned_pit.groupby("AssignedSlot")[["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]].mean()

    team_hit_slots = league_build_team_slot_template(lg.hitter_slots)
    team_pit_slots = league_build_team_slot_template(lg.pitcher_slots)

    base_hit_tot = league_sum_slots(baseline_hit, team_hit_slots)
    base_avg, base_ops = league_team_avg_ops(base_hit_tot)

    base_pit_raw = league_sum_slots(baseline_pit, team_pit_slots)
    base_pit_raw_dict = {k: float(base_pit_raw[k]) for k in ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]}

    rep_rates = league_replacement_pitcher_rates(pit_play.assign(weight=league_initial_pitcher_weight(pit_play)), assigned_pit, n_rep=lg.replacement_pitchers_n)

    rng_hit = np.random.default_rng(rng_seed)
    rng_pit = np.random.default_rng(rng_seed + 1)
    sgp_hit = league_simulate_sgp_hit(assigned_hit, lg, rng_hit)
    sgp_pit = league_simulate_sgp_pit(assigned_pit, lg, rep_rates, rng_pit)

    return {
        "year": year,
        "bat_y": bat_y,
        "pit_y": pit_y,
        "baseline_hit": baseline_hit,
        "baseline_pit": baseline_pit,
        "base_hit_tot": base_hit_tot,
        "base_avg": base_avg,
        "base_ops": base_ops,
        "base_pit_raw": base_pit_raw_dict,
        "rep_rates": rep_rates,
        "sgp_hit": sgp_hit,
        "sgp_pit": sgp_pit,
    }

def league_compute_year_player_values(ctx: dict, lg: LeagueSettings) -> Tuple[pd.DataFrame, pd.DataFrame]:
    year = int(ctx["year"])
    bat_y = ctx["bat_y"]
    pit_y = ctx["pit_y"]

    baseline_hit = ctx["baseline_hit"]
    baseline_pit = ctx["baseline_pit"]
    base_hit_tot = ctx["base_hit_tot"]
    base_avg = float(ctx["base_avg"])
    base_ops = float(ctx["base_ops"])

    base_pit_raw = dict(ctx["base_pit_raw"])
    rep_rates = ctx["rep_rates"]

    sgp_hit = ctx["sgp_hit"]
    sgp_pit = ctx["sgp_pit"]

    base_pit_capped = league_apply_ip_cap(base_pit_raw, ip_cap=lg.ip_max, rep_rates=rep_rates)

    # --- Hitters: best-slot marginal SGP vs average starter in that slot ---
    hit_rows = []
    for row in bat_y.to_dict(orient="records"):
        pos_set = league_parse_hit_positions(row.get("Pos", ""))
        slots = league_eligible_hit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_hit.index:
                continue

            b = baseline_hit.loc[slot]
            new_tot = base_hit_tot.copy()
            for col in LEAGUE_HIT_STAT_COLS:
                new_tot[col] = new_tot[col] - b[col] + float(row.get(col, 0.0))

            new_avg, new_ops = league_team_avg_ops(new_tot)

            delta_R = float(new_tot["R"] - base_hit_tot["R"])
            delta_HR = float(new_tot["HR"] - base_hit_tot["HR"])
            delta_RBI = float(new_tot["RBI"] - base_hit_tot["RBI"])
            delta_SB = float(new_tot["SB"] - base_hit_tot["SB"])
            delta_AVG = float(new_avg - base_avg)
            delta_OPS = float(new_ops - base_ops)

            val = (
                (delta_R / sgp_hit["R"] if sgp_hit["R"] else 0.0)
                + (delta_HR / sgp_hit["HR"] if sgp_hit["HR"] else 0.0)
                + (delta_RBI / sgp_hit["RBI"] if sgp_hit["RBI"] else 0.0)
                + (delta_SB / sgp_hit["SB"] if sgp_hit["SB"] else 0.0)
                + (delta_AVG / sgp_hit["AVG"] if sgp_hit["AVG"] else 0.0)
                + (delta_OPS / sgp_hit["OPS"] if sgp_hit["OPS"] else 0.0)
            )

            if val > best_val:
                best_val = val
                best_slot = slot

        hit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "H",
            "MLBTeam": row.get("MLBTeam", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    hit_vals = pd.DataFrame(hit_rows)

    # --- Pitchers: best-slot marginal SGP vs average starter in that slot, with IP cap ---
    pit_rows = []
    for row in pit_y.to_dict(orient="records"):
        pos_set = league_parse_pit_positions(row.get("Pos", ""))
        slots = league_eligible_pit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_pit.index:
                continue

            b = baseline_pit.loc[slot]
            new_raw = dict(base_pit_raw)
            for col in ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]:
                new_raw[col] = float(new_raw[col]) - float(b[col]) + float(row.get(col, 0.0))

            new_capped = league_apply_ip_cap(new_raw, ip_cap=lg.ip_max, rep_rates=rep_rates)

            delta_W = float(new_capped["W"] - base_pit_capped["W"])
            delta_K = float(new_capped["K"] - base_pit_capped["K"])
            delta_SVH = float(new_capped["SVH"] - base_pit_capped["SVH"])
            delta_QA3 = float(new_capped["QA3"] - base_pit_capped["QA3"])

            # Lower is better for ERA/WHIP => improvement = base - new
            delta_ERA = float(base_pit_capped["ERA"] - new_capped["ERA"])
            delta_WHIP = float(base_pit_capped["WHIP"] - new_capped["WHIP"])

            val = (
                (delta_W / sgp_pit["W"] if sgp_pit["W"] else 0.0)
                + (delta_K / sgp_pit["K"] if sgp_pit["K"] else 0.0)
                + (delta_SVH / sgp_pit["SVH"] if sgp_pit["SVH"] else 0.0)
                + (delta_QA3 / sgp_pit["QA3"] if sgp_pit["QA3"] else 0.0)
                + (delta_ERA / sgp_pit["ERA"] if sgp_pit["ERA"] else 0.0)
                + (delta_WHIP / sgp_pit["WHIP"] if sgp_pit["WHIP"] else 0.0)
            )

            if val > best_val:
                best_val = val
                best_slot = slot

        pit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "P",
            "MLBTeam": row.get("MLBTeam", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    pit_vals = pd.DataFrame(pit_rows)
    return hit_vals, pit_vals

def league_compute_replacement_baselines(
    ctx: dict,
    lg: LeagueSettings,
    rostered_players: Set[str],
    n_repl: Optional[int] = None,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Build per-slot replacement-level baselines from the *unrostered* player pool.

    We approximate "replacement at slot" as the mean stat line of the top `n_repl`
    free agents eligible at that slot (default: n_teams).
    """
    n_repl = int(n_repl or lg.n_teams)

    bat_y = ctx["bat_y"].copy()
    pit_y = ctx["pit_y"].copy()

    # Clean numeric NaNs
    for c in LEAGUE_HIT_STAT_COLS:
        if c in bat_y.columns:
            bat_y[c] = bat_y[c].fillna(0.0)
    for c in ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]:
        if c in pit_y.columns:
            pit_y[c] = pit_y[c].fillna(0.0)

    if "ERA" in pit_y.columns:
        pit_y["ERA"] = pit_y["ERA"].fillna(pit_y["ERA"].mean())
    if "WHIP" in pit_y.columns:
        pit_y["WHIP"] = pit_y["WHIP"].fillna(pit_y["WHIP"].mean())

    # Weights for ordering free agents (same rough weights used for starter-pool selection)
    bat_y["weight"] = league_initial_hitter_weight(bat_y)
    pit_y["weight"] = league_initial_pitcher_weight(pit_y)

    # Candidate free-agent pools (must have playing time to be meaningful replacements)
    fa_hit = bat_y[(~bat_y["Player"].isin(rostered_players)) & (bat_y["AB"] > 0)].copy()
    fa_pit = pit_y[(~pit_y["Player"].isin(rostered_players)) & (pit_y["IP"] > 0)].copy()

    fa_hit["elig"] = fa_hit["Pos"].apply(lambda p: league_eligible_hit_slots(league_parse_hit_positions(p)))
    fa_pit["elig"] = fa_pit["Pos"].apply(lambda p: league_eligible_pit_slots(league_parse_pit_positions(p)))

    # Hit replacement baselines per slot
    repl_hit_rows: List[dict] = []
    baseline_hit_avg = ctx["baseline_hit"]
    for slot in baseline_hit_avg.index:
        cand = (
            fa_hit[fa_hit["elig"].apply(lambda s: slot in s)]
            .sort_values("weight", ascending=False)
            .head(n_repl)
        )
        if len(cand) == 0:
            repl = baseline_hit_avg.loc[slot]
        else:
            repl = cand[LEAGUE_HIT_STAT_COLS].mean()

        row = {c: float(repl.get(c, 0.0)) for c in LEAGUE_HIT_STAT_COLS}
        row["AssignedSlot"] = slot
        repl_hit_rows.append(row)

    repl_hit = pd.DataFrame(repl_hit_rows).set_index("AssignedSlot")

    # Pitch replacement baselines per slot
    repl_pit_rows: List[dict] = []
    baseline_pit_avg = ctx["baseline_pit"]
    pit_cols = ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]
    for slot in baseline_pit_avg.index:
        cand = (
            fa_pit[fa_pit["elig"].apply(lambda s: slot in s)]
            .sort_values("weight", ascending=False)
            .head(n_repl)
        )
        if len(cand) == 0:
            repl = baseline_pit_avg.loc[slot]
        else:
            repl = cand[pit_cols].mean()

        row = {c: float(repl.get(c, 0.0)) for c in pit_cols}
        row["AssignedSlot"] = slot
        repl_pit_rows.append(row)

    repl_pit = pd.DataFrame(repl_pit_rows).set_index("AssignedSlot")

    return repl_hit, repl_pit


def league_compute_year_player_values_vs_replacement(
    ctx: dict,
    lg: LeagueSettings,
    repl_hit: pd.DataFrame,
    repl_pit: pd.DataFrame,
) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """
    Compute per-year player values as marginal roto points above *replacement level*
    (instead of above the average starter).

    Implementation detail:
    - We keep the *team context* as an average-starter roster for the other slots.
    - For each candidate slot, we compare "player in that slot" vs
      "replacement player in that slot".
    """
    year = int(ctx["year"])
    bat_y = ctx["bat_y"]
    pit_y = ctx["pit_y"]

    baseline_hit_avg = ctx["baseline_hit"]
    baseline_pit_avg = ctx["baseline_pit"]

    base_hit_tot_avg = ctx["base_hit_tot"]

    base_pit_raw_avg = dict(ctx["base_pit_raw"])
    rep_rates = ctx["rep_rates"]

    sgp_hit = ctx["sgp_hit"]
    sgp_pit = ctx["sgp_pit"]

    # --- Hitters ---
    hit_rows = []
    for row in bat_y.to_dict(orient="records"):
        pos_set = league_parse_hit_positions(row.get("Pos", ""))
        slots = league_eligible_hit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_hit_avg.index or slot not in repl_hit.index:
                continue

            b_avg = baseline_hit_avg.loc[slot]
            b_rep = repl_hit.loc[slot]

            # Base team but with replacement in this slot
            base_tot = base_hit_tot_avg.copy()
            for col in LEAGUE_HIT_STAT_COLS:
                base_tot[col] = base_tot[col] - b_avg[col] + b_rep[col]

            # New team with this player in this slot
            new_tot = base_hit_tot_avg.copy()
            for col in LEAGUE_HIT_STAT_COLS:
                new_tot[col] = new_tot[col] - b_avg[col] + float(row.get(col, 0.0))

            base_avg, base_ops = league_team_avg_ops(base_tot)
            new_avg, new_ops = league_team_avg_ops(new_tot)

            delta_R = float(new_tot["R"] - base_tot["R"])
            delta_HR = float(new_tot["HR"] - base_tot["HR"])
            delta_RBI = float(new_tot["RBI"] - base_tot["RBI"])
            delta_SB = float(new_tot["SB"] - base_tot["SB"])
            delta_AVG = float(new_avg - base_avg)
            delta_OPS = float(new_ops - base_ops)

            val = (
                (delta_R / sgp_hit["R"] if sgp_hit["R"] else 0.0)
                + (delta_HR / sgp_hit["HR"] if sgp_hit["HR"] else 0.0)
                + (delta_RBI / sgp_hit["RBI"] if sgp_hit["RBI"] else 0.0)
                + (delta_SB / sgp_hit["SB"] if sgp_hit["SB"] else 0.0)
                + (delta_AVG / sgp_hit["AVG"] if sgp_hit["AVG"] else 0.0)
                + (delta_OPS / sgp_hit["OPS"] if sgp_hit["OPS"] else 0.0)
            )

            if val > best_val:
                best_val = val
                best_slot = slot

        hit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "H",
            "MLBTeam": row.get("MLBTeam", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    hit_vals = pd.DataFrame(hit_rows)

    # --- Pitchers ---
    pit_rows = []
    pit_cols = ["IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"]

    for row in pit_y.to_dict(orient="records"):
        pos_set = league_parse_pit_positions(row.get("Pos", ""))
        slots = league_eligible_pit_slots(pos_set)
        if not slots:
            continue

        best_val = -1e18
        best_slot = None

        for slot in slots:
            if slot not in baseline_pit_avg.index or slot not in repl_pit.index:
                continue

            b_avg = baseline_pit_avg.loc[slot]
            b_rep = repl_pit.loc[slot]

            # Base team but with replacement in this slot
            base_raw = dict(base_pit_raw_avg)
            for col in pit_cols:
                base_raw[col] = float(base_raw.get(col, 0.0)) - float(b_avg.get(col, 0.0)) + float(b_rep.get(col, 0.0))

            # New team with this player in this slot
            new_raw = dict(base_pit_raw_avg)
            for col in pit_cols:
                new_raw[col] = float(new_raw.get(col, 0.0)) - float(b_avg.get(col, 0.0)) + float(row.get(col, 0.0))

            base_capped = league_apply_ip_cap(base_raw, ip_cap=lg.ip_max, rep_rates=rep_rates)
            new_capped = league_apply_ip_cap(new_raw, ip_cap=lg.ip_max, rep_rates=rep_rates)

            delta_W = float(new_capped["W"] - base_capped["W"])
            delta_K = float(new_capped["K"] - base_capped["K"])
            delta_SVH = float(new_capped["SVH"] - base_capped["SVH"])
            delta_QA3 = float(new_capped["QA3"] - base_capped["QA3"])

            # Lower is better for ERA/WHIP
            delta_ERA = float(base_capped["ERA"] - new_capped["ERA"])
            delta_WHIP = float(base_capped["WHIP"] - new_capped["WHIP"])

            val = (
                (delta_W / sgp_pit["W"] if sgp_pit["W"] else 0.0)
                + (delta_K / sgp_pit["K"] if sgp_pit["K"] else 0.0)
                + (delta_SVH / sgp_pit["SVH"] if sgp_pit["SVH"] else 0.0)
                + (delta_QA3 / sgp_pit["QA3"] if sgp_pit["QA3"] else 0.0)
                + (delta_ERA / sgp_pit["ERA"] if sgp_pit["ERA"] else 0.0)
                + (delta_WHIP / sgp_pit["WHIP"] if sgp_pit["WHIP"] else 0.0)
            )

            if val > best_val:
                best_val = val
                best_slot = slot

        pit_rows.append({
            "Player": row.get("Player"),
            "Year": year,
            "Type": "P",
            "MLBTeam": row.get("MLBTeam", np.nan),
            "Age": row.get("Age", np.nan),
            "Pos": row.get("Pos", np.nan),
            "BestSlot": best_slot,
            "YearValue": float(best_val),
        })

    pit_vals = pd.DataFrame(pit_rows)
    return hit_vals, pit_vals

def league_combine_hitter_pitcher_year(hit_vals: pd.DataFrame, pit_vals: pd.DataFrame, two_way: str) -> pd.DataFrame:
    merged = pd.merge(
        hit_vals[["Player", "Year", "YearValue", "BestSlot", "Pos", "MLBTeam", "Age"]],
        pit_vals[["Player", "Year", "YearValue", "BestSlot", "Pos", "MLBTeam", "Age"]],
        on=["Player", "Year"],
        how="outer",
        suffixes=("_hit", "_pit"),
    )

    combined_val = []
    combined_slot = []

    for _, r in merged.iterrows():
        hv = r.get("YearValue_hit")
        pv = r.get("YearValue_pit")

        if pd.isna(hv) and pd.isna(pv):
            combined_val.append(np.nan)
            combined_slot.append(None)
            continue

        if pd.isna(hv):
            combined_val.append(float(pv))
            combined_slot.append(r.get("BestSlot_pit"))
            continue

        if pd.isna(pv):
            combined_val.append(float(hv))
            combined_slot.append(r.get("BestSlot_hit"))
            continue

        hv = float(hv)
        pv = float(pv)

        if two_way == "sum":
            combined_val.append(hv + pv)
            combined_slot.append(f"{r.get('BestSlot_hit')}+{r.get('BestSlot_pit')}")
        else:
            if hv >= pv:
                combined_val.append(hv)
                combined_slot.append(r.get("BestSlot_hit"))
            else:
                combined_val.append(pv)
                combined_slot.append(r.get("BestSlot_pit"))

    merged["YearValue"] = combined_val
    merged["BestSlot"] = combined_slot

    merged["Pos"] = merged["Pos_hit"].combine_first(merged["Pos_pit"])
    merged["MLBTeam"] = merged["MLBTeam_hit"].combine_first(merged["MLBTeam_pit"])
    merged["Age"] = merged["Age_hit"].combine_first(merged["Age_pit"])

    return merged[["Player", "Year", "YearValue", "BestSlot", "Pos", "MLBTeam", "Age"]]


# ----------------------------
# XLSX output formatting
# ----------------------------

def _xlsx_apply_header_style(ws) -> None:
    """Apply a consistent header style to row 1."""
    max_col = ws.max_column
    if max_col < 1:
        return

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E79")  # dark blue
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9D9D9")
    header_border = Border(bottom=thin)

    ws.row_dimensions[1].height = 22
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border


def _xlsx_set_freeze_filters_and_view(ws, freeze_panes: str, add_autofilter: bool = False) -> None:
    """Freeze panes, optionally add a worksheet AutoFilter, and hide gridlines.

    IMPORTANT:
      Excel Tables include their own AutoFilter. If the worksheet also has a
      worksheet-level AutoFilter over the same range, Excel will often open the
      file in "repair" mode and remove the Table/filters.

    By default we therefore *clear* any worksheet-level filter and rely on the
    Table's built-in filter dropdowns. Pass add_autofilter=True only for sheets
    where you are NOT creating a Table.
    """
    ws.freeze_panes = freeze_panes
    ws.sheet_view.showGridLines = False

    if add_autofilter:
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row >= 1 and max_col >= 1:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            ws.auto_filter.ref = ref
    else:
        # Clear worksheet-level AutoFilter to avoid conflicts with Excel Tables.
        ws.auto_filter.ref = None


def _xlsx_add_table(ws, table_name: str, style_name: str = "TableStyleMedium9") -> None:
    """Wrap the used range in an Excel Table for striping + filter dropdowns.

    Excel Tables carry their own AutoFilter. If a worksheet-level AutoFilter is
    also present (ws.auto_filter.ref), Excel may "repair" the workbook on open
    and remove the table. To prevent that, we clear any sheet-level AutoFilter
    before adding the Table.
    """
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 2 or max_col < 1:
        return

    # Prevent Excel repair: don't mix worksheet AutoFilter + Table AutoFilter.
    ws.auto_filter.ref = None

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    tab = Table(displayName=table_name, ref=ref)
    tab.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)


def _xlsx_set_column_widths(
    ws,
    df: pd.DataFrame,
    overrides: Optional[Dict[str, float]] = None,
    sample_rows: int = 1000,
    min_width: float = 8.0,
    max_width: float = 45.0,
) -> None:
    """Best-effort "auto-fit" widths with sensible caps (fast + readable)."""
    if df is None or df.empty:
        return

    overrides = dict(overrides or {})

    # Common dynamic overrides
    for col in df.columns:
        if isinstance(col, str) and col.startswith("Value_"):
            overrides.setdefault(col, 10.0)

    for i, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(i)

        # Explicit override wins
        if col_name in overrides:
            ws.column_dimensions[letter].width = float(overrides[col_name])
            continue

        s = df[col_name]

        # Dates (including Python date objects often come through as object dtype)
        if str(col_name).lower().endswith("date"):
            ws.column_dimensions[letter].width = 14.0
            continue

        # Numeric: keep compact
        if pd.api.types.is_numeric_dtype(s):
            base = max(len(str(col_name)) + 2, 10)
            ws.column_dimensions[letter].width = float(min(max(base, min_width), 16.0))
            continue

        # Boolean: compact
        if pd.api.types.is_bool_dtype(s):
            ws.column_dimensions[letter].width = float(max(12.0, len(str(col_name)) + 2))
            continue

        # Text/object: use a sample to estimate
        sample = s.dropna().astype(str).head(sample_rows)
        max_len = int(sample.str.len().max()) if not sample.empty else 0
        width = min(max(max_len, len(str(col_name))) + 2, int(max_width))
        width = max(float(width), float(min_width))
        ws.column_dimensions[letter].width = float(width)


def _xlsx_apply_number_formats(ws, df: pd.DataFrame, formats_by_col: Dict[str, str]) -> None:
    """Apply number formats to entire columns (data rows only)."""
    if df is None or df.empty:
        return

    max_row = ws.max_row
    if max_row < 2:
        return

    cols = list(df.columns)
    for col_name, fmt in formats_by_col.items():
        if col_name not in cols:
            continue
        col_idx = cols.index(col_name) + 1
        for r in range(2, max_row + 1):
            ws.cell(row=r, column=col_idx).number_format = fmt


def _xlsx_add_value_color_scale(ws, df: pd.DataFrame, col_name: str) -> None:
    """Add a red-yellow-green color scale on a value column for readability."""
    if df is None or df.empty:
        return
    if col_name not in df.columns:
        return
    max_row = ws.max_row
    if max_row < 3:
        return

    col_idx = list(df.columns).index(col_name) + 1
    col_letter = get_column_letter(col_idx)
    cell_range = f"{col_letter}2:{col_letter}{max_row}"

    rule = ColorScaleRule(
        start_type="min",
        start_color="F8696B",  # red
        mid_type="percentile",
        mid_value=50,
        mid_color="FFEB84",    # yellow
        end_type="max",
        end_color="63BE7B",    # green
    )
    ws.conditional_formatting.add(cell_range, rule)


def _xlsx_format_player_values(ws, df: pd.DataFrame, table_name: str = "PlayerValuesTbl") -> None:
    """Formatting for the summary tab."""
    _xlsx_apply_header_style(ws)
    _xlsx_set_freeze_filters_and_view(ws, freeze_panes="B2")
    _xlsx_add_table(ws, table_name=table_name)

    overrides = {
        "Player": 24.0,
        "Team": 8.0,
        "MLBTeam": 8.0,
        "Pos": 10.0,
        "OldestProjectionDate": 14.0,
        "DynastyValue": 12.0,
        "RawDynastyValue": 14.0,
        "CenteringBaselineMean": 16.0,
    }
    _xlsx_set_column_widths(ws, df, overrides=overrides)

    formats = {
        "ProjectionsUsed": "0",
        "Age": "0",
        "OldestProjectionDate": "yyyy-mm-dd",
        "DynastyValue": "0.00",
        "RawDynastyValue": "0.00",
        "CenteringBaselineMean": "0.00",
    }
    for c in df.columns:
        if isinstance(c, str) and c.startswith("Value_"):
            formats[c] = "0.00"
    _xlsx_apply_number_formats(ws, df, formats)

    # Helpful visual cue: color-scale DynastyValue
    _xlsx_add_value_color_scale(ws, df, "DynastyValue")


def _xlsx_format_detail_sheet(
    ws,
    df: pd.DataFrame,
    *,
    table_name: str,
    is_pitch: bool,
) -> None:
    """Formatting for Bat_Aggregated / Pitch_Aggregated."""
    _xlsx_apply_header_style(ws)
    # Freeze Player + Year (first two columns) + header row
    _xlsx_set_freeze_filters_and_view(ws, freeze_panes="C2")
    _xlsx_add_table(ws, table_name=table_name)

    overrides = {
        "Player": 24.0,
        "Team": 8.0,
        "MLBTeam": 8.0,
        "Pos": 10.0,
        "BestSlot": 10.0,
        "OldestProjectionDate": 14.0,
        "DynastyValue": 12.0,
        "RawDynastyValue": 14.0,
        "YearValue": 10.0,
    }
    _xlsx_set_column_widths(ws, df, overrides=overrides)

    # Core formats
    formats: Dict[str, str] = {
        "Year": "0",
        "Age": "0",
        "ProjectionsUsed": "0",
        "OldestProjectionDate": "yyyy-mm-dd",
        "YearValue": "0.00",
        "DynastyValue": "0.00",
        "RawDynastyValue": "0.00",
        "AVG": "0.000",
        "OBP": "0.000",
        "SLG": "0.000",
        "OPS": "0.000",
        "ERA": "0.00",
        "WHIP": "0.00",
        "IP": "0.0",
    }

    # Apply only formats for columns that exist in this sheet.
    _xlsx_apply_number_formats(ws, df, formats)

    # Color-scale the most important value columns.
    _xlsx_add_value_color_scale(ws, df, "YearValue")
    _xlsx_add_value_color_scale(ws, df, "DynastyValue")


# ----------------------------
# Dynasty aggregation + centering
# ----------------------------

def league_infer_minor_eligible_start(bat_df: pd.DataFrame, pit_df: pd.DataFrame, lg: LeagueSettings, start_year: int) -> pd.DataFrame:
    """Best-effort start-year minor eligibility inference from projections."""
    inferred = _infer_minor_eligibility_by_year(
        bat_df,
        pit_df,
        years=[start_year],
        hitter_usage_max=lg.minor_hitters_career_ab_max,
        pitcher_usage_max=lg.minor_pitchers_career_ip_max,
        hitter_age_max=lg.infer_minor_age_max_hit,
        pitcher_age_max=lg.infer_minor_age_max_pit,
    )
    out = inferred[inferred["Year"] == int(start_year)][["Player", "minor_eligible"]].copy()
    if out.empty:
        return pd.DataFrame(columns=["Player", "minor_eligible"])
    return out.groupby("Player", as_index=False)["minor_eligible"].max()

def calculate_league_dynasty_values(
    excel_path: str,
    lg: LeagueSettings,
    start_year: Optional[int] = None,
    years: Optional[List[int]] = None,
    verbose: bool = True,
    return_details: bool = False,
    seed: int = 0,
    recent_projections: int = 3,
):
    """League-mode dynasty values (your custom categories/rules).

    If return_details=True, also returns (bat_detail, pit_detail) tables that:
      - collapse duplicate (Player, Year) rows by averaging the most-recent 3 projections
      - keep the original input columns in roughly the same order
      - attach YearValue/BestSlot (per side) and DynastyValue to each Player/Year row
    """
    if not HAVE_SCIPY:
        raise ImportError("scipy is required for league mode (linear_sum_assignment not available).")

    bat_raw = normalize_input_schema(pd.read_excel(excel_path, sheet_name="Bat"), LEAGUE_COLUMN_ALIASES)
    pit_raw = normalize_input_schema(pd.read_excel(excel_path, sheet_name="Pitch"), LEAGUE_COLUMN_ALIASES)

    bat_input_cols = list(bat_raw.columns)
    pit_input_cols = list(pit_raw.columns)
    bat_date_col = _find_projection_date_col(bat_raw)
    pit_date_col = _find_projection_date_col(pit_raw)

    bat_raw, pit_raw = _add_player_identity_keys(bat_raw, pit_raw)
    identity_lookup = _build_player_identity_lookup(bat_raw, pit_raw)
    bat_raw["Player"] = bat_raw[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()
    pit_raw["Player"] = pit_raw[PLAYER_ENTITY_KEY_COL].astype("string").fillna("").str.strip()

    # Average *all numeric stat columns* (except derived rates and Age) so the
    # aggregated detail tabs (and category stats like SVH/QA3) reflect the true averaged projections.
    bat_stat_cols = numeric_stat_cols_for_recent_avg(
        bat_raw,
        group_cols=["Player", "Year"],
        exclude_cols={"Age"} | DERIVED_HIT_RATE_COLS,
    )
    pit_stat_cols = numeric_stat_cols_for_recent_avg(
        pit_raw,
        group_cols=["Player", "Year"],
        exclude_cols={"Age"} | DERIVED_PIT_RATE_COLS,
    )

    bat_df = average_recent_projections(bat_raw, bat_stat_cols, max_entries=recent_projections)
    pit_df = average_recent_projections(pit_raw, pit_stat_cols, max_entries=recent_projections)

    bat_df = recompute_league_rates_hit(bat_df)
    pit_df = recompute_league_rates_pit(pit_df)
    pit_df = league_ensure_pitch_cols(pit_df)

    require_cols(
        bat_df,
        ["Player", "Year", "MLBTeam", "Age", "Pos", "AB", "H", "R", "HR", "RBI", "SB", "BB", "HBP", "SF", "2B", "3B"],
        "Bat",
    )
    require_cols(
        pit_df,
        ["Player", "Year", "MLBTeam", "Age", "Pos", "IP", "W", "K", "SVH", "QA3", "ER", "H", "BB"],
        "Pitch",
    )

    if years is None:
        if start_year is None:
            start_year = int(min(bat_df["Year"].min(), pit_df["Year"].min()))
        max_year = int(max(bat_df["Year"].max(), pit_df["Year"].max()))
        years = [y for y in range(start_year, start_year + lg.horizon_years) if y <= max_year]
    else:
        if start_year is None:
            start_year = int(min(years))
        max_year = int(max(bat_df["Year"].max(), pit_df["Year"].max()))
        years = [y for y in years if y <= max_year]

    if not years:
        raise ValueError("No valuation years available after applying start year / horizon to projection file years.")

    # Projection metadata: how many projections were averaged (<=3) and the oldest date used
    proj_meta = projection_meta_for_start_year(bat_df, pit_df, start_year)

    years_set = {int(y) for y in years}
    elig_year_df = _resolve_minor_eligibility_by_year(
        bat_df,
        pit_df,
        years=years,
        hitter_usage_max=lg.minor_hitters_career_ab_max,
        pitcher_usage_max=lg.minor_pitchers_career_ip_max,
        hitter_age_max=lg.infer_minor_age_max_hit,
        pitcher_age_max=lg.infer_minor_age_max_pit,
    )
    start_minor = elig_year_df[elig_year_df["Year"] == int(start_year)][["Player", "minor_eligible"]].copy()
    if start_minor.empty:
        elig_df = pd.DataFrame(columns=["Player", "minor_eligible"])
    else:
        elig_df = start_minor.groupby("Player", as_index=False)["minor_eligible"].max()

    # Roster depth (league-wide)
    active_per_team = sum(lg.hitter_slots.values()) + sum(lg.pitcher_slots.values())  # should be 23
    total_minor_slots = lg.n_teams * lg.minor_slots
    total_mlb_slots = lg.n_teams * (active_per_team + lg.bench_slots + lg.ir_slots)

    # ------------------------------------------------------------------
    # PASS 1: compute average-starter year values (for a "stash score" that
    #         approximates who is rostered in a deep dynasty league).
    # ------------------------------------------------------------------
    year_contexts: Dict[int, dict] = {}
    year_tables_avg: List[pd.DataFrame] = []

    for y in years:
        if verbose:
            print(f"Year {y}: building baseline + SGP + player values (avg-starter pass) ...")
        ctx = league_compute_year_context(y, bat_df, pit_df, lg, rng_seed=seed + y)
        year_contexts[y] = ctx

        hit_vals, pit_vals = league_compute_year_player_values(ctx, lg)  # vs average starter
        combined = league_combine_hitter_pitcher_year(hit_vals, pit_vals, two_way=lg.two_way)
        year_tables_avg.append(combined)

    all_year_avg = pd.concat(year_tables_avg, ignore_index=True)

    # Stash score: optimal keep/drop value on the avg-starter YearValue stream.
    # Minor-eligible players can be stashed in minors (negative years treated as 0)
    # when the league has minors slots.
    minor_eligibility_by_year = (
        {
            (str(row.Player), int(row.Year)): bool(row.minor_eligible)
            for row in elig_year_df.itertuples(index=False)
            if int(row.Year) in years_set
        }
        if not elig_year_df.empty
        else {}
    )
    bench_stash_players = _players_with_playing_time(bat_df, pit_df, years)

    wide_avg = all_year_avg.pivot_table(index="Player", columns="Year", values="YearValue", aggfunc="max").reset_index()

    # Ensure every horizon year exists as a column (missing years => 0 value)
    for y in years:
        if y not in wide_avg.columns:
            wide_avg[y] = 0.0

    def _stash_row(row: pd.Series, bench_penalty_by_player: Dict[str, float]) -> float:
        player = str(row["Player"])
        can_bench_stash = bool(lg.bench_slots and lg.bench_slots > 0 and player in bench_stash_players)
        bench_penalty = float(bench_penalty_by_player.get(player, 1.0))

        vals: List[float] = []
        for y in years:
            v = row.get(y)
            if pd.isna(v):
                v = 0.0
            v = float(v)
            v = _apply_negative_value_stash_rules(
                v,
                can_minor_stash=bool(
                    lg.minor_slots
                    and lg.minor_slots > 0
                    and bool(minor_eligibility_by_year.get((player, int(y)), False))
                ),
                can_bench_stash=can_bench_stash,
                bench_negative_penalty=bench_penalty,
            )
            vals.append(v)

        return dynasty_keep_or_drop_value(vals, years, lg.discount)

    provisional_bench_penalty = {player: 0.0 for player in bench_stash_players}
    wide_avg["StashScore"] = wide_avg.apply(lambda row: _stash_row(row, provisional_bench_penalty), axis=1)
    provisional_stash_sorted = wide_avg[["Player", "StashScore"]].sort_values("StashScore", ascending=False).reset_index(drop=True)
    bench_penalty_by_player = _build_bench_stash_penalty_map(
        provisional_stash_sorted,
        bench_stash_players=bench_stash_players,
        n_teams=lg.n_teams,
        bench_slots=lg.bench_slots,
    )
    wide_avg["StashScore"] = wide_avg.apply(lambda row: _stash_row(row, bench_penalty_by_player), axis=1)
    stash = wide_avg[["Player", "StashScore"]].copy()

    stash = stash.merge(elig_df, on="Player", how="left")
    stash["minor_eligible"] = _fillna_bool(stash["minor_eligible"])

    # Determine rostered set (minors reserved first, then the rest)
    stash_sorted = stash.sort_values("StashScore", ascending=False).reset_index(drop=True)

    minors_pool = stash_sorted[stash_sorted["minor_eligible"]]
    minors_sel = minors_pool.head(total_minor_slots)
    minor_names = set(minors_sel["Player"])

    remaining = stash_sorted[~stash_sorted["Player"].isin(minor_names)]
    extra_minor_needed = max(total_minor_slots - len(minors_sel), 0)
    extra_minors = remaining.head(extra_minor_needed)
    extra_minor_names = set(extra_minors["Player"])

    remaining = remaining[~remaining["Player"].isin(extra_minor_names)]
    mlb_sel = remaining.head(total_mlb_slots)

    rostered_names: Set[str] = set(mlb_sel["Player"]) | minor_names | extra_minor_names

    # ------------------------------------------------------------------
    # PASS 2: compute per-year values vs *replacement* (from unrostered pool).
    # By default, replacement baselines are frozen from start_year.
    # ------------------------------------------------------------------
    year_tables: List[pd.DataFrame] = []
    hit_year_tables: List[pd.DataFrame] = []
    pit_year_tables: List[pd.DataFrame] = []

    frozen_repl_hit: Optional[pd.DataFrame] = None
    frozen_repl_pit: Optional[pd.DataFrame] = None
    if lg.freeze_replacement_baselines:
        start_ctx_for_replacement = year_contexts.get(start_year)
        if start_ctx_for_replacement is None:
            raise ValueError(
                f"Start year {start_year} context is unavailable for replacement baseline calculation."
            )
        frozen_repl_hit, frozen_repl_pit = league_compute_replacement_baselines(
            start_ctx_for_replacement,
            lg,
            rostered_names,
            n_repl=lg.n_teams,
        )

    for y in years:
        if verbose:
            print(f"Year {y}: computing replacement baselines + player values (replacement pass) ...")
        ctx = year_contexts[y]
        if lg.freeze_replacement_baselines:
            # Reuse a fixed replacement baseline from the start year.
            repl_hit = frozen_repl_hit
            repl_pit = frozen_repl_pit
        else:
            repl_hit, repl_pit = league_compute_replacement_baselines(
                ctx,
                lg,
                rostered_names,
                n_repl=lg.n_teams,
            )
        if repl_hit is None or repl_pit is None:
            raise ValueError("Replacement baselines were not initialized.")
        hit_vals, pit_vals = league_compute_year_player_values_vs_replacement(ctx, lg, repl_hit, repl_pit)

        # Store side-specific year values for the detail tabs
        if not hit_vals.empty:
            hit_year_tables.append(hit_vals[["Player", "Year", "BestSlot", "YearValue"]].copy())
        if not pit_vals.empty:
            pit_year_tables.append(pit_vals[["Player", "Year", "BestSlot", "YearValue"]].copy())

        combined = league_combine_hitter_pitcher_year(hit_vals, pit_vals, two_way=lg.two_way)
        year_tables.append(combined)

    all_year_vals = pd.concat(year_tables, ignore_index=True)

    # Wide table (one row per player) with Value_YEAR columns
    wide = all_year_vals.pivot_table(index="Player", columns="Year", values="YearValue", aggfunc="max").reset_index()
    wide.columns = ["Player"] + [f"Value_{int(c)}" for c in wide.columns[1:]]

    # Metadata from start year
    meta = (
        all_year_vals[all_year_vals["Year"] == start_year][["Player", "MLBTeam", "Pos", "Age"]]
        .drop_duplicates("Player")
    )

    out = meta.merge(wide, on="Player", how="right")

    # Attach projection metadata (based on the start-year averaged projections)
    out = out.merge(proj_meta, on="Player", how="left")

    # Raw dynasty value: optimal keep/drop value.
    #
    # - If the player can be stashed in a minors slot (league has minors slots AND player is minors-eligible),
    #   negative years are treated as 0 (no holding penalty while stashed).
    # - Otherwise, negative years *do* count as a cost if you keep the player, but you can drop the player
    #   permanently for 0 at any year boundary.
    raw_vals: List[float] = []
    for _, r in out.iterrows():
        player = str(r.get("Player") or "")
        can_bench_stash = bool(lg.bench_slots and lg.bench_slots > 0 and player in bench_stash_players)
        bench_penalty = float(bench_penalty_by_player.get(player, 1.0))

        vals: List[float] = []
        for y in years:
            col = f"Value_{y}"
            v = r.get(col)
            if pd.isna(v):
                v = 0.0
            v = float(v)
            v = _apply_negative_value_stash_rules(
                v,
                can_minor_stash=bool(
                    lg.minor_slots
                    and lg.minor_slots > 0
                    and bool(minor_eligibility_by_year.get((player, int(y)), False))
                ),
                can_bench_stash=can_bench_stash,
                bench_negative_penalty=bench_penalty,
            )
            vals.append(v)

        raw_vals.append(dynasty_keep_or_drop_value(vals, years, lg.discount))

    out["RawDynastyValue"] = raw_vals

    # Attach minor eligibility (for centering + output)
    out = out.merge(elig_df, on="Player", how="left")
    out["minor_eligible"] = _fillna_bool(out["minor_eligible"])

    # Center so replacement-level rostered cutoff ~= 0 (active + bench + minors + IR)
    out_sorted = out.sort_values("RawDynastyValue", ascending=False).reset_index(drop=True)

    minors_pool = out_sorted[out_sorted["minor_eligible"]]
    minors_sel = minors_pool.head(total_minor_slots)
    minor_names = set(minors_sel["Player"])

    remaining = out_sorted[~out_sorted["Player"].isin(minor_names)]
    extra_minor_needed = max(total_minor_slots - len(minors_sel), 0)
    extra_minors = remaining.head(extra_minor_needed)
    extra_minor_names = set(extra_minors["Player"])

    remaining = remaining[~remaining["Player"].isin(extra_minor_names)]
    mlb_sel = remaining.head(total_mlb_slots)

    rostered = pd.concat([minors_sel, extra_minors, mlb_sel], ignore_index=True)
    baseline_value = float(rostered["RawDynastyValue"].iloc[-1]) if len(rostered) else 0.0

    out["DynastyValue"] = out["RawDynastyValue"] - baseline_value
    out["CenteringBaselineValue"] = baseline_value
    out["CenteringBaselineMean"] = baseline_value

    out = out.sort_values("DynastyValue", ascending=False).reset_index(drop=True)
    out = _attach_identity_columns_to_output(out, identity_lookup)

    if not return_details:
        return out

    # ----------------------------
    # Detail tabs (aggregated projections + value columns)
    # ----------------------------
    hit_year = pd.concat(hit_year_tables, ignore_index=True) if hit_year_tables else pd.DataFrame(columns=["Player", "Year", "BestSlot", "YearValue"])
    pit_year = pd.concat(pit_year_tables, ignore_index=True) if pit_year_tables else pd.DataFrame(columns=["Player", "Year", "BestSlot", "YearValue"])

    player_vals = out[[PLAYER_ENTITY_KEY_COL, "DynastyValue", "RawDynastyValue", "minor_eligible"]].copy()

    bat_detail = bat_df.merge(hit_year, on=["Player", "Year"], how="left")
    bat_detail = bat_detail.merge(
        player_vals,
        left_on="Player",
        right_on=PLAYER_ENTITY_KEY_COL,
        how="left",
    ).drop(columns=[PLAYER_ENTITY_KEY_COL], errors="ignore")

    pit_detail = pit_df.merge(pit_year, on=["Player", "Year"], how="left")
    pit_detail = pit_detail.merge(
        player_vals,
        left_on="Player",
        right_on=PLAYER_ENTITY_KEY_COL,
        how="left",
    ).drop(columns=[PLAYER_ENTITY_KEY_COL], errors="ignore")

    display_by_entity = (
        dict(zip(identity_lookup[PLAYER_ENTITY_KEY_COL], identity_lookup["Player"]))
        if not identity_lookup.empty
        else {}
    )
    if display_by_entity:
        bat_detail["Player"] = bat_detail["Player"].map(display_by_entity).fillna(bat_detail["Player"])
        pit_detail["Player"] = pit_detail["Player"].map(display_by_entity).fillna(pit_detail["Player"])

    extra = ["ProjectionsUsed", "OldestProjectionDate", "BestSlot", "YearValue", "DynastyValue", "RawDynastyValue", "minor_eligible"]
    bat_detail = reorder_detail_columns(bat_detail, bat_input_cols, add_after=bat_date_col, extra_cols=extra)
    pit_detail = reorder_detail_columns(pit_detail, pit_input_cols, add_after=pit_date_col, extra_cols=extra)

    return out, bat_detail, pit_detail

# ----------------------------
# CLI (subcommands)
# ----------------------------

def main() -> None:
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest="mode", required=True)

    common = sub.add_parser("common", help="Run the common 5x5 dynasty roto valuation.")
    common.add_argument(
        "--input",
        default="Dynasty Baseball Projections.xlsx",
        help="Excel file with Bat and Pitch sheets (default: Dynasty Baseball Projections.xlsx).",
    )
    common.add_argument("--start-year", type=int, default=None, help="First valuation year (default: min Year in file).")
    common.add_argument("--teams", type=positive_int_arg, default=12)
    common.add_argument("--sims", type=positive_int_arg, default=200, help="Monte Carlo sims for SGP denominators.")
    common.add_argument("--horizon", type=positive_int_arg, default=10, help="Dynasty horizon years.")
    common.add_argument("--discount", type=discount_arg, default=0.94, help="Annual discount factor in (0, 1].")
    common.add_argument("--seed", type=int, default=0, help="Global random seed offset for deterministic simulations.")
    common.add_argument("--bench", type=non_negative_int_arg, default=6)
    common.add_argument("--minors", type=non_negative_int_arg, default=0)
    common.add_argument("--ir", type=non_negative_int_arg, default=0)
    common.add_argument("--ip-min", type=non_negative_float_arg, default=0.0, help="Optional IP minimum to qualify for ERA/WHIP (default 0).")
    common.add_argument(
        "--ip-max",
        type=optional_non_negative_float_arg,
        default=None,
        help="Optional IP maximum/cap (default none). Accepts numeric values or 'none'.",
    )
    common.add_argument(
        "--dynamic-replacement-baselines",
        action="store_true",
        help="Recompute replacement baselines for each valuation year (legacy behavior).",
    )
    common.add_argument("--out-prefix", default="common_player_values", help="Output prefix for CSV/XLSX.")
    common.add_argument("--recent-projections", type=positive_int_arg, default=3, help="Number of most recent projections to average per player/year.")

    league = sub.add_parser("league", help="Run the custom league valuation from the original my-league script.")
    league.add_argument(
        "--input",
        default="Dynasty Baseball Projections.xlsx",
        help="Excel file with Bat and Pitch sheets (default: Dynasty Baseball Projections.xlsx).",
    )
    league.add_argument("--start-year", type=int, default=None, help="First valuation year (default: min Year in file).")
    league.add_argument("--sims", type=positive_int_arg, default=200, help="Monte Carlo sims for SGP denominators.")
    league.add_argument("--horizon", type=positive_int_arg, default=10, help="Dynasty horizon years.")
    league.add_argument("--discount", type=discount_arg, default=0.94, help="Annual discount factor in (0, 1].")
    league.add_argument("--seed", type=int, default=0, help="Global random seed offset for deterministic simulations.")
    league.add_argument(
        "--dynamic-replacement-baselines",
        action="store_true",
        help="Recompute replacement baselines for each valuation year (legacy behavior).",
    )
    league.add_argument("--out-prefix", default="player_values", help="Output prefix for CSV/XLSX.")
    league.add_argument("--recent-projections", type=positive_int_arg, default=3, help="Number of most recent projections to average per player/year.")

    args = p.parse_args()

    bat_detail = None
    pit_detail = None

    if args.mode == "common":
        validate_ip_bounds(args.ip_min, args.ip_max)
        lg = CommonDynastyRotoSettings(
            n_teams=args.teams,
            sims_for_sgp=args.sims,
            horizon_years=args.horizon,
            discount=args.discount,
            bench_slots=args.bench,
            minor_slots=args.minors,
            ir_slots=args.ir,
            ip_min=args.ip_min,
            ip_max=args.ip_max,
            freeze_replacement_baselines=not args.dynamic_replacement_baselines,
        )

        out, bat_detail, pit_detail = calculate_common_dynasty_values(
            args.input,
            lg,
            start_year=args.start_year,
            verbose=True,
            return_details=True,
            seed=args.seed,
            recent_projections=args.recent_projections,
        )

        year_cols = [c for c in out.columns if c.startswith("Value_")]
        df = out[
            [
                "Player",
                "ProjectionsUsed",
                "OldestProjectionDate",
                "Team",
                "Pos",
                "Age",
                "DynastyValue",
                "RawDynastyValue",
                "minor_eligible",
            ]
            + year_cols
            + ["CenteringBaselineMean"]
        ]

    else:
        lg = LeagueSettings(
            sims_for_sgp=args.sims,
            horizon_years=args.horizon,
            discount=args.discount,
            two_way="max",
            freeze_replacement_baselines=not args.dynamic_replacement_baselines,
        )
        validate_ip_bounds(lg.ip_min, lg.ip_max)

        out, bat_detail, pit_detail = calculate_league_dynasty_values(
            args.input,
            lg,
            start_year=args.start_year,
            verbose=True,
            return_details=True,
            seed=args.seed,
            recent_projections=args.recent_projections,
        )

        year_cols = [c for c in out.columns if c.startswith("Value_")]
        df = out[
            [
                "Player",
                "ProjectionsUsed",
                "OldestProjectionDate",
                "MLBTeam",
                "Pos",
                "Age",
                "DynastyValue",
                "RawDynastyValue",
                "minor_eligible",
            ]
            + year_cols
        ]

    csv_path = f"{args.out_prefix}.csv"
    xlsx_path = f"{args.out_prefix}.xlsx"

    # CSV stays as the compact per-player summary (same as before)
    df.to_csv(csv_path, index=False)

    # XLSX now includes extra detail tabs:
    #   - PlayerValues (summary)
    #   - Bat_Aggregated (aggregated Bat sheet + YearValue + DynastyValue)
    #   - Pitch_Aggregated (aggregated Pitch sheet + YearValue + DynastyValue)
    with pd.ExcelWriter(xlsx_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="PlayerValues", index=False)
        if bat_detail is not None:
            bat_detail.to_excel(writer, sheet_name="Bat_Aggregated", index=False)
        if pit_detail is not None:
            pit_detail.to_excel(writer, sheet_name="Pitch_Aggregated", index=False)

        # ----------------------------
        # Formatting pass (openpyxl)
        # ----------------------------
        try:
            if "PlayerValues" in writer.sheets:
                _xlsx_format_player_values(writer.sheets["PlayerValues"], df, table_name="PlayerValuesTbl")

            if bat_detail is not None and "Bat_Aggregated" in writer.sheets:
                _xlsx_format_detail_sheet(
                    writer.sheets["Bat_Aggregated"],
                    bat_detail,
                    table_name="BatAggregatedTbl",
                    is_pitch=False,
                )

            if pit_detail is not None and "Pitch_Aggregated" in writer.sheets:
                _xlsx_format_detail_sheet(
                    writer.sheets["Pitch_Aggregated"],
                    pit_detail,
                    table_name="PitchAggregatedTbl",
                    is_pitch=True,
                )
        except Exception as e:
            # Formatting should never prevent producing the workbook.
            print(f"WARNING: Failed to apply Excel formatting: {e}")

    print("\nTop 25 by DynastyValue:")
    print(df.head(25).to_string(index=False))
    print(f"\nWrote: {csv_path}")
    print(f"Wrote: {xlsx_path}")

if __name__ == "__main__":
    main()
